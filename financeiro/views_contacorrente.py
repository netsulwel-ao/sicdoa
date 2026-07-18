import json
import html as _html_mod
import io
from collections import OrderedDict
from datetime import datetime, timedelta, date
from decimal import Decimal

from django.shortcuts import get_object_or_404, render
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Case, When, Value, IntegerField, DecimalField, F
from django.db.models.functions import TruncMonth, TruncYear, ExtractYear, ExtractMonth
from django.utils import timezone
from django.views.generic import ListView, TemplateView
from django.utils.decorators import method_decorator

from users.auth_decorators import requer_sessao_ativa
from clientes.models import Cliente
from .models import (
    FacturaCliente, ReciboCliente, NotaCredito, NotaDebito, FacturaRecibo, HistoricoFinanceiro
)
from .views import BaseContextMixin
from utils.format_kz import fmt_kz


def _user_filter_direct_from_request(request):
    """Retorna dict de filtro para isolar clientes de cada Despachante."""
    from .views import _user_tem_acesso_total, _is_colaborador_institucional
    if _user_tem_acesso_total(request) or _is_colaborador_institucional(request):
        return {}
    from users.permissoes import get_usuario_permissoes
    from .views import _tem_escopo_filial
    perm_set = get_usuario_permissoes(request)
    banca_id = request.session.get('banca_id')
    if banca_id:
        filtro = {'banca_id': banca_id}
        filial_id = request.session.get('colaborador_filial_id')
        if _tem_escopo_filial(perm_set, filial_id) and filial_id:
            filtro['filial_id'] = filial_id
        return filtro
    usuario_id = request.session.get('banca_usuario_id') or request.session.get('usuario_id')
    if not usuario_id:
        return {}
    return {'usuario_id': usuario_id}


def _agregados_periodo(cliente_ids, data_inicio, data_fim):
    """Retorna dict com faturacao, recebimentos, creditos, debitos para todos os clientes no período.
    
    Faz 7 queries no total (vs 3N+2 sem bulk)."""
    fat = rec = 0
    mov_por_cliente = _movimentacoes_para_clientes(cliente_ids, data_inicio, data_fim)
    for mov in mov_por_cliente.values():
        ind = _calcular_indicadores(mov)
        fat += ind['total_debitos']
        rec += ind['total_creditos']

    nc_sum = sum(
        float(v) for v in
        NotaCredito.objects.filter(cliente_id__in=cliente_ids, data__gte=data_inicio, data__lte=data_fim, estado='Aprovada')
        .values_list('valor_creditado', flat=True)
    )
    nd_sum = sum(
        float(v) for v in
        NotaDebito.objects.filter(cliente_id__in=cliente_ids, data__gte=data_inicio, data__lte=data_fim, estado='Aprovada')
        .values_list('valor', flat=True)
    )
    return {'faturacao': fat, 'recebimentos': rec, 'creditos': nc_sum, 'debitos': nd_sum}


def _movimentacoes_cliente(cliente, data_inicio=None, data_fim=None):
    """Retorna todas as movimentações financeiras de um cliente, ordenadas cronologicamente."""
    movimentos = []

    q_factura = FacturaCliente.objects.filter(cliente=cliente)
    q_recibo = ReciboCliente.objects.filter(cliente=cliente)
    q_nc = NotaCredito.objects.filter(cliente=cliente)
    q_nd = NotaDebito.objects.filter(cliente=cliente)
    q_fr = FacturaRecibo.objects.filter(cliente=cliente)

    for f in q_factura.only('id', 'data_emissao', 'numero_factura', 'descricao', 'valor_total', 'estado'):
        d = f.data_emissao.date() if hasattr(f.data_emissao, 'date') else f.data_emissao
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        movimentos.append({
            'data': d, 'tipo': 'Factura', 'documento': f.numero_factura,
            'descricao': f.descricao[:80],
            'debito': float(f.valor_total) if f.estado != 'Cancelada' else 0,
            'credito': 0, 'estado': f.estado, 'pk': f.pk,
            'tipo_url': 'factura_detalhe',
        })

    for r in q_recibo.only('id', 'data_pagamento', 'numero_recibo', 'forma_pagamento', 'valor_recebido', 'estado'):
        d = r.data_pagamento
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        movimentos.append({
            'data': d, 'tipo': 'Recibo', 'documento': r.numero_recibo,
            'descricao': f'Pagamento via {r.forma_pagamento}',
            'debito': 0,
            'credito': float(r.valor_recebido) if r.estado != 'Cancelado' else 0,
            'estado': r.estado or 'Pago', 'pk': r.pk,
            'tipo_url': 'recibo_detalhe',
        })

    for nc in q_nc.only('id', 'data', 'numero_nota', 'motivo', 'valor_creditado', 'estado'):
        d = nc.data
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        valor = float(nc.valor_creditado) if nc.estado == 'Aprovada' else 0
        movimentos.append({
            'data': d, 'tipo': 'Nota de Crédito', 'documento': nc.numero_nota,
            'descricao': nc.motivo, 'debito': 0, 'credito': valor,
            'estado': nc.estado, 'pk': nc.pk,
            'tipo_url': 'nota_credito_detalhe',
        })

    for nd in q_nd.only('id', 'data', 'numero_nota', 'motivo', 'valor', 'estado'):
        d = nd.data
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        valor = float(nd.valor) if nd.estado == 'Aprovada' else 0
        movimentos.append({
            'data': d, 'tipo': 'Nota de Débito', 'documento': nd.numero_nota,
            'descricao': nd.motivo, 'debito': valor, 'credito': 0,
            'estado': nd.estado, 'pk': nd.pk,
            'tipo_url': 'nota_debito_detalhe',
        })

    for fr in q_fr.only('id', 'data', 'numero_factura_recibo', 'forma_pagamento', 'valor', 'estado'):
        d = fr.data
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        valor = float(fr.valor) if fr.estado != 'Cancelada' else 0
        movimentos.append({
            'data': d, 'tipo': 'Factura-Recibo', 'documento': fr.numero_factura_recibo,
            'descricao': f'Venda c/ pagamento - {fr.forma_pagamento}',
            'debito': valor, 'credito': valor,
            'estado': fr.estado, 'pk': fr.pk,
            'tipo_url': 'factura_recibo_detalhe',
        })

    movimentos.sort(key=lambda x: (x['data'], x['tipo']))
    return movimentos


def _movimentacoes_para_clientes(cliente_ids, data_inicio=None, data_fim=None, excluir_ncnd=False):
    """Retorna {cliente_id: [movimentos]} para todos os clientes em 5 queries (uma por modelo).

    Elimina o N+1 ao processar N clientes — faz 5 queries no total em vez de 5N.
    O filtro de data é aplicado via Python (cada modelo tem campo de data diferente).
    Se excluir_ncnd=True, NC/ND não são incluídos (evita dupla contagem quando
    são exibidos em colunas separadas).
    """
    filtro_base = {'cliente_id__in': cliente_ids}
    movimentos_por_cliente = {cid: [] for cid in cliente_ids}

    facturas = FacturaCliente.objects.filter(**filtro_base).only(
        'id', 'cliente_id', 'data_emissao', 'numero_factura', 'descricao',
        'valor_total', 'estado'
    )
    for f in facturas:
        d = f.data_emissao.date() if hasattr(f.data_emissao, 'date') else f.data_emissao
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        movimentos_por_cliente[f.cliente_id].append({
            'data': d, 'tipo': 'Factura', 'documento': f.numero_factura,
            'descricao': f.descricao[:80],
            'debito': float(f.valor_total) if f.estado != 'Cancelada' else 0,
            'credito': 0, 'estado': f.estado, 'pk': f.pk,
            'tipo_url': 'factura_detalhe',
        })

    recibos = ReciboCliente.objects.filter(**filtro_base).only(
        'id', 'cliente_id', 'data_pagamento', 'numero_recibo',
        'forma_pagamento', 'valor_recebido', 'estado'
    )
    for r in recibos:
        d = r.data_pagamento
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        movimentos_por_cliente[r.cliente_id].append({
            'data': d, 'tipo': 'Recibo', 'documento': r.numero_recibo,
            'descricao': f'Pagamento via {r.forma_pagamento}',
            'debito': 0,
            'credito': float(r.valor_recebido) if r.estado != 'Cancelado' else 0,
            'estado': r.estado or 'Pago', 'pk': r.pk,
            'tipo_url': 'recibo_detalhe',
        })

    if not excluir_ncnd:
        notas_credito = NotaCredito.objects.filter(**filtro_base).only(
            'id', 'cliente_id', 'data', 'numero_nota', 'motivo', 'valor_creditado', 'estado'
        )
        for nc in notas_credito:
            d = nc.data
            if data_inicio and d < data_inicio:
                continue
            if data_fim and d > data_fim:
                continue
            valor = float(nc.valor_creditado) if nc.estado == 'Aprovada' else 0
            movimentos_por_cliente[nc.cliente_id].append({
                'data': d, 'tipo': 'Nota de Crédito', 'documento': nc.numero_nota,
                'descricao': nc.motivo, 'debito': 0, 'credito': valor,
                'estado': nc.estado, 'pk': nc.pk,
                'tipo_url': 'nota_credito_detalhe',
            })

    if not excluir_ncnd:
        notas_debito = NotaDebito.objects.filter(**filtro_base).only(
            'id', 'cliente_id', 'data', 'numero_nota', 'motivo', 'valor', 'estado'
        )
        for nd in notas_debito:
            d = nd.data
            if data_inicio and d < data_inicio:
                continue
            if data_fim and d > data_fim:
                continue
            valor = float(nd.valor) if nd.estado == 'Aprovada' else 0
            movimentos_por_cliente[nd.cliente_id].append({
                'data': d, 'tipo': 'Nota de Débito', 'documento': nd.numero_nota,
                'descricao': nd.motivo, 'debito': valor, 'credito': 0,
                'estado': nd.estado, 'pk': nd.pk,
                'tipo_url': 'nota_debito_detalhe',
            })

    facturas_recibo = FacturaRecibo.objects.filter(**filtro_base).only(
        'id', 'cliente_id', 'data', 'numero_factura_recibo',
        'forma_pagamento', 'valor', 'estado'
    )
    for fr in facturas_recibo:
        d = fr.data
        if data_inicio and d < data_inicio:
            continue
        if data_fim and d > data_fim:
            continue
        valor = float(fr.valor) if fr.estado != 'Cancelada' else 0
        movimentos_por_cliente[fr.cliente_id].append({
            'data': d, 'tipo': 'Factura-Recibo', 'documento': fr.numero_factura_recibo,
            'descricao': f'Venda c/ pagamento - {fr.forma_pagamento}',
            'debito': valor, 'credito': valor,
            'estado': fr.estado, 'pk': fr.pk,
            'tipo_url': 'factura_recibo_detalhe',
        })

    for cid in movimentos_por_cliente:
        movimentos_por_cliente[cid].sort(key=lambda x: (x['data'], x['tipo']))

    return movimentos_por_cliente


def _calcular_indicadores(movimentos, saldo_inicial=0):
    """Calcula totais de débitos, créditos e saldo corrente a partir dos movimentos."""
    total_debitos = sum(m['debito'] for m in movimentos)
    total_creditos = sum(m['credito'] for m in movimentos)
    saldo_atual = saldo_inicial - total_debitos + total_creditos
    return {
        'total_debitos': total_debitos,
        'total_creditos': total_creditos,
        'saldo_inicial': saldo_inicial,
        'saldo_atual': saldo_atual,
    }


# ─── 3.0 Conta Corrente Home ─────────────────────────────────────────────────

@method_decorator(requer_sessao_ativa, name='dispatch')
class ContaCorrenteHomeView(BaseContextMixin, TemplateView):
    template_name = 'financeiro/conta_corrente_home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'Financeiro'
        context['active_sub'] = 'conta_corrente'
        return context


# ─── 3.1 Conta Corrente por Cliente ──────────────────────────────────────────

@method_decorator(requer_sessao_ativa, name='dispatch')
class ContaCorrenteClienteListView(BaseContextMixin, TemplateView):
    template_name = 'financeiro/conta_corrente_cliente_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        busca = self.request.GET.get('busca', '')
        clientes = Cliente.objects.all()
        filtro = self._get_user_filter_direct()
        if filtro:
            clientes = clientes.filter(**filtro)
        if busca:
            clientes = clientes.filter(
                Q(nome__icontains=busca) | Q(nif__icontains=busca) | Q(telefone__icontains=busca)
            )
        paginator = Paginator(clientes, 20)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['clientes'] = page_obj
        context['page_obj'] = page_obj
        context['busca'] = busca
        context['active_menu'] = 'Financeiro'
        context['active_sub'] = 'conta_corrente_cliente'
        return context


@method_decorator(requer_sessao_ativa, name='dispatch')
class ContaCorrenteClienteView(BaseContextMixin, TemplateView):
    template_name = 'financeiro/conta_corrente_cliente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente_id = self.kwargs.get('pk')
        filtro = self._get_user_filter_direct()
        if filtro:
            cliente = get_object_or_404(Cliente, pk=cliente_id, **filtro)
        else:
            cliente = get_object_or_404(Cliente, pk=cliente_id)

        data_inicio = self.request.GET.get('data_inicio')
        data_fim = self.request.GET.get('data_fim')
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()

        movimentos = _movimentacoes_cliente(cliente, data_inicio, data_fim)
        total_debitos = sum(m['debito'] for m in movimentos)
        total_creditos = sum(m['credito'] for m in movimentos)
        saldo_inicial = float(cliente.saldo_conta_corrente) + total_debitos - total_creditos
        indicadores = _calcular_indicadores(movimentos, saldo_inicial)

        context['cliente'] = cliente
        context['movimentos'] = movimentos
        context['indicadores'] = indicadores
        context['data_inicio'] = self.request.GET.get('data_inicio', '')
        context['data_fim'] = self.request.GET.get('data_fim', '')
        context['active_menu'] = 'Financeiro'
        context['active_sub'] = 'conta_corrente'
        return context


# ─── 3.2 Conta Corrente Geral dos Clientes ───────────────────────────────────

@method_decorator(requer_sessao_ativa, name='dispatch')
class ContaCorrenteGeralView(BaseContextMixin, TemplateView):
    template_name = 'financeiro/conta_corrente_geral.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        busca = self.request.GET.get('busca', '')
        data_inicio = self.request.GET.get('data_inicio')
        data_fim = self.request.GET.get('data_fim')

        clientes = Cliente.objects.all()
        filtro = self._get_user_filter_direct()
        if filtro:
            clientes = clientes.filter(**filtro)
        if busca:
            clientes = clientes.filter(
                Q(nome__icontains=busca) | Q(nif__icontains=busca)
            )

        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()

        dados_clientes = []
        total_a_receber = 0
        total_recebido = 0
        total_divida = 0
        clientes_ativos = 0
        clientes_inativos = 0

        cliente_ids = [c.pk for c in clientes]
        mov_por_cliente = _movimentacoes_para_clientes(cliente_ids, data_inicio, data_fim)

        for cli in clientes:
            mov = mov_por_cliente.get(cli.pk, [])
            tot_debitos = sum(m['debito'] for m in mov)
            tot_creditos = sum(m['credito'] for m in mov)
            saldo_inicial = float(cli.saldo_conta_corrente) + tot_debitos - tot_creditos
            ind = _calcular_indicadores(mov, saldo_inicial)
            divida = abs(ind['saldo_atual']) if ind['saldo_atual'] < 0 else 0

            total_a_receber += ind['total_debitos']
            total_recebido += ind['total_creditos']
            total_divida += divida

            if cli.ativo:
                clientes_ativos += 1
            else:
                clientes_inativos += 1

            dados_clientes.append({
                'cliente': cli,
                'total_debitos': ind['total_debitos'],
                'total_creditos': ind['total_creditos'],
                'saldo': ind['saldo_atual'],
                'divida': divida,
            })

        # Ranking devedores (maiores saldos negativos)
        ranking_devedores = sorted(
            [d for d in dados_clientes if d['saldo'] < 0],
            key=lambda x: x['saldo']
        )[:10]

        # Ranking melhores clientes (maiores saldos positivos / mais créditos)
        ranking_melhores = sorted(
            [d for d in dados_clientes if d['saldo'] >= 0],
            key=lambda x: x['total_creditos'],
            reverse=True
        )[:10]

        context['dados_clientes'] = dados_clientes
        context['total_a_receber'] = total_a_receber
        context['total_recebido'] = total_recebido
        context['total_divida'] = total_divida
        context['clientes_ativos'] = clientes_ativos
        context['clientes_inativos'] = clientes_inativos
        context['ranking_devedores'] = ranking_devedores
        context['ranking_melhores'] = ranking_melhores
        context['busca'] = busca
        context['data_inicio'] = self.request.GET.get('data_inicio', '')
        context['data_fim'] = self.request.GET.get('data_fim', '')
        context['active_menu'] = 'Financeiro'
        context['active_sub'] = 'conta_corrente_geral'
        return context


# ─── 3.3 Conta Corrente Mensal ───────────────────────────────────────────────

@method_decorator(requer_sessao_ativa, name='dispatch')
class ContaCorrenteMensalView(BaseContextMixin, TemplateView):
    template_name = 'financeiro/conta_corrente_mensal.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ano = self.request.GET.get('ano', str(timezone.now().year))
        mes = self.request.GET.get('mes', '')
        cliente_id = self.request.GET.get('cliente')

        try:
            ano = int(ano)
        except ValueError:
            ano = timezone.now().year

        meses_resumo = OrderedDict()
        for m in range(1, 13):
            meses_resumo[m] = {
                'faturacao': 0,
                'recebimentos': 0,
                'creditos_emitidos': 0,
                'debitos_emitidos': 0,
                'saldo': 0,
            }

        clientes_qs = Cliente.objects.all()
        filtro = self._get_user_filter_direct()
        if filtro:
            clientes_qs = clientes_qs.filter(**filtro)
        if cliente_id:
            clientes_qs = clientes_qs.filter(pk=cliente_id)

        cliente_ids = list(clientes_qs.values_list('pk', flat=True))

        data_inicio_ano = date(ano, 1, 1)
        data_fim_ano = date(ano + 1, 1, 1)

        # 5 queries para o ano todo em vez de 12×5=60
        mov_por_cliente = _movimentacoes_para_clientes(cliente_ids, data_inicio_ano, data_fim_ano - timedelta(days=1), excluir_ncnd=True)

        # NC e ND agregadas por mês — 2 queries em vez de 12×2=24
        nc_por_mes = NotaCredito.objects.filter(
            cliente_id__in=cliente_ids, data__gte=data_inicio_ano, data__lt=data_fim_ano, estado='Aprovada'
        ).annotate(mes=TruncMonth('data')).values('mes').annotate(total=Sum('valor_creditado'))
        nc_dict = {r['mes'].month: float(r['total']) for r in nc_por_mes}

        nd_por_mes = NotaDebito.objects.filter(
            cliente_id__in=cliente_ids, data__gte=data_inicio_ano, data__lt=data_fim_ano, estado='Aprovada'
        ).annotate(mes=TruncMonth('data')).values('mes').annotate(total=Sum('valor'))
        nd_dict = {r['mes'].month: float(r['total']) for r in nd_por_mes}

        for m in range(1, 13):
            data_ini = date(ano, m, 1)
            if m == 12:
                data_fim = data_fim_ano
            else:
                data_fim = date(ano, m + 1, 1)
            data_fim_mes = data_fim - timedelta(days=1)

            for cid, mov in mov_por_cliente.items():
                mov_mes = [mv for mv in mov if data_ini <= mv['data'] <= data_fim_mes]
                if mov_mes:
                    ind = _calcular_indicadores(mov_mes)
                    meses_resumo[m]['faturacao'] += ind['total_debitos']
                    meses_resumo[m]['recebimentos'] += ind['total_creditos']
                    meses_resumo[m]['saldo'] += ind['saldo_atual']

            meses_resumo[m]['creditos_emitidos'] = nc_dict.get(m, 0)
            meses_resumo[m]['debitos_emitidos'] = nd_dict.get(m, 0)

        # Filtrar por mês se especificado
        if mes:
            try:
                mes = int(mes)
                if 1 <= mes <= 12:
                    meses_resumo = {mes: meses_resumo[mes]}
            except ValueError:
                pass

        context['meses_resumo'] = meses_resumo
        context['ano'] = ano
        context['mes'] = mes
        context['anos_disponiveis'] = list(range(2020, timezone.now().year + 2))
        context['active_menu'] = 'Financeiro'
        context['active_sub'] = 'conta_corrente_mensal'
        context['cliente_id'] = cliente_id
        return context


@requer_sessao_ativa
def conta_corrente_mensal_excel(request):
    ano = request.GET.get('ano', str(timezone.now().year))
    try:
        ano = int(ano)
    except ValueError:
        ano = timezone.now().year

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Conta Corrente Mensal {ano}'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    sub_header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    headers = ['Mês', 'Faturação', 'Recebimentos', 'Créditos Emitidos', 'Débitos Emitidos', 'Saldo do Mês']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    clientes_base = Cliente.objects.all()
    filtro = _user_filter_direct_from_request(request)
    if filtro:
        clientes_base = clientes_base.filter(**filtro)
    cliente_ids = list(clientes_base.values_list('pk', flat=True))

    row = 2
    for m in range(1, 13):
        data_ini = date(ano, m, 1)
        data_fim = date(ano + 1, 1, 1) if m == 12 else date(ano, m + 1, 1)
        data_fim_mes = data_fim - timedelta(days=1)

        agg = _agregados_periodo(cliente_ids, data_ini, data_fim_mes)
        faturacao = agg['faturacao']
        recebimentos = agg['recebimentos']
        creditos = agg['creditos']
        debitos = agg['debitos']

        saldo = recebimentos - faturacao
        meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                     'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

        vals = [meses_pt[m - 1], faturacao, recebimentos, creditos, debitos, saldo]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = sub_header_fill
            if col > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(2, 7):
        col_sum = sum(ws.cell(row=r, column=col).value or 0 for r in range(2, total_row))
        cell = ws.cell(row=total_row, column=col, value=col_sum)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.border = thin_border

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="conta_corrente_mensal_{ano}.xlsx"'
    wb.save(response)
    return response


def _safe_cc(text):
    """Escapa caracteres HTML/XML especiais para prevenir XSS em PDFs."""
    if not text:
        return ''
    return _html_mod.escape(str(text))


def _build_cc_header(banca, titulo, W):
    """Cabeçalho padrão: logo, despachante, banca, HASH, título + HR azul."""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from reportlab.platypus.flowables import HRFlowable
    from users.models import Usuario

    COR_PRETO    = colors.HexColor('#0f172a')
    COR_CINZA    = colors.HexColor('#64748b')
    COR_VERDE    = colors.HexColor('#059669')
    COR_PRIMARIO = colors.HexColor('#137fec')
    COR_BORDA    = colors.HexColor('#cbd5e1')

    def st(name, **kw):
        d = dict(fontName='Helvetica', fontSize=9, textColor=COR_PRETO, leading=11)
        d.update(kw)
        return ParagraphStyle(name, **d)

    s_small = st('small', fontSize=7, textColor=COR_CINZA, leading=9)
    s_titulo = st('tit_cc', fontSize=14, fontName='Helvetica-Bold', textColor=COR_PRETO)

    flowables = []

    logo_path = None
    if banca and hasattr(banca, 'logo') and banca.logo:
        try:
            logo_path = banca.logo.path
        except Exception:
            logo_path = None
    col_logo = Paragraph('', s_small)
    if logo_path:
        try:
            col_logo = RLImage(logo_path, width=2.2 * cm, height=1.6 * cm)
        except Exception:
            col_logo = Paragraph('', s_small)

    top_line = Table([[col_logo, '']], colWidths=[W - 1.9 * cm, 1.9 * cm])
    top_line.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    flowables.append(top_line)

    nome_txt   = _safe_cc(banca.nome) if banca else 'Despachante Oficial'
    nif_txt    = _safe_cc(banca.nif) if banca else 'N/D'
    cdoa       = _safe_cc(banca.licenca_cdoa) if banca else '—'
    endereco   = _safe_cc(banca.endereco) if banca else '—'
    telefone   = _safe_cc(banca.telefone) if banca else '—'
    email_b    = _safe_cc(banca.email) if banca else '—'

    resp_nome = 'DESPACHANTE OFICIAL'
    resp_nif = resp_ced = resp_tel = resp_email = '—'
    if banca:
        try:
            ub = Usuario.objects.get(id=banca.usuario_id)
            resp_nome   = _safe_cc(ub.nome or 'DESPACHANTE OFICIAL').upper()
            resp_nif    = _safe_cc(ub.nif) or '—'
            resp_ced    = _safe_cc(ub.cedula) or '—'
            resp_tel    = _safe_cc(ub.telefone) or '—'
            resp_email  = _safe_cc(ub.email) or '—'
        except Exception:
            pass

    desp_info = (
        f'<font size="7" color="{COR_VERDE.hexval()}"><b>DESPACHANTE: {resp_nome}</b></font><br/>'
        f'<font size="6.5" color="#64748b">NIF: {resp_nif}</font><br/>'
        f'<font size="6.5" color="#64748b">Cédula CDOA: {resp_ced}</font><br/>'
        f'<font size="6.5" color="#64748b">Tel: {resp_tel} &nbsp;|&nbsp; Email: {resp_email}</font>'
    )
    banca_info = (
        f'<font size="7" color="#475569"><b>{nome_txt}</b></font><br/>'
        f'<font size="6.5" color="#64748b">NIF: {nif_txt} &nbsp;|&nbsp; Licença CDOA: {cdoa}</font><br/>'
        f'<font size="6.5" color="#64748b">{endereco}</font><br/>'
        f'<font size="6.5" color="#64748b">Tel: {telefone} &nbsp;|&nbsp; Email: {email_b}</font>'
    )
    header_body = Table([[
        Paragraph(desp_info, st('desp_cc', fontSize=6.5, leading=9)),
        Paragraph(banca_info, st('banca_cc', fontSize=6.5, leading=9, alignment=TA_RIGHT)),
    ]], colWidths=[W * 0.55, W * 0.45])
    header_body.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    flowables.append(header_body)
    flowables.append(Spacer(1, 0.1 * cm))

    flowables.append(Paragraph(
        f'<font size="6" color="#94a3b8"><b>{nome_txt} - HASH</b> &nbsp;|&nbsp; '
        f'Processado por programa válido nº35/AGT/2019</font>',
        st('hash_cc', fontSize=6)
    ))
    flowables.append(Spacer(1, 0.1 * cm))
    flowables.append(HRFlowable(width=W, thickness=0.5, color=COR_BORDA))
    flowables.append(Spacer(1, 0.1 * cm))

    flowables.append(Paragraph(titulo, s_titulo))
    flowables.append(HRFlowable(width=W, thickness=2, color=COR_PRIMARIO, spaceAfter=12))

    return flowables


def _build_cc_footer(banca, W):
    """Rodapé: dados bancários compactos (espelho do factura_pdf)."""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.platypus.flowables import HRFlowable

    COR_PRETO = colors.HexColor('#0f172a')
    COR_BORDA = colors.HexColor('#cbd5e1')

    def st(name, **kw):
        d = dict(fontName='Helvetica', fontSize=9, textColor=COR_PRETO, leading=11)
        d.update(kw)
        return ParagraphStyle(name, **d)

    flowables = []

    bancos_pdf = []
    if banca:
        try:
            bancos_pdf = json.loads(banca.dados_bancarios_json or '[]')
        except (json.JSONDecodeError, ValueError):
            bancos_pdf = []
        if not isinstance(bancos_pdf, list):
            bancos_pdf = []

    has_bank = False
    if bancos_pdf:
        has_bank = any(b.get('banco') for b in bancos_pdf if isinstance(b, dict))
    elif banca:
        has_bank = bool(banca.banco or banca.numero_conta or banca.iban)

    if has_bank or (banca and banca.instrucoes_pagamento):
        flowables.append(Spacer(1, 0.1 * cm))
        flowables.append(HRFlowable(width=W, thickness=0.3, color=COR_BORDA))
        flowables.append(Spacer(1, 0.08 * cm))
        flowables.append(Paragraph(
            '<font size="5.5" color="#1e293b"><b>Dados Bancários</b></font>',
            st('bank_title_cc', fontSize=5.5)
        ))
        flowables.append(Spacer(1, 0.03 * cm))

        if bancos_pdf:
            for i, b in enumerate(bancos_pdf):
                if not isinstance(b, dict) or not b.get('banco'):
                    continue
                flowables.append(Paragraph(
                    f'<font size="5" color="#475569">'
                    f'<b>{i + 1}.</b> &nbsp;'
                    f'<b>Banco:</b> {b["banco"]} &nbsp;|&nbsp;'
                    f'<b>IBAN:</b> <font name="Courier">{b.get("iban", "—")}</font>'
                    f'</font>',
                    st(f'bank_line_cc_{i}', fontSize=5, leading=6.5, leftIndent=8)
                ))
        elif banca:
            parts = []
            if banca.banco:
                parts.append(f'<b>Banco:</b> {banca.banco}')
            if banca.iban:
                parts.append(f'<b>IBAN:</b> <font name="Courier">{banca.iban}</font>')
            if banca.numero_conta:
                parts.append(f'<b>Conta:</b> {banca.numero_conta}')
            if parts:
                flowables.append(Paragraph(
                    f'<font size="5" color="#475569">{" &nbsp;|&nbsp; ".join(parts)}</font>',
                    st('bank_foot_cc', fontSize=5, leading=6.5, leftIndent=8)
                ))

        if banca and banca.instrucoes_pagamento:
            txt = banca.instrucoes_pagamento.replace('\n', ' ').replace('\r', '')
            flowables.append(Spacer(1, 0.02 * cm))
            flowables.append(Paragraph(
                f'<font size="4.5" color="#64748b"><i>{txt}</i></font>',
                st('bank_inst_cc', fontSize=4.5, leading=6, leftIndent=8)
            ))

    return flowables


@requer_sessao_ativa
def conta_corrente_mensal_pdf(request):
    ano = request.GET.get('ano', str(timezone.now().year))
    try:
        ano = int(ano)
    except ValueError:
        ano = timezone.now().year

    buffer = io.BytesIO()
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.platypus.flowables import HRFlowable

    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title=f'Conta Corrente Mensal {ano}',
    )
    W = landscape(A4)[0] - 3*cm

    cor_primaria = colors.HexColor('#137fec')
    cor_cabecalho = colors.HexColor('#0f172a')
    cor_borda = colors.HexColor('#e2e8f0')
    cor_linha_par = colors.HexColor('#f8fafc')

    s_titulo = ParagraphStyle('tit', fontSize=16, fontName='Helvetica-Bold', textColor=cor_cabecalho)
    s_normal = ParagraphStyle('n', fontSize=8, fontName='Helvetica', textColor=cor_cabecalho, leading=10)
    s_bold = ParagraphStyle('b', fontSize=8, fontName='Helvetica-Bold', textColor=cor_cabecalho, leading=10)
    s_header = ParagraphStyle('h', fontSize=8, fontName='Helvetica-Bold', textColor=colors.white, leading=10)

    story = []
    banca = None
    banca_id = request.session.get('banca_id')
    if banca_id:
        try:
            from rh.models import Banca
            banca = Banca.objects.get(id=banca_id)
        except Exception:
            banca = None
    titulo = f'Conta Corrente Mensal - {ano}'
    story.extend(_build_cc_header(banca, titulo, W))

    meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    headers = ['Mês', 'Faturação', 'Recebimentos', 'Créditos', 'Débitos', 'Saldo']
    t_data = [[Paragraph(h, s_header) for h in headers]]

    clientes_base = Cliente.objects.all()
    filtro = _user_filter_direct_from_request(request)
    if filtro:
        clientes_base = clientes_base.filter(**filtro)
    cliente_ids = list(clientes_base.values_list('pk', flat=True))

    for m in range(1, 13):
        data_ini = date(ano, m, 1)
        data_fim = date(ano + 1, 1, 1) if m == 12 else date(ano, m + 1, 1)
        data_fim_mes = data_fim - timedelta(days=1)

        agg = _agregados_periodo(cliente_ids, data_ini, data_fim_mes)
        fat = agg['faturacao']
        rec = agg['recebimentos']
        cr = agg['creditos']
        db = agg['debitos']

        saldo = rec - fat
        row_data = [
            Paragraph(meses_pt[m - 1], s_normal),
            Paragraph(f'{fmt_kz(fat)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(rec)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(cr)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(db)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(saldo)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica-Bold', alignment=2)),
        ]
        t_data.append(row_data)

    col_widths = [3*cm, (W-3*cm)*0.2, (W-3*cm)*0.2, (W-3*cm)*0.2, (W-3*cm)*0.2, (W-3*cm)*0.2]
    t = Table(t_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), cor_cabecalho),
        ('GRID', (0, 0), (-1, -1), 0.4, cor_borda),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, cor_linha_par]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(t)
    story.extend(_build_cc_footer(banca, W))
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="conta_corrente_mensal_{ano}.pdf"'
    return response


# ─── 3.4 Conta Corrente Periódica ────────────────────────────────────────────

PERIODOS = {
    'diario': 'Diário',
    'semanal': 'Semanal',
    'quinzenal': 'Quinzenal',
    'mensal': 'Mensal',
    'trimestral': 'Trimestral',
    'semestral': 'Semestral',
    'anual': 'Anual',
    'personalizado': 'Personalizado',
}


def _intervalos_periodo(tipo_periodo, ano, mes=None):
    """Gera lista de intervalos (data_ini, data_fim, label) para um dado tipo de período."""
    intervalos = []
    if tipo_periodo == 'diario':
        dias_no_mes = 30 if not mes else (
            (date(ano, mes + 1, 1) - date(ano, mes, 1)).days if mes < 12 else
            (date(ano + 1, 1, 1) - date(ano, 12, 1)).days
        )
        m = mes or 1
        for d in range(1, dias_no_mes + 1):
            try:
                di = date(ano, m, d)
                intervalos.append((di, di, f'{di.strftime("%d/%m/%Y")}'))
            except ValueError:
                pass
    elif tipo_periodo == 'semanal':
        m = mes or 1
        di = date(ano, m, 1)
        while di.month == m:
            df = di + timedelta(days=6)
            if df.month != m:
                df = date(ano, m + 1, 1) - timedelta(days=1) if m < 12 else date(ano, 12, 31)
            intervalos.append((di, df, f'Sem {di.strftime("%d/%m")} - {df.strftime("%d/%m")}'))
            di = df + timedelta(days=1)
    elif tipo_periodo == 'quinzenal':
        m = mes or 1
        di = date(ano, m, 1)
        meio = date(ano, m, 15)
        fim = date(ano, m + 1, 1) - timedelta(days=1) if m < 12 else date(ano, 12, 31)
        intervalos.append((di, meio, f'1ª Quinzena {meses_pt[m-1]}'))
        intervalos.append((meio + timedelta(days=1), fim, f'2ª Quinzena {meses_pt[m-1]}'))
    elif tipo_periodo == 'mensal':
        m = mes or 1
        di = date(ano, m, 1)
        df = date(ano, m + 1, 1) - timedelta(days=1) if m < 12 else date(ano, 12, 31)
        intervalos.append((di, df, meses_pt[m - 1]))
    elif tipo_periodo == 'trimestral':
        for t in range(1, 13, 3):
            di = date(ano, t, 1)
            df = date(ano, t + 2, 1) - timedelta(days=1) if t + 2 < 12 else date(ano, 12, 31)
            intervalos.append((di, df, f'{t}º Trimestre'))
    elif tipo_periodo == 'semestral':
        intervalos.append((date(ano, 1, 1), date(ano, 6, 30), '1º Semestre'))
        intervalos.append((date(ano, 7, 1), date(ano, 12, 31), '2º Semestre'))
    elif tipo_periodo == 'anual':
        intervalos.append((date(ano, 1, 1), date(ano, 12, 31), str(ano)))
    elif tipo_periodo == 'personalizado':
        pass  # Será tratado separadamente

    return intervalos


meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']


@method_decorator(requer_sessao_ativa, name='dispatch')
class ContaCorrentePeriodicaView(BaseContextMixin, TemplateView):
    template_name = 'financeiro/conta_corrente_periodica.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tipo_periodo = self.request.GET.get('periodo', 'mensal')
        ano = self.request.GET.get('ano', str(timezone.now().year))
        mes = self.request.GET.get('mes', '')
        data_personalizada_ini = self.request.GET.get('data_ini', '')
        data_personalizada_fim = self.request.GET.get('data_fim', '')
        cliente_id = self.request.GET.get('cliente')

        try:
            ano = int(ano)
        except ValueError:
            ano = timezone.now().year

        if mes:
            try:
                mes = int(mes)
            except ValueError:
                mes = None

        clientes_qs = Cliente.objects.all()
        filtro = self._get_user_filter_direct()
        if filtro:
            clientes_qs = clientes_qs.filter(**filtro)
        if cliente_id:
            clientes_qs = clientes_qs.filter(pk=cliente_id)

        if tipo_periodo == 'personalizado' and data_personalizada_ini and data_personalizada_fim:
            try:
                di = datetime.strptime(data_personalizada_ini, '%Y-%m-%d').date()
                df = datetime.strptime(data_personalizada_fim, '%Y-%m-%d').date()
                intervalos = [(di, df, f'{di.strftime("%d/%m/%Y")} a {df.strftime("%d/%m/%Y")}')]
            except ValueError:
                intervalos = []
        else:
            intervalos = _intervalos_periodo(tipo_periodo, ano, mes)

        cliente_ids = list(clientes_qs.values_list('pk', flat=True))
        periodos = []
        for data_ini, data_fim, label in intervalos:
            agg = _agregados_periodo(cliente_ids, data_ini, data_fim)
            fat = agg['faturacao']
            rec = agg['recebimentos']
            cr = agg['creditos']
            db = agg['debitos']

            saldo = rec - fat
            periodos.append({
                'label': label,
                'data_ini': data_ini,
                'data_fim': data_fim,
                'faturacao': fat,
                'recebimentos': rec,
                'creditos': cr,
                'debitos': db,
                'saldo': saldo,
            })

        context['periodos'] = periodos
        context['tipo_periodo'] = tipo_periodo
        context['ano'] = ano
        context['mes'] = str(mes) if mes else ''
        context['data_ini'] = data_personalizada_ini
        context['data_fim'] = data_personalizada_fim
        context['anos_disponiveis'] = list(range(2020, timezone.now().year + 2))
        context['periodos_disponiveis'] = PERIODOS
        context['active_menu'] = 'Financeiro'
        context['active_sub'] = 'conta_corrente_periodica'
        context['cliente_id'] = cliente_id
        context['total_faturacao'] = sum(p['faturacao'] for p in periodos)
        context['total_recebimentos'] = sum(p['recebimentos'] for p in periodos)
        context['total_saldo'] = sum(p['saldo'] for p in periodos)
        return context


@requer_sessao_ativa
def conta_corrente_periodica_json(request):
    """API que retorna dados em JSON para gráficos de evolução financeira."""
    tipo_periodo = request.GET.get('periodo', 'mensal')
    ano = request.GET.get('ano', str(timezone.now().year))
    try:
        ano = int(ano)
    except ValueError:
        ano = timezone.now().year

    if tipo_periodo == 'personalizado':
        data_ini_str = request.GET.get('data_ini', '')
        data_fim_str = request.GET.get('data_fim', '')
        intervalos = []
        if data_ini_str and data_fim_str:
            try:
                di = datetime.strptime(data_ini_str, '%Y-%m-%d').date()
                df = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                intervalos = [(di, df, f'{di.strftime("%d/%m")} a {df.strftime("%d/%m")}')]
            except ValueError:
                pass
    else:
        intervalos = _intervalos_periodo(tipo_periodo, ano)

    data = []
    clientes_base = Cliente.objects.all()
    filtro = _user_filter_direct_from_request(request)
    if filtro:
        clientes_base = clientes_base.filter(**filtro)
    cliente_ids = list(clientes_base.values_list('pk', flat=True))
    for data_ini, data_fim, label in intervalos:
        agg = _agregados_periodo(cliente_ids, data_ini, data_fim)
        fat = agg['faturacao']
        rec = agg['recebimentos']
        cr = agg['creditos']
        db = agg['debitos']
        saldo = rec - fat
        data.append({
            'label': label,
            'faturacao': round(fat, 2),
            'recebimentos': round(rec, 2),
            'creditos': round(cr, 2),
            'debitos': round(db, 2),
            'saldo': round(saldo, 2),
        })

    return JsonResponse({'data': data})


@requer_sessao_ativa
def conta_corrente_periodica_excel(request):
    tipo_periodo = request.GET.get('periodo', 'mensal')
    ano = request.GET.get('ano', str(timezone.now().year))
    try:
        ano = int(ano)
    except ValueError:
        ano = timezone.now().year

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Conta Corrente {tipo_periodo}'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    sub_header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    headers = ['Período', 'Faturação', 'Recebimentos', 'Créditos', 'Débitos', 'Saldo']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    if tipo_periodo == 'personalizado':
        data_ini_str = request.GET.get('data_ini', '')
        data_fim_str = request.GET.get('data_fim', '')
        intervalos = []
        if data_ini_str and data_fim_str:
            try:
                di = datetime.strptime(data_ini_str, '%Y-%m-%d').date()
                df = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                intervalos = [(di, df, f'{di.strftime("%d/%m/%Y")} a {df.strftime("%d/%m/%Y")}')]
            except ValueError:
                pass
    else:
        intervalos = _intervalos_periodo(tipo_periodo, ano)

    clientes_base = Cliente.objects.all()
    filtro = _user_filter_direct_from_request(request)
    if filtro:
        clientes_base = clientes_base.filter(**filtro)
    cliente_ids = list(clientes_base.values_list('pk', flat=True))
    row = 2
    for data_ini, data_fim, label in intervalos:
        agg = _agregados_periodo(cliente_ids, data_ini, data_fim)
        fat = agg['faturacao']
        rec = agg['recebimentos']
        cr = agg['creditos']
        db = agg['debitos']
        saldo = rec - fat

        vals = [label, fat, rec, cr, db, saldo]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = sub_header_fill
            if col > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 25

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="conta_corrente_{tipo_periodo}_{ano}.xlsx"'
    wb.save(response)
    return response


@requer_sessao_ativa
def conta_corrente_periodica_pdf(request):
    tipo_periodo = request.GET.get('periodo', 'mensal')
    ano = request.GET.get('ano', str(timezone.now().year))
    try:
        ano = int(ano)
    except ValueError:
        ano = timezone.now().year

    data_personalizada_ini = request.GET.get('data_ini', '')
    data_personalizada_fim = request.GET.get('data_fim', '')

    if tipo_periodo == 'personalizado' and data_personalizada_ini and data_personalizada_fim:
        try:
            di = datetime.strptime(data_personalizada_ini, '%Y-%m-%d').date()
            df = datetime.strptime(data_personalizada_fim, '%Y-%m-%d').date()
            intervalos = [(di, df, f'{di.strftime("%d/%m/%Y")} a {df.strftime("%d/%m/%Y")}')]
        except ValueError:
            intervalos = []
    else:
        intervalos = _intervalos_periodo(tipo_periodo, ano)

    buffer = io.BytesIO()
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.platypus.flowables import HRFlowable

    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title=f'Conta Corrente {PERIODOS.get(tipo_periodo, tipo_periodo)}',
    )
    W = landscape(A4)[0] - 3*cm

    cor_primaria = colors.HexColor('#137fec')
    cor_cabecalho = colors.HexColor('#0f172a')
    cor_borda = colors.HexColor('#e2e8f0')
    cor_linha_par = colors.HexColor('#f8fafc')

    s_titulo = ParagraphStyle('tit', fontSize=16, fontName='Helvetica-Bold', textColor=cor_cabecalho)
    s_normal = ParagraphStyle('n', fontSize=8, fontName='Helvetica', textColor=cor_cabecalho, leading=10)
    s_bold = ParagraphStyle('b', fontSize=8, fontName='Helvetica-Bold', textColor=cor_cabecalho, leading=10)
    s_header = ParagraphStyle('h', fontSize=8, fontName='Helvetica-Bold', textColor=colors.white, leading=10)

    story = []
    banca = None
    banca_id = request.session.get('banca_id')
    if banca_id:
        try:
            from rh.models import Banca
            banca = Banca.objects.get(id=banca_id)
        except Exception:
            banca = None
    titulo = f'Conta Corrente {PERIODOS.get(tipo_periodo, tipo_periodo)} - {ano}'
    story.extend(_build_cc_header(banca, titulo, W))

    clientes_base = Cliente.objects.all()
    filtro = _user_filter_direct_from_request(request)
    if filtro:
        clientes_base = clientes_base.filter(**filtro)
    cliente_ids = list(clientes_base.values_list('pk', flat=True))

    headers = ['Período', 'Faturação', 'Recebimentos', 'Créditos', 'Débitos', 'Saldo']
    t_data = [[Paragraph(h, s_header) for h in headers]]

    for data_ini, data_fim, label in intervalos:
        agg = _agregados_periodo(cliente_ids, data_ini, data_fim)
        fat = agg['faturacao']
        rec = agg['recebimentos']
        cr = agg['creditos']
        db = agg['debitos']
        saldo = rec - fat
        row_data = [
            Paragraph(label, s_normal),
            Paragraph(f'{fmt_kz(fat)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(rec)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(cr)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(db)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica', alignment=2)),
            Paragraph(f'{fmt_kz(saldo)}', ParagraphStyle('r', fontSize=8, fontName='Helvetica-Bold', alignment=2)),
        ]
        t_data.append(row_data)

    col_widths = [4*cm, (W-4*cm)*0.2, (W-4*cm)*0.2, (W-4*cm)*0.2, (W-4*cm)*0.2, (W-4*cm)*0.2]
    t = Table(t_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), cor_cabecalho),
        ('GRID', (0, 0), (-1, -1), 0.4, cor_borda),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, cor_linha_par]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(t)
    story.extend(_build_cc_footer(banca, W))
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    periodo_label = tipo_periodo.replace('_', '-')
    response['Content-Disposition'] = f'inline; filename="conta_corrente_{periodo_label}_{ano}.pdf"'
    return response
