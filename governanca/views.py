import json
import logging
import random
import hashlib
import hmac
import time
import urllib.parse

logger = logging.getLogger(__name__)

from django.conf import settings
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ValidationError

from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from django.core.cache import cache
from utils.cache_utils import cache_get_or_set, safe_cache_key, cache_invalidate_prefix

from users.models import Usuario
from users.auth_decorators import sessao_expirada, limpar_sessao
from utils.email_utils import _enviar
from utils.format_kz import fmt_kz, parse_kz
from .models import (
    QuotaConfig, QuotaGerada, PagamentoQuota, EstadoFinanceiro,
    CertidaoRegularidade, CarteiraProfissional,
    CategoriaMembro, TipoQuota, IsencaoMembro, HistoricoQuota,
    Assembleia, PautaVotacao, PresencaAssembleia,
    Procuracao, Voto, ReciboVoto, ManifestoIntegridade, AtaDigital, Notificacao,
    DocumentoAssembleia, MembroMesa, MensagemChat,
    ConsultaPublica, ArtigoDocumento, Comentario,
    VotacaoConsulta, VotoConsulta, RelatorioConsulta,
    Convocatoria, RespostaPresenca, LogAssembleia,
)


# â”€â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _get_usuario(request):
    if not request.session.get('usuario_id'):
        return None
    if sessao_expirada(request):
        limpar_sessao(request)
        return None
    return request.session.get('usuario')


def _requer_login(view_func):
    def wrapper(request, *args, **kwargs):
        usuario = _get_usuario(request)
        if not usuario:
            return redirect('login')
        uid = (request.session.get('banca_usuario_id')
               or request.session.get('usuario_id'))
        if not uid:
            return redirect('login')
        request.session['login_time'] = time.time()
        request.session.modified = True
        request.usuario_obj = Usuario.objects.get(id=uid)
        return view_func(request, *args, **kwargs)
    return wrapper


def _criar_notificacao(usuario_id, tipo, titulo, mensagem='', link=''):
    Notificacao.objects.create(
        usuario_id=usuario_id,
        tipo=tipo,
        titulo=titulo,
        mensagem=mensagem,
        link=link,
    )
    cache_invalidate_prefix(f'dash_governanca_{usuario_id}')


def _notificar_para_papel(papel, tipo, titulo, mensagem='', link=''):
    usuarios = list(Usuario.objects.filter(papel=papel, status='Ativo').values_list('id', flat=True))
    if not usuarios:
        return
    if getattr(settings, 'REDIS_ENABLED', False):
        from governanca.tasks import notificar_utilizadores_task
        notificar_utilizadores_task.delay(usuarios, tipo, titulo, mensagem, link)
    else:
        Notificacao.objects.bulk_create([
            Notificacao(usuario_id=u_id, tipo=tipo, titulo=titulo, mensagem=mensagem, link=link)
            for u_id in usuarios
        ])
        for u_id in usuarios:
            cache_invalidate_prefix(f'dash_governanca_{u_id}')


def _enviar_convocatorias_email(assembleia):
    """Envia email das convocatórias publicadas aos membros quando a assembleia inicia."""
    from django.conf import settings as dj_settings
    from utils.email_utils import _enviar
    convocatorias = list(assembleia.convocatorias.filter(status='Publicada'))
    if not convocatorias:
        return
    topicos = '\n'.join(f'  • {c.titulo}' for c in convocatorias)
    data_hora = assembleia.data_hora.strftime('%d/%m/%Y às %H:%M')
    site_url = dj_settings.SITE_URL
    assunto = f'Assembleia em Curso: {assembleia.titulo}'
    texto = (
        f'Prezado(a) Membro,\n\n'
        f'A assembleia "{assembleia.titulo}" teve início às {data_hora}.\n\n'
        f'Tema: {assembleia.descricao or "—"}\n'
        f'Local: {assembleia.local}\n\n'
        f'Convocatórias publicadas:\n{topicos}\n\n'
        f'Aceda à sala virtual: {site_url}/governanca/assembleia/{assembleia.pk}/sala/\n\n'
        f'Atenciosamente,\nCDOA'
    )
    destinatarios = list(
        Usuario.objects.filter(
            status='Ativo',
            papel__in=('Administrador', 'Despachante Oficial'),
        ).exclude(email='').values_list('email', flat=True)
    )
    if destinatarios:
        _enviar(assunto, texto, None, destinatarios)


def _gerar_otp():
    return f'{random.randint(100000, 999999)}'


def _b64url(data):
    return __import__('base64').urlsafe_b64encode(data).rstrip(b'=').decode()

def _livekit_token(room_name, identity):
    api_key = settings.LIVEKIT_API_KEY
    api_secret = settings.LIVEKIT_API_SECRET
    header = {'alg': 'HS256', 'typ': 'JWT'}
    now = int(time.time())
    payload = {
        'iss': api_key,
        'sub': identity,
        'nbf': now - 10,
        'exp': now + 3600,
        'identity': identity,
        'video': {
            'room': room_name,
            'roomJoin': True,
            'canPublish': True,
            'canSubscribe': True,
            'canPublishSources': ['camera', 'microphone', 'screen_share'],
        },
    }
    header_b64 = _b64url(json.dumps(header, separators=(',', ':')).encode())
    payload_b64 = _b64url(json.dumps(payload, separators=(',', ':')).encode())
    sig = hmac.new(api_secret.encode(), f'{header_b64}.{payload_b64}'.encode(), hashlib.sha256).digest()
    sig_b64 = _b64url(sig)
    return f'{header_b64}.{payload_b64}.{sig_b64}'


def _broadcast_ws(assembleia_id, event_type, data):
    """Envia evento WebSocket para todos os conectados na sala da assembleia."""
    try:
        layer = get_channel_layer()
        if layer is None:
            logger.warning('channel_layer is None')
            return
        async_to_sync(layer.group_send)(
            f'assembleia_{assembleia_id}',
            {'type': event_type, 'data': data},
        )
    except Exception as e:
        logger.exception('Erro no broadcast WS: %s', e)


def _get_client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def _log_assembleia(assembleia_id, usuario_id, acao, detalhes=None, ip=''):
    LogAssembleia.objects.create(
        assembleia_id=assembleia_id,
        usuario_id=usuario_id,
        acao=acao,
        detalhes=detalhes or {},
        ip=ip or '',
    )


def _verificar_elegibilidade(usuario_id):
    ef = EstadoFinanceiro.objects.filter(despachante_id=usuario_id).first()
    if ef and ef.estado in ('Irregular', 'Suspenso'):
        return False, 'Estado financeiro irregular — direito de voto suspenso. Acesso ao streaming autorizado.'
    return True, ''


# â”€â”€â”€ Páginas Principais â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@_requer_login
def index(request):
    usuario_obj = request.usuario_obj
    hoje = timezone.now()

    from users.permissoes import _is_admin_ou_acesso_total
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel == 'Administrador':
        qs_quotas = QuotaGerada.objects.all()
        quotas_pendentes = qs_quotas.filter(status='Pendente').count()
        quotas_pagas = qs_quotas.filter(status='Paga').count()
    else:
        qs_quotas = QuotaGerada.objects.filter(despachante=usuario_obj)
        quotas_pendentes = qs_quotas.filter(status='Pendente').count()
        quotas_pagas = qs_quotas.filter(status='Paga').count()
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleias': Assembleia.objects.all()[:5],
        'proximas': Assembleia.objects.filter(status='Agendada', data_hora__gte=hoje).order_by('data_hora')[:5],
        'assembleias_em_curso': Assembleia.objects.filter(status='Em Curso')[:5],
        'total_assembleias': Assembleia.objects.count(),
        'documentos_recentes': DocumentoAssembleia.objects.filter(publicado=True).select_related('assembleia', 'created_by').order_by('-created_at')[:5],
        'total_docs_publicados': DocumentoAssembleia.objects.filter(publicado=True).count(),
        'notificacoes_nao_lidas': Notificacao.objects.filter(usuario=usuario_obj, lida=False).count(),
        'quotas_pendentes': quotas_pendentes,
        'quotas_pagas': quotas_pagas,
        'assembleias_concluidas': Assembleia.objects.filter(status='Concluida').order_by('-data_hora'),
    }
    return render(request, 'governanca/index.html', context)


@_requer_login
def lista_assembleias(request):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_assembleia') and not usuario_tem_permissao(request, 'gerir_governanca'):
        return redirect('governanca_index')

    status_filtro = request.GET.get('status', '')
    qs = Assembleia.objects.all()
    agora = timezone.now()
    qs.filter(status='Agendada', data_hora__lte=agora).update(status='Em Curso')
    
    # Contagem por status
    from django.db.models import Count
    total_assembleias = Assembleia.objects.count()
    agendadas_count = Assembleia.objects.filter(status='Agendada').count()
    em_curso_count = Assembleia.objects.filter(status='Em Curso').count()
    concluidas_count = Assembleia.objects.filter(status='Concluida').count()
    canceladas_count = Assembleia.objects.filter(status='Cancelada').count()
    
    if status_filtro:
        qs = qs.filter(status=status_filtro)
    
    STATUS_CHOICES = [
        ('', 'Todas'),
        ('Agendada', 'Agendadas'),
        ('Em Curso', 'Em Curso'),
        ('Concluida', 'Concluidas'),
        ('Cancelada', 'Canceladas'),
    ]
    
    # Usar status_counts com chaves que correspondem aos valores do STATUS_CHOICES
    STATUS_COUNTS = [
        ('Agendada', agendadas_count),
        ('Em Curso', em_curso_count),
        ('Concluida', concluidas_count),
        ('Cancelada', canceladas_count),
    ]
    
    paginator = Paginator(qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'assembleias',
        'assembleias': page_obj,
        'page_obj': page_obj,
        'status_atual': status_filtro,
        'status_choices': STATUS_CHOICES,
        'status_counts': STATUS_COUNTS,
    }
    return render(request, 'governanca/lista_assembleias.html', context)


@_requer_login
def nova_assembleia(request):
    from users.permissoes import usuario_tem_permissao
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia') and not usuario_tem_permissao(request, 'gerir_governanca'):
        messages.error(request, 'Sem permissão para criar assembleias.')
        return redirect('governanca_assembleias')

    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        descricao = request.POST.get('descricao', '').strip()
        data_hora_str = request.POST.get('data_hora', '').strip()
        livekit_room = request.POST.get('livekit_room', '').strip()
        iniciar_agora = request.POST.get('iniciar_agora') == 'on'
        quorum_minimo = request.POST.get('quorum_minimo', '').strip()

        if not titulo:
            messages.error(request, 'Preencha todos os campos obrigatórios.')
            return render(request, 'governanca/nova_assembleia.html', {**locals()})

        if iniciar_agora:
            data_hora = timezone.now()
        else:
            if not data_hora_str:
                messages.error(request, 'Preencha a data e hora.')
                return render(request, 'governanca/nova_assembleia.html', {**locals()})
            from zoneinfo import ZoneInfo
            from django.utils.dateparse import parse_datetime
            from django.utils.timezone import make_aware
            data_hora = parse_datetime(data_hora_str)
            if not data_hora:
                messages.error(request, 'Data/hora inválida.')
                return render(request, 'governanca/nova_assembleia.html', {**locals()})
            if timezone.is_naive(data_hora):
                data_hora = make_aware(data_hora, timezone=ZoneInfo('Africa/Luanda'))
            if data_hora <= timezone.now():
                messages.error(request, 'A data e hora deve ser posterior ao momento atual.')
                return render(request, 'governanca/nova_assembleia.html', {**locals()})

        if not livekit_room:
            livekit_room = f'assembleia-{int(time.time())}'

        if iniciar_agora:
            status = 'Em Curso'
        else:
            status = 'Agendada'

        total_ativos = Usuario.objects.filter(status='Ativo', papel__in=['Administrador', 'Despachante Oficial', 'Colaborador Institucional']).count()
        try:
            quorum_minimo_val = int(quorum_minimo) if quorum_minimo else total_ativos
        except (ValueError, TypeError):
            quorum_minimo_val = total_ativos

        try:
            assembleia = Assembleia.objects.create(
                titulo=titulo,
                descricao=descricao,
                data_hora=data_hora,
                status=status,
                local='Sala Virtual CDOA',
                livekit_room=livekit_room,
                quorum_minimo=quorum_minimo_val,
                total_eleitores=total_ativos,
                max_procuracao=1,
                created_by=request.usuario_obj,
            )

            if iniciar_agora:
                _notificar_para_papel(
                    'Administrador', 'assembleia_iniciada',
                    f'Assembleia em curso: {titulo}',
                    f'Assembleia iniciada instantaneamente. Entre na sala virtual!',
                    f'/governanca/assembleia/{assembleia.pk}/sala/'
                )
                _notificar_para_papel(
                    'Despachante Oficial', 'assembleia_iniciada',
                    f'Assembleia em curso: {titulo}',
                    f'Assembleia iniciada instantaneamente. Entre na sala virtual!',
                    f'/governanca/assembleia/{assembleia.pk}/sala/'
                )
            else:
                _notificar_para_papel(
                    'Administrador', 'assembleia_agendada',
                    f'Nova Assembleia: {titulo}',
                    f'Foi agendada uma nova assembleia para {data_hora:%d/%m/%Y às %H:%M}.',
                    f'/governanca/assembleia/{assembleia.pk}/'
                )
                _notificar_para_papel(
                    'Despachante Oficial', 'assembleia_agendada',
                    f'Assembleia Agendada: {titulo}',
                    f'Foi agendada uma assembleia para {data_hora:%d/%m/%Y às %H:%M}. Participe!',
                    f'/governanca/assembleia/{assembleia.pk}/'
                )

            for u in Usuario.objects.filter(status='Ativo', papel__in=['Administrador', 'Despachante Oficial']).exclude(email=''):
                if iniciar_agora:
                    assunto = f'Assembleia em curso: {titulo}'
                    corpo = f'Prezado(a) {u.nome},\n\nA assembleia "{titulo}" foi iniciada e já está em curso.\n\nEntre na sala virtual: {settings.SITE_URL}/governanca/assembleia/{assembleia.pk}/sala/\n\nAtenciosamente,\nCDOA'
                else:
                    assunto = f'Assembleia Agendada: {titulo}'
                    corpo = f'Prezado(a) {u.nome},\n\nFoi agendada uma nova assembleia:\n\n  Título: {titulo}\n  Data: {data_hora:%d/%m/%Y às %H:%M}\n  Descrição: {descricao}\n\nParticipe em: {settings.SITE_URL}/governanca/assembleia/{assembleia.pk}/\n\nAtenciosamente,\nCDOA'
                _enviar(assunto, corpo, None, [u.email])

            pautas_titulos = request.POST.getlist('pauta_titulo[]')
            pautas_descricoes = request.POST.getlist('pauta_descricao[]')
            pautas_tipos = request.POST.getlist('pauta_tipo[]')
            for i, titulo in enumerate(pautas_titulos):
                if titulo.strip():
                    PautaVotacao.objects.create(
                        assembleia=assembleia,
                        titulo=titulo.strip(),
                        descricao=(pautas_descricoes[i] if i < len(pautas_descricoes) else ''),
                        tipo_votacao=(pautas_tipos[i] if i < len(pautas_tipos) else 'Aberta'),
                        ordem=i + 1,
                    )

            messages.success(request, 'Assembleia criada com sucesso!')
            if iniciar_agora:
                return redirect('governanca_detalhe', pk=assembleia.pk)
            else:
                messages.info(request, 'Agora crie a convocatória para esta assembleia.')
                return redirect('governanca_criar_convocatoria', pk=assembleia.pk)
        except ValidationError as e:
            if hasattr(e, 'message_dict'):
                for field, messages_list in e.message_dict.items():
                    for msg in messages_list:
                        messages.error(request, msg)
            else:
                messages.error(request, str(e))
            return render(request, 'governanca/nova_assembleia.html', {**locals()})

    total_ativos = Usuario.objects.filter(status='Ativo', papel__in=['Administrador', 'Despachante Oficial', 'Colaborador Institucional']).count()
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'assembleias',
        'total_ativos': total_ativos,
    }
    return render(request, 'governanca/nova_assembleia.html', context)


@_requer_login
def detalhe_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_assembleia') and not usuario_tem_permissao(request, 'gerir_votacoes') and not usuario_tem_permissao(request, 'gerir_convocatorias'):
        return redirect('governanca_index')
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if assembleia.status == 'Agendada' and assembleia.data_hora <= timezone.now():
        assembleia.status = 'Em Curso'
        assembleia.save(update_fields=['status'])
    usuario_id = request.session['usuario_id']
    minhas_procuracao = Procuracao.objects.filter(outorgante_id=usuario_id, assembleia=assembleia)
    procuracao_recebidas = Procuracao.objects.filter(outorgado_id=usuario_id, assembleia=assembleia)
    tenho_presenca = PresencaAssembleia.objects.filter(assembleia=assembleia, usuario_id=usuario_id, presente_em__isnull=False).exists()
    ja_votei_pautas = set(
        Voto.objects.filter(pauta__assembleia=assembleia, usuario_id=usuario_id, em_delegacao=False)
        .values_list('pauta_id', flat=True)
    )
    documentos = assembleia.documentos.filter(publicado=True)

    minha_resposta_obj = RespostaPresenca.objects.filter(assembleia=assembleia, usuario_id=usuario_id).first()

    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleia': assembleia,
        'minhas_procuracao': minhas_procuracao,
        'procuracao_recebidas': procuracao_recebidas,
        'tenho_presenca': tenho_presenca,
        'ja_votei_pautas': ja_votei_pautas,
        'despachantes': Usuario.objects.filter(papel='Despachante Oficial', status='Ativo').exclude(id=usuario_id),
        'documentos': documentos,
        'minha_resposta': minha_resposta_obj.resposta if minha_resposta_obj else None,
    }
    return render(request, 'governanca/detalhe_assembleia.html', context)


@_requer_login
def editar_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        messages.error(request, 'Sem permissão.')
        return redirect('governanca_detalhe', pk=pk)

    if request.method == 'POST':
        assembleia.titulo = request.POST.get('titulo', assembleia.titulo)
        assembleia.descricao = request.POST.get('descricao', assembleia.descricao)
        data_hora_str = request.POST.get('data_hora', '').strip()
        if data_hora_str:
            from django.utils.dateparse import parse_datetime
            nova_data = parse_datetime(data_hora_str)
            if not nova_data:
                messages.error(request, 'Formato de data inválido. Use AAAA-MM-DD HH:MM.')
                return redirect('governanca_editar', pk=pk)
            if nova_data <= timezone.now():
                messages.error(request, 'A data da assembleia deve ser posterior ao momento atual.')
                return redirect('governanca_editar', pk=pk)
            assembleia.data_hora = nova_data
        assembleia.link_streaming = request.POST.get('link_streaming', assembleia.link_streaming)
        assembleia.local = request.POST.get('local', '').strip() or 'Sala Virtual CDOA'
        assembleia.livekit_room = request.POST.get('livekit_room', assembleia.livekit_room)
        quorum_str = request.POST.get('quorum_minimo', '').strip()
        if quorum_str:
            try:
                assembleia.quorum_minimo = int(quorum_str)
            except (ValueError, TypeError):
                pass
        assembleia.save()

        pautas_titulos = request.POST.getlist('pauta_titulo[]')
        pautas_descricoes = request.POST.getlist('pauta_descricao[]')
        pautas_tipos = request.POST.getlist('pauta_tipo[]')
        for i, titulo in enumerate(pautas_titulos):
            if titulo.strip():
                PautaVotacao.objects.create(
                    assembleia=assembleia,
                    titulo=titulo.strip(),
                    descricao=(pautas_descricoes[i] if i < len(pautas_descricoes) else ''),
                    tipo_votacao=(pautas_tipos[i] if i < len(pautas_tipos) else 'Aberta'),
                    ordem=i + 1,
                )

        messages.success(request, 'Assembleia atualizada!')
        return redirect('governanca_detalhe', pk=pk)

    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleia': assembleia,
    }
    return render(request, 'governanca/editar_assembleia.html', context)


# â”€â”€â”€ Sala da Assembleia (Fase 2) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@_requer_login
def sala_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_assembleia') and not usuario_tem_permissao(request, 'gerir_votacoes') and not usuario_tem_permissao(request, 'gerir_convocatorias'):
        return redirect('governanca_index')
    assembleia = get_object_or_404(Assembleia, pk=pk)
    usuario_id = request.session['usuario_id']

    if assembleia.status == 'Agendada':
        if assembleia.data_hora <= timezone.now():
            assembleia.status = 'Em Curso'
            assembleia.save()
        else:
            messages.warning(request, 'A assembleia ainda não começou.')
            return redirect('governanca_detalhe', pk=pk)

    presencia, created = PresencaAssembleia.objects.get_or_create(
        assembleia=assembleia,
        usuario_id=usuario_id,
        defaults={'presente_em': timezone.now()},
    )
    if not presencia.presente_em:
        presencia.presente_em = timezone.now()
        presencia.save()

    Assembleia.objects.filter(pk=assembleia.pk).update(ultima_actividade=timezone.now())

    pauta_ativa = assembleia.pautas.filter(status='Em Votacao').first()

    if not MembroMesa.objects.filter(assembleia=assembleia).exists() and assembleia.created_by:
        MembroMesa.objects.get_or_create(
            assembleia=assembleia, usuario=assembleia.created_by,
            defaults={'funcao': 'Presidente', 'ordem': 0},
        )

    minhas_procuracao = Procuracao.objects.filter(
        outorgado_id=usuario_id, assembleia=assembleia, status='Confirmada'
    ).select_related('outorgante')
    # minhas_procuracao: procurações onde sou outorgado (recebi o poder de voto)

    pautas_ja_votadas = set(
        Voto.objects.filter(usuario_id=usuario_id, em_delegacao=False)
        .values_list('pauta_id', flat=True)
    )
    pautas_voto_delegado = set(
        Voto.objects.filter(
            usuario_id=usuario_id, em_delegacao=True, delegado_de__isnull=False
        ).values_list('pauta_id', 'delegado_de_id')
    )

    livekit_token = ''
    if assembleia.livekit_room:
        livekit_token = _livekit_token(
            assembleia.livekit_room,
            request.session['usuario']['nome'],
        )

    elegivel, _ = _verificar_elegibilidade(usuario_id)

    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleia': assembleia,
        'pautas': assembleia.pautas.with_vote_counts(),
        'pauta_ativa': pauta_ativa,
        'minhas_procuracao': minhas_procuracao,
        'pautas_ja_votadas': pautas_ja_votadas,
        'pautas_voto_delegado': pautas_voto_delegado,
        'livekit_token': livekit_token,
        'livekit_url': settings.LIVEKIT_URL,
        'ws_url': f'{"ws" if not request.is_secure() else "wss"}://{request.get_host()}/ws/assembleia/{pk}/',
        'presentes': assembleia.presencas.filter(presente_em__isnull=False).select_related('usuario'),
        'mesa': MembroMesa.objects.filter(assembleia=assembleia).select_related('usuario'),
        'despachantes': Usuario.objects.filter(
            papel='Despachante Oficial', status='Ativo'
        ).exclude(id=usuario_id) if request.session['usuario']['papel'] == 'Administrador' else [],
        'elegivel': elegivel,
    }
    return render(request, 'governanca/sala_assembleia.html', context)


# â”€â”€â”€ Gestão / Mesa (Admin) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@_requer_login
def gerir_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        messages.error(request, 'Sem permissão.')
        return redirect('governanca_detalhe', pk=pk)

    from rh.models import CargoMesa
    mesa_existing = set(MembroMesa.objects.filter(assembleia=assembleia).values_list('usuario_id', flat=True))
    membros_direcao = CargoMesa.objects.select_related('usuario').all()
    for cargo_mesa in membros_direcao:
        u = cargo_mesa.usuario
        if u.status == 'Ativo' and u.id not in mesa_existing:
            MembroMesa.objects.get_or_create(
                assembleia=assembleia, usuario=u,
                defaults={'funcao': cargo_mesa.funcao, 'ordem': MembroMesa.objects.filter(assembleia=assembleia).count()},
            )

    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': papel,
        'active_menu': 'Governanca',
        'assembleia': assembleia,
        'pautas': assembleia.pautas.with_vote_counts(),
        'presentes': assembleia.presencas.filter(presente_em__isnull=False).select_related('usuario'),
        'procuracao': Procuracao.objects.filter(assembleia=assembleia).select_related('outorgante', 'outorgado'),
        'documentos': assembleia.documentos.all(),
        'mesa': MembroMesa.objects.filter(assembleia=assembleia).select_related('usuario'),
        'despachantes': Usuario.objects.filter(
            papel__in=['Administrador', 'Despachante Oficial'], status='Ativo'
        ).exclude(id=request.session['usuario_id']),
        'tem_convocatoria_publicada': assembleia.convocatorias.filter(status='Publicada').exists(),
        'direcao_ids': list(membros_direcao.values_list('usuario_id', flat=True)),
    }
    return render(request, 'governanca/gerir_assembleia.html', context)


# â”€â”€â”€ Repositório de Atas â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@_requer_login
def repositorio_atas(request):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_atas'):
        return redirect('governanca_index')

    page_number = request.GET.get('page', '1')
    cache_key = safe_cache_key('repositorio_atas', page_number)

    def _compute():
        atas = AtaDigital.objects.filter(publicado_em__isnull=False).select_related('assembleia', 'assinado_por')
        documentos = DocumentoAssembleia.objects.filter(publicado=True).select_related('assembleia', 'created_by')
        paginator = Paginator(atas, 8)
        page_obj = paginator.get_page(page_number)
        return {
            'atas': page_obj,
            'page_obj': page_obj,
            'documentos': list(documentos),
        }

    cached = cache_get_or_set(cache_key, _compute, timeout=300)
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'atas',
        **cached,
    }
    return render(request, 'governanca/repositorio_atas.html', context)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Presença
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_http_methods(['POST'])
@_requer_login
def api_registar_presenca(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    usuario_id = request.session['usuario_id']
    obj, created = PresencaAssembleia.objects.get_or_create(
        assembleia=assembleia, usuario_id=usuario_id,
        defaults={'presente_em': timezone.now()},
    )
    if not obj.presente_em:
        obj.presente_em = timezone.now()
        obj.save()

    Assembleia.objects.filter(pk=assembleia.pk).update(ultima_actividade=timezone.now())

    _broadcast_ws(pk, 'quorum_update', {
        'presentes': assembleia.presentes_count,
        'quorum_minimo': assembleia.quorum_minimo,
        'atingido': assembleia.quorum_atingido,
        'total_eleitores': assembleia.total_eleitores,
    })

    _log_assembleia(assembleia.id, usuario_id, 'entrada', {}, ip=_get_client_ip(request))

    return JsonResponse({
        'status': 'ok', 'presente': True,
        'presentes_count': assembleia.presentes_count,
        'quorum': assembleia.quorum_minimo,
        'quorum_atingido': assembleia.quorum_atingido,
    })


@_requer_login
def api_listar_presencas(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    presencas = assembleia.presencas.filter(presente_em__isnull=False).select_related('usuario')
    data = [{
        'id': p.usuario.id,
        'nome': p.usuario.nome,
        'email': p.usuario.email,
        'presente_em': p.presente_em.isoformat() if p.presente_em else None,
    } for p in presencas]
    return JsonResponse({
        'presentes': data,
        'total': len(data),
        'quorum_minimo': assembleia.quorum_minimo,
        'quorum_atingido': assembleia.quorum_atingido,
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Procuração
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_http_methods(['POST'])
@_requer_login
def api_solicitar_procuracao(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    outorgante_id = request.session['usuario_id']

    if assembleia.status not in ('Agendada', 'Em Curso'):
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode delegar voto numa assembleia que já foi concluída ou cancelada.'
        }, status=400)

    if assembleia.data_hora < timezone.now() and assembleia.status == 'Agendada':
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode delegar voto para uma assembleia com data já passada.'
        }, status=400)

    existing = Procuracao.objects.filter(assembleia=assembleia, outorgante_id=outorgante_id).first()
    if existing and existing.status != 'Cancelada':
        return JsonResponse({'status': 'error', 'message': 'Já possui uma procuração ativa para esta assembleia.'}, status=400)
    if existing and existing.status == 'Cancelada':
        existing.delete()

    elegivel, msg_elig = _verificar_elegibilidade(outorgante_id)
    if not elegivel:
        return JsonResponse({'status': 'error', 'message': msg_elig}, status=403)

    outorgante = Usuario.objects.get(id=outorgante_id)
    if outorgante.papel != 'Despachante Oficial' or outorgante.status != 'Ativo':
        return JsonResponse({
            'status': 'error',
            'message': 'Apenas despachantes oficiais ativos podem delegar voto.'
        }, status=403)

    data = json.loads(request.body)
    outorgado_id = data.get('outorgado_id')
    if not outorgado_id:
        return JsonResponse({'status': 'error', 'message': 'Selecione um despachante para receber a delegação de voto.'}, status=400)

    if int(outorgado_id) == outorgante_id:
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode delegar voto a si mesmo. Selecione outro colega.'
        }, status=400)

    elegivel_outorgado, msg_outorgado = _verificar_elegibilidade(outorgado_id)
    if not elegivel_outorgado:
        return JsonResponse({
            'status': 'error',
            'message': f'O despachante selecionado não pode receber delegação de voto: {msg_outorgado}'
        }, status=400)

    outorgado = Usuario.objects.get(id=outorgado_id)
    if outorgado.papel != 'Despachante Oficial' or outorgado.status != 'Ativo':
        return JsonResponse({
            'status': 'error',
            'message': 'Apenas despachantes oficiais ativos podem receber delegação de voto.'
        }, status=403)

    # Validar limite de procurações por outorgado
    procuracao_ativas = Procuracao.objects.filter(
        assembleia=assembleia, outorgado_id=outorgado_id, status='Confirmada'
    ).count()
    max_proc = assembleia.max_procuracao
    if procuracao_ativas >= max_proc:
        return JsonResponse({
            'status': 'error',
            'message': f'Limite máximo de {max_proc} procuração(ões) atingido para este membro.'
        }, status=400)

    otp = _gerar_otp()
    otp_hash = hashlib.sha256(otp.encode()).hexdigest()
    procuracao = Procuracao.objects.create(
        assembleia=assembleia,
        outorgante_id=outorgante_id,
        outorgado_id=outorgado_id,
        codigo_otp=otp_hash,
    )
    _criar_notificacao(
        outorgado_id, 'procuracao_solicitada',
        'Procuração Solicitada',
        f'{request.session["usuario"]["nome"]} solicitou-lhe procuração para {assembleia.titulo}.',
        f'/governanca/assembleia/{assembleia.pk}/'
    )
    request.session['otp_plaintext'] = otp
    request.session['otp_procuracao_id'] = procuracao.id

    if outorgante.email:
        _enviar(
            'Código OTP â€” Procuração â€” CDOA',
            f'Prezado(a) {outorgante.nome},\n\n'
            f'O seu código OTP para confirmar a procuração na assembleia "{assembleia.titulo}" é:\n\n'
            f'  {otp}\n\n'
            f'O código expira após a primeira utilização.\n\n'
            f'Atenciosamente,\nSICDOA',
            None,
            [outorgante.email],
        )
    else:
        logger = __import__('logging').getLogger(__name__)
        logger.warning('Outorgante %s não tem email cadastrado â€” OTP não enviado por email', outorgante.nome)
    _log_assembleia(assembleia.id, outorgante_id, 'procuracao_solicitada', {
        'outorgado_id': outorgado_id,
        'procuracao_id': procuracao.id,
    }, ip=_get_client_ip(request))

    return JsonResponse({
        'status': 'ok',
        'procuracao_id': procuracao.id,
        'message': 'Código OTP enviado para o seu email. Verifique a sua caixa de entrada.',
    })


@require_http_methods(['POST'])
@_requer_login
def api_confirmar_procuracao(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    data = json.loads(request.body)
    codigo = data.get('codigo_otp', '')
    procuracao_id = data.get('procuracao_id')
    procuracao = get_object_or_404(Procuracao, pk=procuracao_id, assembleia=assembleia, outorgante_id=request.session['usuario_id'])

    if procuracao.status != 'Pendente':
        return JsonResponse({'status': 'error', 'message': 'Procuração já foi processada.'}, status=400)

    if assembleia.status not in ('Agendada', 'Em Curso'):
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode confirmar procuração para uma assembleia que já foi concluída ou cancelada.'
        }, status=400)

    if assembleia.data_hora < timezone.now() and assembleia.status == 'Agendada':
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode confirmar procuração para uma assembleia com data já passada.'
        }, status=400)

    otp_hash_input = hashlib.sha256(codigo.encode()).hexdigest()
    otp_session = request.session.get('otp_plaintext', '')
    if procuracao.codigo_otp != otp_hash_input and otp_session != codigo:
        return JsonResponse({'status': 'error', 'message': 'Código OTP inválido.'}, status=400)

    if procuracao.created_at and (timezone.now() - procuracao.created_at).total_seconds() > 900:
        procuracao.status = 'Cancelada'
        procuracao.save()
        return JsonResponse({
            'status': 'error',
            'message': 'O código OTP expirou (válido por 15 minutos). Solicite uma nova procuração.'
        }, status=400)

    procuracao.status = 'Confirmada'
    procuracao.confirmado_em = timezone.now()
    procuracao.save()

    _criar_notificacao(
        procuracao.outorgado_id, 'procuracao_confirmada',
        'Procuração Confirmada',
        f'{request.session["usuario"]["nome"]} confirmou a procuração para {assembleia.titulo}. Agora tem um voto delegado.',
        f'/governanca/assembleia/{assembleia.pk}/'
    )

    _broadcast_ws(assembleia.id, 'procuracao_confirmada', {
        'procuracao_id': procuracao.id,
        'outorgante_id': procuracao.outorgante_id,
        'outorgante_nome': request.session['usuario']['nome'],
        'outorgado_id': procuracao.outorgado_id,
    })

    _log_assembleia(assembleia.id, request.session['usuario_id'], 'procuracao_confirmada', {
        'outorgado_id': procuracao.outorgado_id,
        'procuracao_id': procuracao.id,
    }, ip=_get_client_ip(request))
    return JsonResponse({'status': 'ok', 'message': 'Procuração confirmada com sucesso!'})


@_requer_login
def api_minhas_procuracao(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    usuario_id = request.session['usuario_id']
    como_outorgante = Procuracao.objects.filter(assembleia=assembleia, outorgante_id=usuario_id).select_related('outorgado')
    como_outorgado = Procuracao.objects.filter(assembleia=assembleia, outorgado_id=usuario_id).select_related('outorgante')
    data = {
        'como_outorgante': [{
            'id': p.id, 'nome': p.outorgado.nome,
            'status': p.status, 'confirmado_em': p.confirmado_em.isoformat() if p.confirmado_em else None,
        } for p in como_outorgante],
        'como_outorgado': [{
            'id': p.id, 'nome': p.outorgante.nome,
            'outorgante_id': p.outorgante_id,
            'status': p.status, 'confirmado_em': p.confirmado_em.isoformat() if p.confirmado_em else None,
        } for p in como_outorgado],
    }
    return JsonResponse(data)


@require_http_methods(['POST'])
@_requer_login
def api_cancelar_procuracao(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    usuario_id = request.session['usuario_id']
    data = json.loads(request.body)
    procuracao_id = data.get('procuracao_id')

    if not procuracao_id:
        return JsonResponse({'status': 'error', 'message': 'Procuração não identificada.'}, status=400)

    procuracao = get_object_or_404(Procuracao, pk=procuracao_id, assembleia=assembleia, outorgante_id=usuario_id)

    if procuracao.status not in ('Pendente', 'Confirmada'):
        return JsonResponse({'status': 'error', 'message': 'Esta procuração já foi cancelada ou expirou.'}, status=400)

    if assembleia.status not in ('Agendada', 'Em Curso'):
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode cancelar procuração de uma assembleia que já foi concluída ou cancelada.'
        }, status=400)

    procuracao.status = 'Cancelada'
    procuracao.save()

    _criar_notificacao(
        procuracao.outorgado_id, 'procuracao_cancelada',
        'Procuração Cancelada',
        f'{request.session["usuario"]["nome"]} cancelou a procuração para {assembleia.titulo}. Já não tem voto delegado.',
        f'/governanca/assembleia/{assembleia.pk}/'
    )

    _log_assembleia(assembleia.id, usuario_id, 'procuracao_cancelada', {
        'outorgado_id': procuracao.outorgado_id,
        'procuracao_id': procuracao.id,
    }, ip=_get_client_ip(request))

    request.session.pop('otp_plaintext', None)
    request.session.pop('otp_procuracao_id', None)

    return JsonResponse({'status': 'ok', 'message': 'Procuração cancelada com sucesso.'})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Votação
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_http_methods(['POST'])
@_requer_login
def api_iniciar_votacao(request, pk):
    from users.permissoes import usuario_tem_permissao
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_votacoes'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores podem abrir votações.'}, status=403)
    if pauta.assembleia.status != 'Em Curso':
        return JsonResponse({'status': 'error', 'message': 'Assembleia não está em curso.'}, status=400)
    if pauta.status == 'Concluida':
        return JsonResponse({'status': 'error', 'message': 'Votação já foi concluída.'}, status=400)
    if pauta.assembleia.presentes_count < pauta.assembleia.quorum_minimo:
        return JsonResponse({'status': 'error', 'message': 'Quórum mínimo não atingido para iniciar votação.'}, status=400)

    pauta.status = 'Em Votacao'
    pauta.iniciado_em = timezone.now()
    pauta.save()

    Assembleia.objects.filter(pk=pauta.assembleia_id).update(ultima_actividade=timezone.now())

    _broadcast_ws(pauta.assembleia_id, 'votacao_aberta', {
        'pauta_id': pauta.id,
        'titulo': pauta.titulo,
        'tipo_votacao': pauta.tipo_votacao,
    })

    _log_assembleia(pauta.assembleia_id, request.session['usuario_id'], 'votacao_aberta', {
        'pauta_id': pauta.id, 'pauta_titulo': pauta.titulo,
    }, ip=_get_client_ip(request))

    return JsonResponse({
        'status': 'ok',
        'pauta_id': pauta.id,
        'titulo': pauta.titulo,
        'tipo_votacao': pauta.tipo_votacao,
    })


@require_http_methods(['POST'])
@_requer_login
def api_encerrar_votacao(request, pk):
    from users.permissoes import usuario_tem_permissao
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_votacoes'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores podem encerrar votações.'}, status=403)
    pauta.status = 'Concluida'
    pauta.encerrado_em = timezone.now()
    pauta.apurar_resultado()

    Assembleia.objects.filter(pk=pauta.assembleia_id).update(ultima_actividade=timezone.now())

    _broadcast_ws(pauta.assembleia_id, 'votacao_encerrada', {
        'pauta_id': pauta.id,
        'resultado_final': pauta.resultado_final,
        'favor': pauta.votos_favor,
        'contra': pauta.votos_contra,
        'abstencao': pauta.votos_abstencao,
        'quorum': pauta.assembleia.presentes_count,
        'quorum_minimo': pauta.assembleia.quorum_minimo,
    })

    _log_assembleia(pauta.assembleia_id, request.session['usuario_id'], 'votacao_encerrada', {
        'pauta_id': pauta.id, 'pauta_titulo': pauta.titulo,
        'resultado': pauta.resultado_final,
    }, ip=_get_client_ip(request))

    return JsonResponse({
        'status': 'ok',
        'resultado': pauta.resultado_final,
        'favor': pauta.votos_favor,
        'contra': pauta.votos_contra,
        'abstencao': pauta.votos_abstencao,
        'total': pauta.total_votos,
    })


@require_http_methods(['POST'])
@_requer_login
def api_votar(request, pk):
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    assembleia = pauta.assembleia

    logger.info('api_votar pauta=%s status=%s tipo=%s', pk, pauta.status, pauta.tipo_votacao)

    if pauta.status != 'Em Votacao':
        logger.warning('api_votar status=%s != Em Votacao', pauta.status)
        return JsonResponse({'status': 'error', 'message': 'Votação não está ativa.'}, status=400)

    if assembleia.status != 'Em Curso':
        return JsonResponse({'status': 'error', 'message': 'Assembleia não está em curso.'}, status=400)

    usuario_id = request.session['usuario_id']

    elegivel, msg = _verificar_elegibilidade(usuario_id)
    if not elegivel:
        logger.warning('api_votar usuario %s não elegível: %s', usuario_id, msg)
        return JsonResponse({'status': 'error', 'message': msg}, status=403)

    data = json.loads(request.body)
    opcao = data.get('opcao', '')
    if opcao not in ('Favor', 'Contra', 'Abstencao'):
        logger.warning('api_votar opção inválida: %s', opcao)
        return JsonResponse({'status': 'error', 'message': 'Opção inválida.'}, status=400)

    em_delegacao = data.get('em_delegacao', False)
    delegado_de_id = data.get('delegado_de_id')

    if em_delegacao:
        if not delegado_de_id:
            return JsonResponse({'status': 'error', 'message': 'Identifique o membro em nome de quem está a votar.'}, status=400)
        procuracao_valida = Procuracao.objects.filter(
            assembleia=assembleia,
            outorgante_id=delegado_de_id,
            outorgado_id=usuario_id,
            status='Confirmada',
        ).exists()
        if not procuracao_valida:
            return JsonResponse({
                'status': 'error',
                'message': 'Não possui uma procuração válida para votar em nome deste membro.'
            }, status=403)

    filtro_dup = {'pauta': pauta, 'usuario_id': usuario_id, 'em_delegacao': em_delegacao}
    if em_delegacao and delegado_de_id:
        filtro_dup['delegado_de_id'] = delegado_de_id
    if Voto.objects.filter(**filtro_dup).exists():
        tipo = 'procuração' if em_delegacao else 'pessoal'
        logger.warning('api_votar usuario %s já votou %s pauta=%s', usuario_id, tipo, pk)
        return JsonResponse({'status': 'error', 'message': 'Já votou nesta pauta.'}, status=400)

    with transaction.atomic():
        voto = Voto.objects.create(
            pauta=pauta,
            usuario_id=usuario_id,
            opcao=opcao,
            em_delegacao=em_delegacao,
            delegado_de_id=delegado_de_id if em_delegacao else None,
        )
        if pauta.tipo_votacao == 'Secreta':
            ReciboVoto.objects.create(
                voto=voto,
                recibo_hash=voto.recibo_hash,
                pauta_titulo=pauta.titulo,
                data_voto=voto.votado_em,
            )
    Assembleia.objects.filter(pk=assembleia.pk).update(ultima_actividade=timezone.now())
    log_detalhes = {
        'pauta_id': pauta.id, 'pauta_titulo': pauta.titulo,
        'em_delegacao': em_delegacao,
    }
    if pauta.tipo_votacao == 'Aberta':
        log_detalhes['opcao'] = opcao
    _log_assembleia(assembleia.id, usuario_id, 'votacao', log_detalhes, ip=_get_client_ip(request))

    if pauta.tipo_votacao == 'Aberta':
        _broadcast_ws(assembleia.id, 'voto_registado', {
            'action': 'voto_registado',
            'pauta_id': pauta.id,
            'nome': request.usuario_obj.nome,
            'opcao': opcao,
            'tipo_votacao': 'Aberta',
        })

    _broadcast_ws(assembleia.id, 'resultados_update', {
        'pauta_id': pauta.id,
        'titulo': pauta.titulo,
        'favor': pauta.votos_favor,
        'contra': pauta.votos_contra,
        'abstencao': pauta.votos_abstencao,
        'total': pauta.total_votos,
        'status': pauta.status,
    })

    return JsonResponse({
        'status': 'ok',
        'hash_auditoria': voto.hash_auditoria,
        'recibo_hash': voto.recibo_hash,
        'total_votos': pauta.total_votos,
        'tipo_votacao': pauta.tipo_votacao,
    })


@_requer_login
def api_resultados_pauta(request, pk):
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    return JsonResponse({
        'pauta_id': pauta.id,
        'titulo': pauta.titulo,
        'status': pauta.status,
        'resultado_final': pauta.resultado_final,
        'favor': pauta.votos_favor,
        'contra': pauta.votos_contra,
        'abstencao': pauta.votos_abstencao,
        'total': pauta.total_votos,
        'votos_delegados': pauta.votos_delegados,
        'tipo_votacao': pauta.tipo_votacao,
    })


@_requer_login
def api_votos_pauta(request, pk):
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    votos = Voto.objects.filter(pauta=pauta).select_related('usuario', 'delegado_de').order_by('votado_em')
    data = []
    for v in votos:
        item = {'votado_em': v.votado_em.isoformat(), 'em_delegacao': v.em_delegacao}
        if pauta.tipo_votacao == 'Aberta':
            if v.em_delegacao and v.delegado_de:
                item['nome'] = f"{v.delegado_de.nome} (delegado por {v.usuario.nome})"
            else:
                item['nome'] = v.usuario.nome
            item['opcao'] = v.opcao
        else:
            item['nome'] = '***'
            item['opcao'] = ''
        data.append(item)
    return JsonResponse({
        'votos': data,
        'tipo_votacao': pauta.tipo_votacao,
        'total': pauta.total_votos,
        'favor': pauta.votos_favor,
        'contra': pauta.votos_contra,
        'abstencao': pauta.votos_abstencao,
    })


@_requer_login
def api_verificar_voto(request, pk):
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    recibo_hash = request.GET.get('recibo_hash', '')
    if not recibo_hash:
        return JsonResponse({'status': 'error', 'message': 'recibo_hash é obrigatório.'}, status=400)

    voto = Voto.objects.filter(pauta=pauta, recibo_hash=recibo_hash).first()
    if not voto:
        return JsonResponse({'status': 'error', 'message': 'Voto não encontrado.'}, status=404)

    return JsonResponse({
        'status': 'ok',
        'verificado': True,
        'pauta_titulo': pauta.titulo,
        'opcao': voto.opcao if pauta.tipo_votacao == 'Aberta' else '***',
        'hash_auditoria': voto.hash_auditoria,
        'votado_em': voto.votado_em.isoformat(),
        'em_delegacao': voto.em_delegacao,
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Assembleia (Status / Iniciar / Concluir / Cancelar)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@_requer_login
def api_status_assembleia(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    pauta_ativa = assembleia.pautas.filter(status='Em Votacao').first()
    return JsonResponse({
        'id': assembleia.id,
        'status': assembleia.status,
        'presentes': assembleia.presentes_count,
        'quorum_minimo': assembleia.quorum_minimo,
        'quorum_atingido': assembleia.quorum_atingido,
        'total_pautas': assembleia.total_pautas,
        'pautas_concluidas': assembleia.pautas_concluidas,
        'pauta_ativa_id': pauta_ativa.id if pauta_ativa else None,
        'pauta_ativa_titulo': pauta_ativa.titulo if pauta_ativa else '',
        'pauta_ativa_tipo': pauta_ativa.tipo_votacao if pauta_ativa else '',
    })


@require_http_methods(['POST'])
@_requer_login
def api_iniciar_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores podem iniciar assembleias.'}, status=403)
    if assembleia.status != 'Agendada':
        return JsonResponse({'status': 'error', 'message': 'Assembleia já foi iniciada ou concluída.'}, status=400)
    if not assembleia.convocatorias.filter(status='Publicada').exists():
        return JsonResponse({'status': 'error', 'message': 'É necessário publicar pelo menos uma Convocatória antes de iniciar a assembleia.'}, status=400)
    assembleia.status = 'Em Curso'
    assembleia.save()
    _enviar_convocatorias_email(assembleia)
    _notificar_para_papel('Administrador', 'assembleia_iniciada', f'Assembleia em curso: {assembleia.titulo}', 'A assembleia já está em curso. Entre na sala virtual!', f'/governanca/assembleia/{assembleia.pk}/sala/')
    _notificar_para_papel('Despachante Oficial', 'assembleia_iniciada', f'Assembleia em curso: {assembleia.titulo}', 'A assembleia já está em curso. Entre na sala virtual!', f'/governanca/assembleia/{assembleia.pk}/sala/')
    _log_assembleia(assembleia.id, request.session['usuario_id'], 'assembleia_iniciada', {}, ip=_get_client_ip(request))
    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
@_requer_login
def api_concluir_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores podem concluir assembleias.'}, status=403)
    if assembleia.status != 'Em Curso':
        return JsonResponse({'status': 'error', 'message': 'Assembleia não está em curso.'}, status=400)

    pautas_em_votacao = assembleia.pautas.with_vote_counts().filter(status='Em Votacao')
    for pauta in pautas_em_votacao:
        pauta.status = 'Concluida'
        pauta.encerrado_em = timezone.now()
        pauta.apurar_resultado()

    assembleia.status = 'Concluida'
    assembleia.data_encerramento = timezone.now()
    assembleia.hash_integridade = assembleia.gerar_hash_integridade()
    assembleia.save()

    todas_pautas = assembleia.pautas.with_vote_counts()
    manifesto = ManifestoIntegridade.objects.create(
        assembleia=assembleia,
        hash_consolidado=assembleia.hash_integridade,
        dados_json=json.dumps({
            'presentes': assembleia.presentes_count,
            'total_pautas': assembleia.total_pautas,
            'pautas': [{'id': p.id, 'titulo': p.titulo, 'resultado': p.resultado_final} for p in todas_pautas],
        }, ensure_ascii=False),
        gerado_por=request.usuario_obj,
    )

    _notificar_para_papel('Administrador', 'resultado_publicado', f'Assembleia Concluida: {assembleia.titulo}', 'Os resultados já estão disponíveis.', f'/governanca/assembleia/{assembleia.pk}/')
    _log_assembleia(assembleia.id, request.session['usuario_id'], 'assembleia_concluida', {
        'hash': assembleia.hash_integridade,
        'total_pautas': assembleia.total_pautas,
    }, ip=_get_client_ip(request))
    return JsonResponse({'status': 'ok', 'hash': assembleia.hash_integridade})


@require_http_methods(['POST'])
@_requer_login
def api_cancelar_assembleia(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores podem cancelar assembleias.'}, status=403)
    assembleia.status = 'Cancelada'
    assembleia.save()
    _log_assembleia(assembleia.id, request.session['usuario_id'], 'assembleia_cancelada', {}, ip=_get_client_ip(request))
    return JsonResponse({'status': 'ok'})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Manifesto / Ata
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_http_methods(['POST'])
@_requer_login
def api_gerar_manifesto(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if assembleia.status != 'Concluida':
        return JsonResponse({'status': 'error', 'message': 'Assembleia precisa estar concluída.'}, status=400)
    hash_val = assembleia.gerar_hash_integridade()
    manifesto, created = ManifestoIntegridade.objects.get_or_create(
        assembleia=assembleia,
        defaults={
            'hash_consolidado': hash_val,
            'dados_json': json.dumps({'gerado_em': str(timezone.now())}, ensure_ascii=False),
            'gerado_por': request.usuario_obj,
        },
    )
    return JsonResponse({
        'status': 'ok',
        'hash': manifesto.hash_consolidado,
        'gerado_em': manifesto.gerado_em.isoformat(),
    })


@require_http_methods(['POST'])
@_requer_login
def api_publicar_ata(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if assembleia.status != 'Concluida':
        return JsonResponse({
            'status': 'error',
            'message': 'Apenas assembleias concluídas podem ter ata.'
        }, status=403)
    data = json.loads(request.body)
    conteudo = data.get('conteudo', '')
    if not conteudo:
        return JsonResponse({'status': 'error', 'message': 'Conteúdo da ata é obrigatório.'}, status=400)

    raw = f'{assembleia.id}-{conteudo}-{timezone.now().isoformat()}'
    assinatura = hashlib.sha256(raw.encode()).hexdigest()

    ata = AtaDigital.objects.create(
        assembleia=assembleia,
        conteudo=conteudo,
        assinatura_hash=assinatura,
        assinado_por=request.usuario_obj,
        assinado_em=timezone.now(),
        assinatura_hash_presidente=assinatura,
        assinado_presidente_em=timezone.now(),
        status_assinatura='Aguardando Secretario',
    )
    _log_assembleia(assembleia.id, request.session['usuario_id'], 'criacao', {
        'ata_id': ata.id, 'status_assinatura': ata.status_assinatura,
    }, ip=_get_client_ip(request))

    _notificar_para_papel('Administrador', 'ata_publicada', f'Ata publicada: {assembleia.titulo}', 'A ata da assembleia foi publicada no repositório.', f'/governanca/atas/')
    _notificar_para_papel('Despachante Oficial', 'ata_publicada', f'Ata publicada: {assembleia.titulo}', 'A ata da assembleia está disponível para consulta.', f'/governanca/atas/')
    return JsonResponse({'status': 'ok', 'ata_id': ata.id, 'assinatura': assinatura, 'status_assinatura': ata.get_status_assinatura_display()})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Documentos
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_http_methods(['POST'])
@_requer_login
def api_upload_documento(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if not usuario_tem_permissao(request, 'gerir_documentos'):
        return JsonResponse({'status': 'error', 'message': 'Sem permissão para gerir documentos.'}, status=403)
    titulo = request.POST.get('titulo', '').strip()
    tipo = request.POST.get('tipo', 'ata')
    if assembleia.status != 'Concluida':
        return JsonResponse({
            'status': 'error',
            'message': 'Apenas assembleias concluídas podem ter documentos.'
        }, status=403)
    if not titulo:
        return JsonResponse({'status': 'error', 'message': 'Título obrigatório.'}, status=400)
    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'status': 'error', 'message': 'Ficheiro obrigatório.'}, status=400)
    MAX_SIZE = 10 * 1024 * 1024
    if arquivo.size > MAX_SIZE:
        return JsonResponse({'status': 'error', 'message': 'O ficheiro excede o tamanho máximo de 10 MB.'}, status=400)
    ALLOWED_TYPES = [
        'application/pdf', 'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'image/jpeg', 'image/png',
    ]
    if arquivo.content_type not in ALLOWED_TYPES:
        return JsonResponse({'status': 'error', 'message': 'Tipo de ficheiro não permitido. Use PDF, DOC, DOCX, JPG ou PNG.'}, status=400)
    doc = DocumentoAssembleia.objects.create(
        assembleia=assembleia,
        tipo=tipo, titulo=titulo,
        descricao=request.POST.get('descricao', ''),
        arquivo=arquivo,
        created_by=request.usuario_obj,
    )
    return JsonResponse({'status': 'ok', 'id': doc.id, 'titulo': doc.titulo, 'tipo': doc.tipo})


@_requer_login
def api_listar_documentos(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    docs = assembleia.documentos.select_related('created_by').all()
    data = [{
        'id': d.id, 'tipo': d.tipo, 'titulo': d.titulo,
        'descricao': d.descricao,
        'arquivo': d.arquivo.url if d.arquivo else '',
        'publicado': d.publicado,
        'publicado_em': d.publicado_em.isoformat() if d.publicado_em else None,
        'created_at': d.created_at.isoformat(),
        'created_by': d.created_by.nome if d.created_by else '',
    } for d in docs]
    return JsonResponse({'documentos': data})


@require_http_methods(['POST'])
@_requer_login
def api_publicar_documento(request, pk, doc_pk):
    from users.permissoes import usuario_tem_permissao
    doc = get_object_or_404(DocumentoAssembleia, pk=doc_pk, assembleia_id=pk)
    if not usuario_tem_permissao(request, 'gerir_documentos'):
        return JsonResponse({'status': 'error', 'message': 'Sem permissão para gerir documentos.'}, status=403)
    doc.publicado = True
    doc.publicado_em = timezone.now()
    doc.save()

    _notificar_para_papel(
        'Administrador', 'ata_publicada',
        f'Documento publicado: {doc.titulo}',
        f'Foi publicado o documento "{doc.titulo}" na assembleia {doc.assembleia.titulo}.',
        f'/governanca/assembleia/{pk}/'
    )
    _notificar_para_papel(
        'Despachante Oficial', 'ata_publicada',
        f'Documento disponível: {doc.titulo}',
        f'O documento "{doc.titulo}" está disponível para consulta.',
        f'/governanca/assembleia/{pk}/'
    )

    for u in Usuario.objects.filter(status='Ativo', papel='Despachante Oficial').exclude(email=''):
        _enviar(
            f'Documento publicado: {doc.titulo}',
            f'Prezado(a) {u.nome},\n\n'
            f'Foi publicado o documento "{doc.titulo}" referente à assembleia "{doc.assembleia.titulo}".\n\n'
            f'Aceda em: {settings.SITE_URL}/governanca/assembleia/{pk}/\n\n'
            f'Atenciosamente,\nCDOA',
            None, [u.email],
        )

    for u in Usuario.objects.filter(status='Ativo', papel='Administrador').exclude(email=''):
        _enviar(
            f'Documento publicado: {doc.titulo}',
            f'Prezado(a) {u.nome},\n\n'
            f'O documento "{doc.titulo}" foi publicado na assembleia "{doc.assembleia.titulo}".\n\n'
            f'Aceda em: {settings.SITE_URL}/governanca/assembleia/{pk}/\n\n'
            f'Atenciosamente,\nCDOA',
            None, [u.email],
        )

    return JsonResponse({'status': 'ok', 'id': doc.id})


@require_http_methods(['POST'])
@_requer_login
def api_remover_documento(request, pk, doc_pk):
    from users.permissoes import usuario_tem_permissao
    doc = get_object_or_404(DocumentoAssembleia, pk=doc_pk, assembleia_id=pk)
    papel = request.session['usuario']['papel']
    usuario_id = request.session['usuario_id']
    from users.permissoes import _is_admin_ou_acesso_total
    is_admin = _is_admin_ou_acesso_total(request)
    is_criador = doc.created_by_id == usuario_id if doc.created_by_id else False
    usuario_obj = Usuario.objects.filter(pk=usuario_id).first()
    is_secretario = MembroMesa.objects.filter(
        assembleia=doc.assembleia,
        usuario_id=usuario_id,
        funcao__in=('1º Secretário', '2º Secretário')
    ).exists() or (usuario_obj and (usuario_obj.is_secretario or usuario_obj.is_vice_secretario))
    if not is_admin and not is_criador and not is_secretario and not usuario_tem_permissao(request, 'gerir_documentos'):
        return JsonResponse({'status': 'error', 'message': 'Sem permissão para remover este documento.'}, status=403)
    doc.delete()
    return JsonResponse({'status': 'ok'})


def _pode_admin_ou_secretario(papel, usuario_obj, request=None):
    from users.permissoes import _is_admin_ou_acesso_total
    if _is_admin_ou_acesso_total(request) if request else papel == 'Administrador':
        return True
    if usuario_obj and (usuario_obj.is_secretario or usuario_obj.is_vice_secretario):
        return True
    if usuario_obj:
        from rh.models import CargoMesa
        return CargoMesa.objects.filter(
            usuario=usuario_obj,
            funcao__in=('1º Secretário', '2º Secretário', 'Secretário', 'Vice-Presidente'),
        ).exists()
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# API: Gerar documento a partir de assembleia
# ═══════════════════════════════════════════════════════════════════════════════

@require_http_methods(['POST'])
@_requer_login
def api_gerar_documento(request):
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'status': 'error', 'message': 'Sem permissão.'}, status=403)

    assembleia_id = request.POST.get('assembleia_id', '').strip()
    tipo = request.POST.get('tipo', 'ata')
    if not assembleia_id:
        return JsonResponse({'status': 'error', 'message': 'Assembleia não especificada.'}, status=400)
    assembleia = get_object_or_404(Assembleia, pk=assembleia_id)
    if tipo not in dict(DocumentoAssembleia.TIPOS):
        return JsonResponse({'status': 'error', 'message': 'Tipo inválido.'}, status=400)

    from .utils import gerar_conteudo_documento
    titulo_map = {'ata': f'Ata — {assembleia.titulo}', 'relatorio': f'Relatório — {assembleia.titulo}', 'decreto': f'Decreto — {assembleia.titulo}'}
    titulo = titulo_map.get(tipo, f'Documento — {assembleia.titulo}')
    conteudo = gerar_conteudo_documento(assembleia, tipo, created_by=usuario_obj)

    doc = DocumentoAssembleia.objects.create(
        assembleia=assembleia,
        tipo=tipo,
        titulo=titulo,
        conteudo=conteudo,
        created_by=usuario_obj,
    )
    return JsonResponse({'status': 'ok', 'id': doc.id, 'titulo': doc.titulo, 'tipo': doc.tipo})


# ═══════════════════════════════════════════════════════════════════════════════
# Visualizar documento gerado (conteúdo HTML)
# ═══════════════════════════════════════════════════════════════════════════════

@_requer_login
def visualizar_documento(request, pk):
    doc = get_object_or_404(DocumentoAssembleia, pk=pk)
    context = {
        'usuario': request.session.get('usuario', {}),
        'nome': request.session.get('usuario', {}).get('nome', ''),
        'papel': request.session.get('usuario', {}).get('papel', ''),
        'active_menu': 'Governanca',
        'doc': doc,
    }
    return render(request, 'governanca/visualizar_documento.html', context)


# ═══════════════════════════════════════════════════════════════════════════════
# Secretário - Gestão de Actas, Decretos e Relatórios
# ═══════════════════════════════════════════════════════════════════════════════

@_requer_login
def secretario_documentos(request):
    usuario = _get_usuario(request)
    if not usuario:
        return redirect('login')
    papel = usuario.get('papel', '')
    usuario_id = usuario.get('id')
    usuario_obj = Usuario.objects.filter(pk=usuario_id).first()
    from users.permissoes import usuario_tem_permissao
    from rh.models import CargoMesa
    if not usuario_obj or not (CargoMesa.objects.filter(usuario=usuario_obj, funcao__in=['Secretário', '1º Secretário', '2º Secretário', 'Vice-Presidente']).exists() or usuario_tem_permissao(request, 'ver_secretaria') or usuario_tem_permissao(request, 'gerir_documentos')):
        return redirect('governanca_index')

    from django.db.models import Count, Q, Prefetch
    todas_qs = DocumentoAssembleia.objects.select_related('assembleia', 'created_by')
    rascunhos = todas_qs.filter(publicado=False).order_by('-created_at')
    publicados = todas_qs.filter(publicado=True).order_by('-publicado_em')
    stats = todas_qs.aggregate(
        total=Count('id'),
        total_rascunhos=Count('id', filter=Q(publicado=False)),
        total_publicados=Count('id', filter=Q(publicado=True)),
        atas=Count('id', filter=Q(tipo='ata')),
        decretos=Count('id', filter=Q(tipo='decreto')),
        relatorios=Count('id', filter=Q(tipo='relatorio')),
    )
    assembleias = Assembleia.objects.prefetch_related(
        Prefetch('documentos', queryset=DocumentoAssembleia.objects.select_related('created_by').order_by('-created_at'), to_attr='todos_docs')
    ).annotate(
        total_docs=Count('documentos'),
        docs_rascunho=Count('documentos', filter=Q(documentos__publicado=False)),
    ).order_by('-data_hora')

    return render(request, 'governanca/secretario_documentos.html', {
        'usuario': usuario,
        'nome': usuario.get('nome', ''),
        'papel': papel,
        'stats': stats,
        'rascunhos': rascunhos[:20],
        'publicados': publicados[:20],
        'assembleias': assembleias,
        'active_menu': 'Governanca',
        'active_sub': 'secretario_docs',
        'TIPOS_DOC': DocumentoAssembleia.TIPOS,
    })


@_requer_login
def api_secretario_assembleias(request):
    usuario = _get_usuario(request)
    if not usuario:
        return JsonResponse({'status': 'error', 'message': 'Sessão expirada.'}, status=401)
    usuario_id = usuario.get('id')
    usuario_obj = Usuario.objects.filter(pk=usuario_id).first()
    if not usuario_obj or not (usuario_obj.has_cargo('secretario') or usuario_obj.has_cargo('vice-secretario')):
        return JsonResponse({'status': 'error', 'message': 'Apenas Secretário e Vice-Secretário.'}, status=403)
    from django.db.models import Prefetch
    assembleias = Assembleia.objects.prefetch_related(
        Prefetch('documentos', queryset=DocumentoAssembleia.objects.select_related('created_by'))
    ).order_by('-data_hora')
    data = []
    for assem in assembleias:
        docs = [{
            'id': d.id, 'tipo': d.tipo, 'titulo': d.titulo,
            'descricao': d.descricao,
            'arquivo': d.arquivo.url if d.arquivo else '',
            'publicado': d.publicado,
            'publicado_em': d.publicado_em.isoformat() if d.publicado_em else None,
            'created_at': d.created_at.isoformat(),
            'created_by': d.created_by.nome if d.created_by else '',
        } for d in assem.documentos.all()]
        data.append({
            'id': assem.id,
            'titulo': assem.titulo,
            'data_hora': assem.data_hora.isoformat(),
            'status': assem.status,
            'local': assem.local,
            'documentos': docs,
        })
    return JsonResponse({'assembleias': data})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Mesa da Assembleia
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_http_methods(['POST'])
@_requer_login
def api_mesa_adicionar(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores.'}, status=403)
    data = json.loads(request.body)
    usuario_id = data.get('usuario_id')
    funcao = data.get('funcao', '').strip()
    if not usuario_id or not funcao:
        return JsonResponse({'status': 'error', 'message': 'Usuário e função obrigatórios.'}, status=400)
    valida = dict(MembroMesa.FUNCOES)
    if funcao not in valida:
        return JsonResponse({'status': 'error', 'message': 'Função inválida.'}, status=400)
    membro, created = MembroMesa.objects.get_or_create(
        assembleia=assembleia, usuario_id=usuario_id,
        defaults={'funcao': funcao, 'ordem': MembroMesa.objects.filter(assembleia=assembleia).count()},
    )
    if not created:
        return JsonResponse({'status': 'error', 'message': 'Usuário já faz parte da mesa.'}, status=400)
    usuario = Usuario.objects.get(id=usuario_id)
    return JsonResponse({
        'status': 'ok',
        'id': membro.id,
        'nome': usuario.nome,
        'funcao': membro.funcao,
    })


@require_http_methods(['POST'])
@_requer_login
def api_mesa_remover(request, pk, membro_pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores.'}, status=403)
    membro = get_object_or_404(MembroMesa, pk=membro_pk, assembleia=assembleia)
    membro.delete()
    return JsonResponse({'status': 'ok'})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - Chat
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@_requer_login
def api_chat_historico(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    msgs = MensagemChat.objects.filter(assembleia=assembleia).select_related('usuario').order_by('-created_at')[:100]
    lista = []
    for m in reversed(msgs):
        item = {
            'id': m.id, 'tipo': m.tipo,
            'nome': m.usuario.nome, 'user_id': m.usuario.id,
            'created_at': m.created_at.isoformat(),
        }
        if m.tipo == 'reacao':
            emojis = {'mao': 'ðŸ–ï¸', 'palmas': 'ðŸ‘', 'coracao': 'â¤ï¸'}
            item['reacao'] = m.reacao
            item['emoji'] = emojis.get(m.reacao, 'â¤ï¸')
        else:
            item['texto'] = m.texto
        lista.append(item)
    return JsonResponse({'mensagens': lista})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API - LiveKit Token & Control
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _livekit_rest_call(method, body=None):
    host = settings.LIVEKIT_URL.replace('wss://', 'https://').replace('ws://', 'http://')
    api_key = settings.LIVEKIT_API_KEY
    api_secret = settings.LIVEKIT_API_SECRET
    header = {'alg': 'HS256', 'typ': 'JWT'}
    now = int(__import__('time').time())
    payload = {
        'iss': api_key,
        'nbf': now - 10,
        'exp': now + 3600,
        'video': {'roomCreate': True, 'roomAdmin': True},
    }
    header_b64 = __import__('base64').urlsafe_b64encode(
        json.dumps(header).encode()).rstrip(b'=').decode()
    payload_b64 = __import__('base64').urlsafe_b64encode(
        json.dumps(payload).encode()).rstrip(b'=').decode()
    sig = __import__('hmac').new(
        api_secret.encode(), f'{header_b64}.{payload_b64}'.encode(),
        __import__('hashlib').sha256).digest()
    sig_b64 = __import__('base64').urlsafe_b64encode(sig).rstrip(b'=').decode()
    token = f'{header_b64}.{payload_b64}.{sig_b64}'

    url = f'{host}/twirp/livekit.RoomService/{method}'
    try:
        import requests
        r = requests.post(url, json=body or {},
                          headers={'Authorization': f'Bearer {token}',
                                   'Content-Type': 'application/json'},
                          timeout=5)
        return r.json() if r.status_code < 300 else {'error': r.text}
    except Exception as e:
        return {'error': str(e)}


@require_http_methods(['POST'])
@_requer_login
def api_livekit_mute(request):
    """Muta ou remove um participante do LiveKit."""
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores.'}, status=403)
    data = json.loads(request.body)
    room = data.get('room', '')
    identity = data.get('identity', '')
    acao = data.get('acao', 'mute')
    if not room or not identity:
        return JsonResponse({'status': 'error', 'message': 'Room e identity obrigatórios.'}, status=400)

    if acao == 'mute':
        resp = _livekit_rest_call('MutePublishedTrack', {
            'room': room, 'identity': identity,
        })
    elif acao == 'unmute':
        resp = _livekit_rest_call('UpdateParticipant', {
            'room': room, 'identity': identity,
            'permission': {'canPublish': True, 'canSubscribe': True},
        })
    elif acao == 'remove':
        resp = _livekit_rest_call('RemoveParticipant', {
            'room': room, 'identity': identity,
        })
    elif acao in ('camera_off', 'camera_on'):
        participants = _livekit_rest_call('ListParticipants', {'room': room})
        track_sid = None
        for p in participants.get('participants', []):
            if p.get('identity') == identity:
                for track in p.get('tracks', []):
                    if track.get('type') == 'video' and track.get('kind') == 'video' or track.get('source') == 'camera':
                        track_sid = track.get('sid')
                        break
                break
        if acao == 'camera_off' and track_sid:
            resp = _livekit_rest_call('MutePublishedTrack', {
                'room': room, 'identity': identity, 'track_sid': track_sid,
            })
        elif acao == 'camera_on':
            resp = _livekit_rest_call('UpdateParticipant', {
                'room': room, 'identity': identity,
                'permission': {'canPublish': True, 'canSubscribe': True},
            })
            if 'error' not in resp:
                resp = {'status': 'ok', 'info': 'Permissão reativada. Participante precisa ligar a câmara manualmente.'}
        else:
            resp = {'error': 'Track de vídeo não encontrado'}
    else:
        return JsonResponse({'status': 'error', 'message': 'Ação inválida.'}, status=400)

    if 'error' in resp:
        return JsonResponse({'status': 'error', 'message': resp['error']}, status=500)
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_livekit_participants(request):
    """Lista participantes ativos numa sala LiveKit."""
    room = request.GET.get('room', '')
    if not room:
        return JsonResponse({'status': 'error', 'message': 'Room obrigatória.'}, status=400)
    resp = _livekit_rest_call('ListParticipants', {'room': room})
    if 'error' in resp:
        return JsonResponse({'status': 'error', 'participants': [], 'message': resp['error']})
    participants = resp.get('participants', [])
    return JsonResponse({
        'status': 'ok',
        'participants': [{
            'identity': p.get('identity', ''),
            'name': p.get('name', ''),
            'isMuted': p.get('isMuted', False),
            'joinedAt': p.get('joinedAt', ''),
            'tracks': [{
                'sid': t.get('sid'),
                'type': t.get('type'),
                'source': t.get('source'),
                'muted': t.get('muted', False),
                'kind': t.get('kind'),
            } for t in p.get('tracks', [])],
            'camera_muted': any(
                t.get('muted', False)
                for t in p.get('tracks', [])
                if t.get('type') in ('video',) or t.get('kind') == 'video'
            ),
            'audio_muted': any(
                t.get('muted', False)
                for t in p.get('tracks', [])
                if t.get('type') in ('audio',) or t.get('kind') == 'audio'
            ),
        } for p in participants],
    })


@_requer_login
def api_livekit_token(request):
    room = request.GET.get('room', '')
    identity = request.session['usuario']['nome']
    token = _livekit_token(room, identity)
    return JsonResponse({
        'token': token,
        'url': settings.LIVEKIT_URL,
        'room': room,
        'identity': identity,
    })


@_requer_login
def api_livekit_refresh_token(request):
    """Refreshes the LiveKit token (extends session)."""
    room = request.GET.get('room', '')
    if not room:
        return JsonResponse({'status': 'error', 'message': 'Room obrigatória.'}, status=400)
    identity = request.session['usuario']['nome']
    token = _livekit_token(room, identity)
    return JsonResponse({'token': token, 'identity': identity})


@require_http_methods(['POST'])
@_requer_login
def api_recording_start(request):
    """Inicia gravação da assembleia via LiveKit Egress."""
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores.'}, status=403)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)
    room = body.get('room', '')
    assembleia_id = body.get('assembleia_id', '')
    if not room:
        return JsonResponse({'status': 'error', 'message': 'Room obrigatória.'}, status=400)
    # Chama o Egress do LiveKit para iniciar gravação
    payload = {
        'room_name': room,
        'output': {
            'file_type': 'mp4',
            'filepath': f'/recordings/{room}_{int(time.time())}.mp4',
        },
    }
    resp = _livekit_rest_call('Egress.StartRoomCompositeEgress', payload)
    if 'error' in resp:
        return JsonResponse({'status': 'error', 'message': resp['error']}, status=500)
    egress_id = resp.get('egress_id', '')
    return JsonResponse({'status': 'ok', 'egress_id': egress_id})


@require_http_methods(['POST'])
@_requer_login
def api_recording_stop(request):
    """Para a gravação da assembleia."""
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'status': 'error', 'message': 'Apenas administradores.'}, status=403)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)
    egress_id = body.get('egress_id', '')
    if not egress_id:
        return JsonResponse({'status': 'error', 'message': 'egress_id obrigatório.'}, status=400)
    resp = _livekit_rest_call('Egress.StopEgress', {'egress_id': egress_id})
    if 'error' in resp:
        return JsonResponse({'status': 'error', 'message': resp['error']}, status=500)
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_mesa_listar(request, pk):
    """Lista membros da mesa de uma assembleia."""
    mesa = MembroMesa.objects.filter(assembleia_id=pk).select_related('usuario')
    return JsonResponse({
        'membros': [{
            'id': m.id,
            'usuario_id': m.usuario.id,
            'usuario_nome': m.usuario.nome,
            'funcao': m.funcao,
        } for m in mesa]
    })


@_requer_login
def api_presencas_listar(request, pk):
    """Lista presenças de uma assembleia."""
    presencas = PresencaAssembleia.objects.filter(
        assembleia_id=pk, presente_em__isnull=False
    ).select_related('usuario')
    return JsonResponse({
        'presentes': [{
            'usuario_id': p.usuario.id,
            'nome': p.usuario.nome,
            'presente_em': p.presente_em.isoformat() if p.presente_em else None,
        } for p in presencas]
    })


@_requer_login
def api_assembleia_dados(request, pk):
    """Retorna dados completos da assembleia para o frontend."""
    a = get_object_or_404(Assembleia, pk=pk)
    pautas = a.pautas.with_vote_counts().order_by('ordem')
    user_id = request.session.get('usuario_id')
    votos_usuario = set(
        Voto.objects.filter(pauta__assembleia=a, usuario_id=user_id)
        .values_list('pauta_id', flat=True)
    )
    primeira_ata = a.atas.first()
    return JsonResponse({
        'assembleia': {
            'id': a.id,
            'titulo': a.titulo,
            'descricao': a.descricao,
            'data_hora': a.data_hora.isoformat(),
            'status': a.status,
            'quorum_minimo': a.quorum_minimo,
            'total_eleitores': a.total_eleitores,
            'presentes_count': a.presentes_count,
            'livekit_room': a.livekit_room,
        },
        'ata': {
            'id': primeira_ata.id,
            'status_assinatura': primeira_ata.status_assinatura,
            'assinado_presidente': bool(primeira_ata.assinatura_hash_presidente),
            'assinado_secretario': bool(primeira_ata.assinatura_hash_secretario),
            'created_at': primeira_ata.created_at.isoformat(),
        } if primeira_ata else None,
        'pautas': [{
            'id': p.id,
            'titulo': p.titulo,
            'descricao': p.descricao,
            'ordem': p.ordem,
            'status': p.status,
            'tipo_votacao': p.tipo_votacao,
            'resultado_final': p.resultado_final,
            'votos_favor': p.votos_favor,
            'votos_contra': p.votos_contra,
            'votos_abstencao': p.votos_abstencao,
            'total_votos': p.total_votos,
            'ja_votou': p.id in votos_usuario,
        } for p in pautas],
    })


# ═══════════════════════════════════════════════════════════════════════════════
# API - Notificações
# ═══════════════════════════════════════════════════════════════════════════════
# Submódulo: Gestão Financeira de Quotas
# ═══════════════════════════════════════════════════════════════════════════════

import datetime as _dt
import uuid as _uuid
from decimal import Decimal as _Decimal
from utils.email_utils import _enviar as _email

def _get_estado_financeiro(despachante_id):
    ef, _ = EstadoFinanceiro.objects.get_or_create(despachante_id=despachante_id, defaults={'estado': 'Regular'})
    return ef

def _registrar_historico(membro, quota, pagamento, acao, descricao, request=None, utilizador_id=None):
    """Regista um evento no histórico de quotas."""
    HistoricoQuota.objects.create(
        membro=membro,
        quota=quota,
        pagamento=pagamento,
        acao=acao,
        descricao=descricao,
        utilizador_id=utilizador_id or (request.session.get('usuario_id') if request else None),
        ip=request.META.get('REMOTE_ADDR') if request else None,
    )


def _calcular_multa_quota(quota, config_override=None):
    """Calcula multa por atraso para uma QuotaGerada.
    Usa dias_carencia da config. Retorna dict com dias_atraso, multa_valor, total_sugerido.
    """
    if not quota:
        return {'dias_atraso': 0, 'multa_valor': 0, 'total_sugerido': 0}
    ano = quota.ano
    mes = quota.mes
    config = None
    if ano and mes:
        config = config_override or QuotaConfig.objects.filter(ano=ano, mes=mes).first()
    vencimento = config.data_vencimento if config else quota.data_vencimento
    if not config or not config.multa_percentual or not vencimento:
        valor_base = quota.valor_original or quota.valor
        return {'dias_atraso': 0, 'multa_valor': 0, 'total_sugerido': valor_base}
    hoje = timezone.now().date()
    dias_desde_venc = (hoje - vencimento).days
    carencia = config.dias_carencia or 0
    if dias_desde_venc <= carencia:
        valor_base = quota.valor_original or quota.valor
        return {'dias_atraso': 0, 'multa_valor': 0, 'total_sugerido': valor_base}
    dias_atraso = dias_desde_venc - carencia
    valor_original = quota.valor_original or quota.valor
    multa_valor = valor_original * (config.multa_percentual / _Decimal(100)) * dias_atraso
    multa_valor = min(multa_valor, valor_original * 12)
    total = valor_original + multa_valor
    return {'dias_atraso': dias_atraso, 'multa_valor': multa_valor, 'total_sugerido': total}


def _calcular_multa(pagamento, config_override=None):
    """Calcula multa por atraso para um PagamentoQuota. (deprecated — usa _calcular_multa_quota)"""
    if not pagamento or not pagamento.quota:
        return {'dias_atraso': 0, 'multa_valor': 0, 'total_sugerido': 0}
    return _calcular_multa_quota(pagamento.quota, config_override)

def _atualizar_estado_financeiro(despachante_id, registar_historico=False, request=None):
    ef = _get_estado_financeiro(despachante_id)
    if ef.estado == 'Suspenso':
        return ef
    pendentes = QuotaGerada.objects.filter(despachante_id=despachante_id, status__in=['Pendente','Atrasada','Pendente Confirmacao']).count()
    novo_estado = 'Regular' if pendentes == 0 else 'Irregular'
    if novo_estado != ef.estado:
        estado_anterior = ef.estado
        ef.estado = novo_estado
        ef.save(update_fields=['estado', 'ultima_atualizacao'])
        if registar_historico:
            from users.models import Usuario
            membro = Usuario.objects.filter(id=despachante_id).first()
            if membro:
                _registrar_historico(
                    membro=membro, quota=None, pagamento=None,
                    acao='ESTADO_FINANCEIRO_ALTERADO',
                    descricao=f'{estado_anterior} → {novo_estado}',
                    request=request,
                )
    return ef


# ─── Páginas HTML ───────────────────────────────────────────────────────────

@_requer_login
def quotas_dashboard(request):
    from users.permissoes import usuario_tem_permissao
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    papel = request.session['usuario']['papel']
    if papel not in ('Administrador', 'Despachante Oficial') and not (usuario_tem_permissao(request, 'gerir_quotas') or usuario_tem_permissao(request, 'ver_quotas')):
        return redirect('governanca_index')
    ef = _get_estado_financeiro(usuario_id)
    quotas_pendentes = QuotaGerada.objects.filter(despachante_id=usuario_id, status__in=['Pendente','Atrasada','Pendente Confirmacao']).count()
    total_quotas = QuotaGerada.objects.filter(despachante_id=usuario_id).count()
    ultimas_quotas_qs = QuotaGerada.objects.filter(despachante_id=usuario_id).order_by('-ano','-mes')
    paginator = Paginator(ultimas_quotas_qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    carteira = CarteiraProfissional.objects.filter(despachante_id=usuario_id).first()
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': papel, 'active_menu': 'Governanca', 'active_sub': 'quotas',
        'estado_financeiro': ef, 'quotas_pendentes': quotas_pendentes,
        'total_quotas': total_quotas, 'ultimas_quotas': page_obj, 'page_obj': page_obj, 'carteira': carteira,
    }
    return render(request, 'governanca/quotas/dashboard.html', context)


@_requer_login
def quotas_faturas(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    papel = request.session['usuario']['papel']
    from users.permissoes import _is_admin_ou_acesso_total
    if papel == 'Administrador':
        quotas = QuotaGerada.objects.all().select_related('despachante').order_by('-ano','-mes')
    else:
        quotas = QuotaGerada.objects.filter(despachante_id=usuario_id).order_by('-ano','-mes')
    paginator = Paginator(quotas, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': papel, 'active_menu': 'Governanca', 'active_sub': 'quotas',
        'quotas': page_obj, 'page_obj': page_obj, 'is_admin': papel == 'Administrador',
    }
    return render(request, 'governanca/quotas/faturas.html', context)


@_requer_login
def quotas_fatura_detalhe(request, fatura_uuid):
    from users.permissoes import _is_admin_ou_acesso_total
    quota = get_object_or_404(QuotaGerada, fatura_uuid=fatura_uuid)
    uid = request.session.get('banca_usuario_id') or request.session['usuario_id']
    papel = request.session.get('usuario', {}).get('papel', '')
    if quota.despachante_id != uid and papel != 'Administrador':
        return redirect('governanca_quotas_dashboard')
    pagamentos_qs = PagamentoQuota.objects.filter(quota=quota).order_by('-data_pagamento')
    paginator = Paginator(pagamentos_qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    config = QuotaConfig.objects.filter(ano=quota.ano, mes=quota.mes).first() if quota.ano and quota.mes else None
    multa_info = _calcular_multa_quota(quota, config)
    historico = HistoricoQuota.objects.filter(quota=quota).select_related('utilizador').order_by('-created_at')[:20]
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'], 'active_menu': 'Governanca', 'active_sub': 'quotas',
        'quota': quota, 'pagamentos': page_obj, 'page_obj': page_obj,
        'multa_info': multa_info, 'config_multa': config.multa_percentual if config else 0,
        'config_dias_carencia': config.dias_carencia if config else 5,
        'historico': historico,
    }
    return render(request, 'governanca/quotas/quota_detalhe.html', context)


@_requer_login
def quotas_certidao(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    ef = _get_estado_financeiro(usuario_id)
    certidoes_qs = CertidaoRegularidade.objects.filter(despachante_id=usuario_id).order_by('-data_emissao')
    paginator = Paginator(certidoes_qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'], 'active_menu': 'Governanca', 'active_sub': 'quotas',
        'estado_financeiro': ef, 'certidoes': page_obj, 'page_obj': page_obj,
        'pode_emitir': ef.estado == 'Regular',
    }
    return render(request, 'governanca/quotas/certidao.html', context)


@_requer_login
def quotas_carteira(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    carteira = CarteiraProfissional.objects.filter(despachante_id=usuario_id).first()
    ef = _get_estado_financeiro(usuario_id)
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'], 'active_menu': 'Governanca', 'active_sub': 'quotas',
        'carteira': carteira, 'estado_financeiro': ef,
    }
    return render(request, 'governanca/quotas/carteira.html', context)


# ─── Admin ──────────────────────────────────────────────────────────────────

@_requer_login
def quotas_admin_dashboard(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return redirect('governanca_quotas_dashboard')

    from django.db.models import Count, Q
    stats = QuotaGerada.objects.aggregate(
        total=Count('id'),
        pendentes=Count('id', filter=Q(status__in=['Pendente','Atrasada','Pendente Confirmacao'])),
        pagas=Count('id', filter=Q(status='Paga')),
    )
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': 'Administrador', 'active_menu': 'Governanca', 'active_sub': 'quotas_admin',
        'total_quotas': stats['total'],
        'pendentes': stats['pendentes'],
        'pagas': stats['pagas'],
        'pagamentos_pendentes': PagamentoQuota.objects.filter(status='Pendente Confirmacao').count(),
        'config': QuotaConfig.objects.order_by('-ano','-mes').first(),
        'categorias': CategoriaMembro.objects.all(),
        'tipos_quota': TipoQuota.objects.all(),
        'membros_isencao': Usuario.objects.filter(papel='Despachante Oficial', status='Ativo').order_by('nome'),
    }
    return render(request, 'governanca/quotas/admin_dashboard.html', context)


@_requer_login
def quotas_admin_pagamentos(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return redirect('governanca_quotas_dashboard')
    pagamentos_qs = PagamentoQuota.objects.all().select_related('despachante','quota').order_by('-data_pagamento')
    paginator = Paginator(pagamentos_qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    configs = {(c.ano, c.mes): c for c in QuotaConfig.objects.all()}
    pagamentos_com_multa = []
    for p in page_obj:
        config = configs.get((p.quota.ano, p.quota.mes))
        multa_info = _calcular_multa(p, config_override=config)
        pagamentos_com_multa.append({
            'pagamento': p,
            'dias_atraso': multa_info['dias_atraso'],
            'multa_valor': multa_info['multa_valor'],
            'total_sugerido': multa_info['total_sugerido'],
        })
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': 'Administrador', 'active_menu': 'Governanca', 'active_sub': 'quotas_admin',
        'pagamentos': page_obj, 'page_obj': page_obj,
        'pagamentos_com_multa': pagamentos_com_multa,
    }
    return render(request, 'governanca/quotas/admin_pagamentos.html', context)


@_requer_login
def quotas_admin_config(request):
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return redirect('governanca_quotas_dashboard')
    configs_qs = QuotaConfig.objects.order_by('-ano','-mes')
    paginator = Paginator(configs_qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': 'Administrador', 'active_menu': 'Governanca', 'active_sub': 'quotas_admin',
        'configs': page_obj, 'page_obj': page_obj,
        'categorias': CategoriaMembro.objects.filter(usuario__papel='Despachante Oficial', usuario__status='Ativo').distinct(),
        'tipos_quota': TipoQuota.objects.all(),
    }
    return render(request, 'governanca/quotas/admin_config.html', context)


@_requer_login
def quotas_admin_relatorios(request):
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return redirect('governanca_quotas_dashboard')

    from django.db.models import Sum, Count, Q
    total_arrecadado = QuotaGerada.objects.filter(
        status='Paga'
    ).aggregate(total=Sum('valor'))['total'] or 0
    stats = QuotaGerada.objects.aggregate(
        atrasadas=Count('id', filter=Q(status='Atrasada')),
        pendentes=Count('id', filter=Q(status='Pendente')),
    )
    inadimplentes = list(
        Usuario.objects.filter(papel='Despachante Oficial', status='Ativo')
        .annotate(
            quotas_pend=Count('quotas', filter=Q(quotas__status__in=['Pendente', 'Atrasada'])),
            total_devido=Sum('quotas__valor', filter=Q(quotas__status__in=['Pendente', 'Atrasada']), distinct=True),
        )
        .filter(quotas_pend__gt=0)
        .order_by('-total_devido')
    )[:20]
    historico = list(
        PagamentoQuota.objects.filter(status='Confirmado').select_related(
            'quota', 'despachante', 'confirmado_por'
        ).order_by('-confirmado_em')[:50]
    )
    resumo_mensal = list(
        QuotaGerada.objects.values('ano', 'mes')
        .annotate(
            total=Count('id'),
            pagas=Count('id', filter=Q(status='Paga')),
            pendentes=Count('id', filter=Q(status='Pendente')),
            atrasadas=Count('id', filter=Q(status='Atrasada')),
            pend_conf=Count('id', filter=Q(status='Pendente Confirmacao')),
            isentas=Count('id', filter=Q(status='Isenta')),
            canceladas=Count('id', filter=Q(status='Cancelada')),
            valor_total=Sum('valor'),
            arrecadado=Sum('valor', filter=Q(status='Paga')),
        )
        .order_by('-ano', '-mes')[:12]
    )
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': 'Administrador', 'active_menu': 'Governanca', 'active_sub': 'quotas_admin',
        'total_arrecadado': total_arrecadado,
        'quotas_em_atraso': stats['atrasadas'],
        'quotas_pendentes': stats['pendentes'],
        'inadimplentes': inadimplentes,
        'historico': historico,
        'resumo_mensal': resumo_mensal,
    }
    return render(request, 'governanca/quotas/admin_relatorios.html', context)


@_requer_login
def quotas_admin_gerar_retroativo(request):
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return redirect('governanca_quotas_dashboard')
    hoje = timezone.now()
    config_qs = QuotaConfig.objects.filter(ativa=True).select_related('tipo', 'categoria').order_by('-ano', '-mes')
    configs_json = []
    for c in config_qs:
        configs_json.append({
            'id': c.id, 'tipo_id': c.tipo_id, 'tipo_nome': c.tipo.nome if c.tipo else '',
            'ano': c.ano, 'mes': c.mes,
            'valor': float(str(c.valor)) if c.valor else 0,
            'data_vencimento': str(c.data_vencimento) if c.data_vencimento else '',
            'multa_percentual': float(str(c.multa_percentual)) if c.multa_percentual else None,
            'dias_carencia': c.dias_carencia,
            'ativa': c.ativa,
        })
    context = {
        'usuario': request.session['usuario'], 'nome': request.session['usuario']['nome'],
        'papel': 'Administrador', 'active_menu': 'Governanca', 'active_sub': 'quotas_admin',
        'tipos_quota': TipoQuota.objects.all(),
        'ano_actual': hoje.year,
        'mes_actual': hoje.month,
        'configs_json': json.dumps(configs_json),
    }
    return render(request, 'governanca/quotas/admin_gerar_retroativo.html', context)


# ─── APIs ───────────────────────────────────────────────────────────────────

@require_http_methods(['POST'])
@_requer_login
def api_quotas_pagar(request, fatura_uuid):
    quota = get_object_or_404(QuotaGerada, fatura_uuid=fatura_uuid)
    pode_pagar = (
        quota.despachante_id == request.session['usuario_id']
        or (
            request.session.get('tipo_usuario') == 'colaborador'
            and request.session.get('banca_usuario_id') == quota.despachante_id
        )
    )
    if not pode_pagar:
        return JsonResponse({'erro': 'Esta quota não lhe pertence'}, status=403)
    if quota.status in ('Paga', 'Cancelada', 'Isenta'):
        return JsonResponse({'erro': 'Quota não pode ser paga'}, status=400)
    if quota.status == 'Pendente Confirmacao':
        return JsonResponse({'erro': 'Já existe um pagamento pendente de confirmação para esta quota'}, status=400)
    metodo = request.POST.get('metodo', '')
    metodos_validos = [m[0] for m in PagamentoQuota.METODOS]
    if metodo not in metodos_validos:
        return JsonResponse({'erro': 'Método de pagamento inválido'}, status=400)
    comprovativo = request.FILES.get('comprovativo')
    observacao = request.POST.get('observacao', '')
    multa_info = _calcular_multa_quota(quota)
    valor_pago = multa_info['total_sugerido']
    pag = PagamentoQuota(
        quota=quota, despachante_id=quota.despachante_id,
        metodo=metodo, valor_pago=valor_pago,
        status_anterior_quota=quota.status,
    )
    if comprovativo:
        pag.comprovativo = comprovativo
    if observacao:
        pag.observacoes = observacao
    pag.save()
    quota.status = 'Pendente Confirmacao'
    quota.save(update_fields=['status'])
    _registrar_historico(
        membro=quota.despachante, quota=quota, pagamento=pag,
        acao='PAGAMENTO_SUBMETIDO',
        descricao=f'Pagamento submetido via {metodo} — Kz {fmt_kz(valor_pago)}',
        request=request,
    )
    return JsonResponse({'status': 'ok', 'pagamento_id': pag.id, 'mensagem': 'Pagamento submetido com sucesso. Aguarde confirmação.'})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_confirmar_pagamento(request, pk):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    pag = get_object_or_404(PagamentoQuota, pk=pk)
    acao = request.POST.get('acao', 'confirmar')
    if acao == 'confirmar':
        multa_info = _calcular_multa(pag)
        pag.status = 'Confirmado'
        pag.confirmado_por_id = request.session['usuario_id']
        pag.confirmado_em = timezone.now()
        pag.save(update_fields=['status', 'confirmado_por_id', 'confirmado_em'])
        pag.quota.status = 'Paga'
        pag.quota.data_pagamento = timezone.now()
        pag.quota.valor_multa = multa_info['multa_valor']
        pag.quota.valor_total = multa_info['total_sugerido']
        pag.quota.save(update_fields=['status', 'data_pagamento', 'valor_multa', 'valor_total'])
        _registrar_historico(
            membro=pag.despachante, quota=pag.quota, pagamento=pag,
            acao='PAGAMENTO_APROVADO',
            descricao=f'Pagamento aprovado por admin. Multa: Kz {fmt_kz(multa_info["multa_valor"])}.',
            request=request,
        )
        _atualizar_estado_financeiro(pag.despachante_id, registar_historico=True, request=request)
        multa_msg = ''
        if multa_info['dias_atraso'] > 0:
            multa_msg = f' ({multa_info["dias_atraso"]} dias de atraso, multa de Kz {fmt_kz(multa_info["multa_valor"])})'
        Notificacao.objects.create(
            usuario=pag.despachante, tipo='pagamento_confirmado',
            titulo='Pagamento Confirmado',
            mensagem=f'O pagamento da quota {pag.quota.mes:02d}/{pag.quota.ano} foi confirmado.{multa_msg}',
            link='/governanca/quotas/',
        )
        if pag.despachante.email:
            _email('Pagamento Confirmado',
                f'Olá {pag.despachante.nome},\n\nO pagamento da sua quota {pag.quota.mes:02d}/{pag.quota.ano} foi confirmado.{multa_msg}\n\nCDOA Angola', None, [pag.despachante.email])
    elif acao == 'rejeitar':
        motivo = request.POST.get('motivo_rejeicao', '').strip()
        if not motivo:
            return JsonResponse({'erro': 'Motivo da rejeição é obrigatório'}, status=400)
        pag.status = 'Rejeitado'
        pag.confirmado_por_id = request.session['usuario_id']
        pag.confirmado_em = timezone.now()
        if motivo:
            pag.observacoes = (pag.observacoes + '\n' + motivo).strip()
        pag.save(update_fields=['status', 'confirmado_por_id', 'confirmado_em', 'observacoes'])
        quota = pag.quota
        estado_anterior = pag.status_anterior_quota or 'Pendente'
        quota.status = estado_anterior
        quota.save(update_fields=['status'])
        _atualizar_estado_financeiro(pag.despachante_id)
        _registrar_historico(
            membro=pag.despachante, quota=quota, pagamento=pag,
            acao='PAGAMENTO_REJEITADO',
            descricao=f'Pagamento rejeitado. Motivo: {motivo}. Quota revertida para: {estado_anterior}',
            request=request,
        )
        Notificacao.objects.create(
            usuario=pag.despachante, tipo='pagamento_rejeitado',
            titulo='Pagamento Rejeitado',
            mensagem=f'O pagamento da quota {pag.quota.mes:02d}/{pag.quota.ano} foi rejeitado. Motivo: {motivo}.',
            link='/governanca/quotas/',
        )
        if pag.despachante.email:
            _email('Pagamento Rejeitado',
                f'Olá {pag.despachante.nome},\n\nO pagamento da sua quota {pag.quota.mes:02d}/{pag.quota.ano} foi rejeitado.\nMotivo: {motivo}\n\nPor favor, envie um novo comprovativo.\n\nCDOA Angola', None, [pag.despachante.email])
    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_emitir_certidao(request):
    despachante_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    ef = _get_estado_financeiro(despachante_id)
    if ef.estado != 'Regular':
        return JsonResponse({'erro': 'Estado financeiro irregular. Regularize as suas quotas primeiro.'}, status=400)
    despachante = Usuario.objects.filter(id=despachante_id).first()
    if not despachante:
        return JsonResponse({'erro': 'Utilizador não encontrado'}, status=404)
    from utils.pdf_quotas import gerar_certidao_pdf
    result = gerar_certidao_pdf(despachante, request.session['usuario']['nome'])
    validade = timezone.now().date() + _dt.timedelta(days=90)
    cert = CertidaoRegularidade.objects.create(
        despachante_id=despachante_id, codigo_certidao=result['codigo'],
        data_validade=validade, arquivo_pdf=result['pdf_path'],
        assinatura_hash=result['hash'], emitido_por_id=request.session['usuario_id'],
    )
    Notificacao.objects.create(
        usuario_id=despachante_id, tipo='certidao_emitida',
        titulo='Certidão de Regularidade Emitida',
        mensagem='A sua certidão de regularidade foi emitida com sucesso.',
        link='/governanca/quotas/certidao/',
    )
    return JsonResponse({'status': 'ok', 'codigo': result['codigo'], 'url': result['pdf_url']})


@_requer_login
def api_quotas_definir_estado(request, pk):
    """Admin define estado financeiro de um despachante."""
    from users.permissoes import usuario_tem_permissao
    papel = request.session['usuario']['papel']
    if papel != 'Administrador' and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'erro': 'Método não permitido'}, status=405)
    if papel != 'Administrador':
        uid = request.session.get('banca_usuario_id') or request.session.get('usuario_id')
        if int(pk) != uid:
            return JsonResponse({'erro': 'Sem permissão para alterar estado de outro despachante'}, status=403)
    estado = request.POST.get('estado', '').strip()
    if estado not in dict(EstadoFinanceiro.ESTADOS):
        return JsonResponse({'erro': 'Estado inválido'}, status=400)
    ef, _ = EstadoFinanceiro.objects.get_or_create(despachante_id=pk, defaults={'estado': estado})
    ef.estado = estado
    ef.observacoes = request.POST.get('observacoes', '')
    ef.save(update_fields=['estado', 'observacoes', 'ultima_atualizacao'])
    return JsonResponse({'status': 'ok', 'estado': ef.estado})


@_requer_login
def api_quotas_buscar_membros(request):
    """Busca membros (despachantes) para o admin definir estado financeiro."""
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    q = request.GET.get('q', '').strip()
    membros = Usuario.objects.filter(papel='Despachante Oficial', status='Ativo')
    if q:
        membros = membros.filter(
            Q(nome__icontains=q) | Q(email__icontains=q) | Q(nif__icontains=q)
        )
    membros = membros.order_by('nome')[:20]
    data = []
    for m in membros:
        ef = EstadoFinanceiro.objects.filter(despachante_id=m.id).first()
        data.append({
            'id': m.id,
            'nome': m.nome,
            'cedula': '',
            'estado_financeiro': ef.estado if ef else 'Regular',
        })
    return JsonResponse({'membros': data})


@_requer_login
def api_quotas_verificar_estado(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    ef = _get_estado_financeiro(usuario_id)
    pendentes = QuotaGerada.objects.filter(despachante_id=usuario_id, status__in=['Pendente','Atrasada','Pendente Confirmacao']).count()
    return JsonResponse({
        'estado': ef.estado, 'quotas_pendentes': pendentes,
        'pode_votar': ef.estado == 'Regular',
        'pode_emitir_certidao': ef.estado == 'Regular',
    })


@_requer_login
def api_quotas_listar(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    papel = request.session['usuario']['papel']
    ano = request.GET.get('ano')
    from users.permissoes import _is_admin_ou_acesso_total
    e_admin = papel == 'Administrador'
    if e_admin:
        quotas = QuotaGerada.objects.select_related('despachante').order_by('-ano','-mes')
        if ano:
            quotas = quotas.filter(ano=ano)
        else:
            from datetime import date
            quotas = quotas.filter(ano=date.today().year)
    else:
        quotas = QuotaGerada.objects.filter(despachante_id=usuario_id).order_by('-ano','-mes')
    data = []
    for q in quotas:
        data.append({
            'id': q.id, 'fatura_uuid': q.fatura_uuid, 'ano': q.ano, 'mes': q.mes,
            'valor': str(q.valor), 'data_vencimento': str(q.data_vencimento),
            'status': q.status, 'despachante_nome': q.despachante.nome if e_admin else None,
        })
    return JsonResponse({'quotas': data})


@_requer_login
def api_quotas_dashboard(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    papel = request.session['usuario']['papel']
    ef = _get_estado_financeiro(usuario_id)
    pendentes = QuotaGerada.objects.filter(despachante_id=usuario_id, status__in=['Pendente','Atrasada','Pendente Confirmacao']).count()
    pagas = QuotaGerada.objects.filter(despachante_id=usuario_id, status='Paga').count()
    return JsonResponse({
        'estado': ef.estado, 'quotas_pendentes': pendentes, 'quotas_pagas': pagas,
    })


# ─── Carteira Profissional ──────────────────────────────────────────────────

@require_http_methods(['POST'])
@_requer_login
def api_quotas_renovar_carteira(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    ef = _get_estado_financeiro(usuario_id)
    if ef.estado != 'Regular':
        return JsonResponse({'erro': 'Estado financeiro irregular. Regularize as suas quotas primeiro.'}, status=400)
    carteira = CarteiraProfissional.objects.filter(despachante_id=usuario_id).first()
    if carteira:
        despachante = carteira.despachante
    else:
        despachante = Usuario.objects.filter(id=usuario_id).first()
    if not despachante:
        return JsonResponse({'erro': 'Utilizador não encontrado'}, status=404)
    if not carteira:
        from datetime import date as _ddate
        hoje = _ddate.today()
        validade = _ddate(hoje.year + 2, hoje.month, hoje.day)
        from uuid import uuid4 as _uuid4
        numero = f'CDOA-{despachante.cedula or despachante.id}-{hoje.year}'
        carteira = CarteiraProfissional.objects.create(
            despachante=despachante, numero_carteira=numero,
            data_emissao=hoje, data_validade=validade,
        )
    carteira.data_renovacao = timezone.now().date()
    carteira.status = 'Activa'
    carteira.save()
    from utils.pdf_quotas import gerar_carteira_pdf
    pdf_path, pdf_url = gerar_carteira_pdf(despachante, carteira, request.session['usuario']['nome'])
    carteira.arquivo_pdf = pdf_path
    carteira.save(update_fields=['arquivo_pdf'])
    return JsonResponse({'status': 'ok', 'numero_carteira': carteira.numero_carteira, 'validade': str(carteira.data_validade), 'pdf_url': pdf_url})


@_requer_login
def api_quotas_carteira(request):
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    carteira = CarteiraProfissional.objects.filter(despachante_id=usuario_id).first()
    if not carteira:
        return JsonResponse({'carteira': None})
    return JsonResponse({
        'carteira': {
            'numero': carteira.numero_carteira, 'data_emissao': str(carteira.data_emissao),
            'data_validade': str(carteira.data_validade), 'status': carteira.status,
        }
    })


@_requer_login
def api_quotas_listar_despachantes(request):
    """Lista todos os Despachante Oficial ativos para seleção pré-publicação."""
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    despachantes = Usuario.objects.filter(papel='Despachante Oficial', status='Ativo').order_by('nome')
    data = []
    for d in despachantes:
        data.append({
            'id': d.id,
            'nome': d.nome,
            'email': d.email or '',
            'nif': d.nif or '',
            'categoria': d.categoria.nome if d.categoria else '',
        })
    return JsonResponse({'despachantes': data, 'total': len(data)})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_salvar_config(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    try:
        ano = int(request.POST.get('ano', 0))
        mes_s = request.POST.get('mes', '')
        mes = int(mes_s) if mes_s else None
        valor = _Decimal(parse_kz(request.POST.get('valor', '0')) or '0')
        vencimento = request.POST.get('data_vencimento', '')
        multa_percentual = _Decimal(parse_kz(request.POST.get('multa_percentual', '0.50')) or '0.50')
        dias_carencia = int(request.POST.get('dias_carencia', 5))
        ativa = request.POST.get('ativa', '0') == '1'
        publicar = request.POST.get('publicar', '0') == '1'
        categoria_id = request.POST.get('categoria_id', '') or None
        tipo_id = request.POST.get('tipo_id', '') or None
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Dados inválidos'}, status=400)
    if ano < 2000 or ano > 2100 or valor <= 0:
        return JsonResponse({'erro': 'Ano ou valor inválidos'}, status=400)
    if mes and (mes < 1 or mes > 12):
        return JsonResponse({'erro': 'Mês inválido'}, status=400)
    defaults = {
        'valor': valor,
        'multa_percentual': multa_percentual,
        'dias_carencia': max(dias_carencia, 0),
        'ativa': ativa,
    }
    if categoria_id:
        defaults['categoria_id'] = int(categoria_id)
    if tipo_id:
        defaults['tipo_id'] = int(tipo_id)
    if vencimento:
        try:
            data_venc = _dt.datetime.strptime(vencimento, '%Y-%m-%d').date()
            if data_venc <= timezone.now().date():
                return JsonResponse({
                    'erro': 'A data de vencimento deve ser posterior ao dia de hoje.'
                }, status=400)
            defaults['data_vencimento'] = data_venc
        except (ValueError, TypeError):
            return JsonResponse({'erro': 'Data de vencimento inválida (use AAAA-MM-DD)'}, status=400)
    elif not QuotaConfig.objects.filter(ano=ano, mes=mes).exists():
        return JsonResponse({'erro': 'Data de vencimento é obrigatória na primeira configuração'}, status=400)

    lookup = {'ano': ano, 'mes': mes}
    if tipo_id:
        lookup['tipo_id'] = int(tipo_id)
    else:
        lookup['tipo_id'] = None
    config, created = QuotaConfig.objects.update_or_create(**lookup, defaults=defaults)

    if publicar and ativa:
        despachantes = Usuario.objects.filter(papel='Despachante Oficial', status='Ativo').order_by('nome')
        tipo = TipoQuota.objects.filter(id=tipo_id).first() if tipo_id else None
        if not tipo:
            tipo = TipoQuota.objects.filter(slug='mensal').first()
        lista = []
        for d in despachantes:
            lista.append({
                'id': d.id,
                'nome': d.nome,
                'email': d.email or '',
                'nif': d.nif or '',
                'categoria': d.categoria.nome if d.categoria else '',
            })
        return JsonResponse({
            'status': 'listar_despachantes',
            'despachantes': lista,
            'config_id': config.id,
            'total': len(lista),
        })
    label = f'{mes:02d}/{ano}' if mes else f'{ano}'
    msg = f'Configuração de {label} salva: Kz {fmt_kz(valor)}'
    messages.success(request, msg)
    return JsonResponse({'status': 'ok', 'mensagem': msg})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_publicar(request):
    """
    Gera quotas para todos os Despachante Oficial ativos,
    exceto os listados em 'excluidos[]'.
    """
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    try:
        config_id = int(request.POST.get('config_id', 0))
        ano = int(request.POST.get('ano', 0))
        mes_s = request.POST.get('mes', '')
        mes = int(mes_s) if mes_s else None
        tipo_id = request.POST.get('tipo_id', '') or None
        valor = _Decimal(parse_kz(request.POST.get('valor', '0')) or '0')
        vencimento = request.POST.get('data_vencimento', '')
        multa_percentual = _Decimal(parse_kz(request.POST.get('multa_percentual', '0.50')) or '0.50')
        dias_carencia = int(request.POST.get('dias_carencia', 5))
        excluidos_str = request.POST.getlist('excluidos[]')
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Dados inválidos'}, status=400)

    excluidos = set()
    for e in excluidos_str:
        try:
            excluidos.add(int(e))
        except ValueError:
            pass

    config = QuotaConfig.objects.filter(id=config_id, ano=ano).first()
    if not config:
        return JsonResponse({'erro': 'Configuração não encontrada'}, status=404)

    if mes is not None:
        config.mes = mes
        config.save(update_fields=['mes'])

    tipo = TipoQuota.objects.filter(id=tipo_id).first() if tipo_id else TipoQuota.objects.filter(slug='mensal').first()
    if not tipo:
        return JsonResponse({'erro': 'Tipo de quota não encontrado'}, status=404)

    slug_tipo = tipo.slug.upper()
    descricao = f'{tipo.nome} {mes:02d}/{ano}' if mes else f'{tipo.nome} {ano}'

    despachantes = Usuario.objects.filter(papel='Despachante Oficial', status='Ativo').order_by('nome')
    geradas = 0
    hoje = timezone.now().date()
    seq = QuotaGerada.objects.filter(ano=ano, mes=mes, tipo=tipo).count() + 1

    for d in despachantes:
        if d.id in excluidos:
            continue
        if tipo.recorrente and mes:
            if QuotaGerada.objects.filter(despachante=d, tipo=tipo, ano=ano, mes=mes).exists():
                continue
        elif not tipo.recorrente:
            if QuotaGerada.objects.filter(despachante=d, tipo=tipo, ano=ano).exists():
                continue
        pi = _dt.date(ano, mes, 1) if mes else None
        pf = None
        if pi and tipo.dias_intervalo:
            from dateutil.relativedelta import relativedelta
            pf = pi + relativedelta(months=(tipo.dias_intervalo // 30)) - _dt.timedelta(days=1)
        referencia = f'QUOTA-{slug_tipo}-{mes:02d}-{ano}-{seq:05d}' if mes else f'QUOTA-{slug_tipo}-{ano}-{seq:05d}'
        seq += 1
        for _attempt in range(5):
            try:
                q = QuotaGerada.objects.create(
                    despachante=d, tipo=tipo, ano=ano, mes=mes,
                    periodo_inicio=pi, periodo_fim=pf,
                    descricao=descricao,
                    valor=config.valor, valor_original=config.valor, valor_total=config.valor,
                    data_vencimento=config.data_vencimento,
                    data_envio=hoje,
                    referencia=referencia,
                )
                break
            except (IntegrityError, ValidationError):
                referencia = f'QUOTA-{slug_tipo}-{mes:02d}-{ano}-{seq:05d}' if mes else f'QUOTA-{slug_tipo}-{ano}-{seq:05d}'
                seq += 1
                continue
        _registrar_historico(
            membro=d, quota=q, pagamento=None,
            acao='QUOTA_GERADA',
            descricao=f'Quota gerada via publicação. Config: {descricao}. Referência: {referencia}',
            request=request,
        )
        multa_str = f' Multa de {config.multa_percentual}%/dia após o vencimento.' if config.multa_percentual else ''
        carencia_str = f' Período de carência: {config.dias_carencia} dias.' if config.dias_carencia else ''
        Notificacao.objects.create(
            usuario=d, tipo='quota_gerada',
            titulo=descricao + ' — Pagamento Disponível',
            mensagem=f'Foi publicada a sua {descricao} no valor de Kz {config.valor}. Vencimento: {config.data_vencimento}.{multa_str}{carencia_str}',
            link='/governanca/quotas/',
        )
        if d.email:
            texto_multa = f"Multa de {config.multa_percentual}%/dia após o vencimento.\n" if config.multa_percentual else ""
            _enviar(
                'Quota Associativa — Pagamento Disponível',
                f'Olá {d.nome},\n\nA sua {descricao} no valor de Kz {config.valor} foi publicada.\n'
                f'Data de vencimento: {config.data_vencimento}\n{texto_multa}'
                f'\nAceda ao sistema para efetuar o pagamento dentro do prazo.\n\nCDOA Angola',
                None, [d.email],
            )
        geradas += 1

    msg = f'Configuração de {descricao} publicada. {geradas} quotas geradas.'
    messages.success(request, msg)
    return JsonResponse({'status': 'ok', 'mensagem': msg, 'geradas': geradas})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_gerar_retroativo(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    try:
        mes = int(request.POST.get('mes', 0))
        ano = int(request.POST.get('ano', 0))
        data_inicio = request.POST.get('data_inicio', '')
        data_fim = request.POST.get('data_fim', '')
        despachante_id = request.POST.get('despachante_id', '')
        despachante_ids = request.POST.get('despachante_ids', '')
        todos = request.POST.get('todos', '1') == '1'
        force = request.POST.get('force', '0') == '1'
        tipo_id = request.POST.get('tipo_id', '') or None
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Dados inválidos'}, status=400)

    meses_para_gerar = []
    if mes and ano:
        meses_para_gerar.append((ano, mes))
    elif data_inicio and data_fim:
        try:
            di = _dt.datetime.strptime(data_inicio, '%Y-%m-%d').date()
            df = _dt.datetime.strptime(data_fim, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'erro': 'Datas inválidas (use AAAA-MM-DD)'}, status=400)
        if di > df:
            return JsonResponse({
                'erro': 'A data de início não pode ser posterior à data de fim.'
            }, status=400)
        m = di.replace(day=1)
        while m <= df:
            meses_para_gerar.append((m.year, m.month))
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)
    else:
        return JsonResponse({'erro': 'Informe mês/ano ou um intervalo de datas'}, status=400)

    if todos:
        despachantes = Usuario.objects.filter(papel__in=['Despachante Oficial', 'Administrador'], status='Ativo')
    elif despachante_ids:
        try:
            ids = [int(x.strip()) for x in despachante_ids.split(',') if x.strip()]
            despachantes = Usuario.objects.filter(id__in=ids, papel__in=['Despachante Oficial', 'Administrador'], status='Ativo')
            if not despachantes.exists():
                return JsonResponse({'erro': 'Nenhum despachante encontrado para os IDs fornecidos'}, status=404)
        except ValueError:
            return JsonResponse({'erro': 'IDs de despachante inválidos'}, status=400)
    elif despachante_id:
        try:
            despachantes = Usuario.objects.filter(id=int(despachante_id), papel__in=['Despachante Oficial', 'Administrador'], status='Ativo')
            if not despachantes.exists():
                return JsonResponse({'erro': 'Despachante não encontrado ou inativo'}, status=404)
        except ValueError:
            return JsonResponse({'erro': 'ID de despachante inválido'}, status=400)
    else:
        despachantes = Usuario.objects.filter(papel__in=['Despachante Oficial', 'Administrador'], status='Ativo')

    tipo = TipoQuota.objects.filter(id=tipo_id).first() if tipo_id else TipoQuota.objects.filter(slug='mensal').first()
    slug_tipo = tipo.slug.upper() if tipo else 'QUOTA'
    erros = []
    geradas = 0
    hoje = timezone.now().date()
    admin_id = request.session['usuario_id']

    for aa, mm in meses_para_gerar:
        config = QuotaConfig.objects.filter(ano=aa, mes=mm, ativa=True).first()
        if not config:
            erros.append(f'{mm:02d}/{aa}: sem configuração ativa')
            continue
        seq = QuotaGerada.objects.filter(ano=aa, mes=mm, tipo=tipo).count() + 1
        for d in despachantes:
            existente = QuotaGerada.objects.filter(despachante=d, tipo=tipo, ano=aa, mes=mm).first()
            if existente:
                if force:
                    existente.delete()
                else:
                    continue
            descricao = f'{tipo.nome} {mm:02d}/{aa}'
            referencia = f'QUOTA-{slug_tipo}-{mm:02d}-{aa}-{seq:05d}'
            seq += 1
            for _attempt in range(5):
                try:
                    q = QuotaGerada.objects.create(
                        despachante=d, tipo=tipo, ano=aa, mes=mm,
                        descricao=descricao,
                        valor=config.valor, valor_original=config.valor, valor_total=config.valor,
                        data_vencimento=config.data_vencimento,
                        data_envio=hoje,
                        referencia=referencia,
                    )
                    break
                except (IntegrityError, ValidationError):
                    referencia = f'QUOTA-{slug_tipo}-{mm:02d}-{aa}-{seq:05d}'
                    seq += 1
                    continue
            _registrar_historico(
                membro=d, quota=q, pagamento=None,
                acao='QUOTA_GERADA',
                descricao=f'Quota gerada retroativamente para {mm:02d}/{aa}. Referência: {referencia}',
                request=request,
            )
            multa_str = f' Multa de {config.multa_percentual}%/dia após o vencimento.' if config.multa_percentual else ''
            Notificacao.objects.create(
                usuario=d, tipo='quota_gerada',
                titulo=f'{descricao} — Gerada Retroativamente',
                mensagem=f'Foi gerada retroativamente a sua {descricao} no valor de Kz {config.valor}.{multa_str}',
                link='/governanca/quotas/',
            )
            if d.email:
                texto_multa = f"Multa de {config.multa_percentual}%/dia após o vencimento.\n" if config.multa_percentual else ""
                _enviar(
                    'Quota Associativa — Geração Retroativa',
                    f'Olá {d.nome},\n\nA sua {descricao} no valor de Kz {config.valor} foi gerada retroativamente.\n'
                    f'Vencimento: {config.data_vencimento}\n{texto_multa}'
                    f'\nAceda ao sistema para efetuar o pagamento.\n\nCDOA Angola',
                    None, [d.email],
                )
            geradas += 1

    msg = f'{geradas} quotas geradas.'
    if erros:
        msg += ' Erros: ' + '; '.join(erros)
    return JsonResponse({'status': 'ok', 'mensagem': msg, 'geradas': geradas, 'erros': erros})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_marcar_paga(request, fatura_uuid):
    from users.permissoes import usuario_tem_permissao
    papel = request.session['usuario']['papel']
    if papel != 'Administrador' and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    quota = get_object_or_404(QuotaGerada, fatura_uuid=fatura_uuid)
    if papel != 'Administrador':
        uid = request.session.get('banca_usuario_id') or request.session.get('usuario_id')
        if quota.despachante_id != uid:
            return JsonResponse({'erro': 'Sem permissão para alterar quotas de outro despachante'}, status=403)
    if quota.status == 'Paga':
        return JsonResponse({'erro': 'Esta quota já está paga'}, status=400)
    quota.status = 'Paga'
    quota.data_pagamento = timezone.now()
    quota.save(update_fields=['status', 'data_pagamento'])
    _registrar_historico(
        membro=quota.despachante, quota=quota, pagamento=None,
        acao='PAGAMENTO_APROVADO',
        descricao=f'Quota marcada como paga pela administração',
        request=request,
    )
    _atualizar_estado_financeiro(quota.despachante_id, registar_historico=True, request=request)
    Notificacao.objects.create(
        usuario=quota.despachante, tipo='pagamento_confirmado',
        titulo='Quota Marcada como Paga',
        mensagem=f'A sua quota {quota.mes:02d}/{quota.ano} foi marcada como paga pela administração.',
        link='/governanca/quotas/',
    )
    if quota.despachante.email:
        _email('Quota Marcada como Paga',
            f'Olá {quota.despachante.nome},\n\nA sua quota {quota.mes:02d}/{quota.ano} foi marcada como paga pela administração.\n\nCDOA Angola', None, [quota.despachante.email])
    return JsonResponse({'status': 'ok', 'mensagem': f'Quota {quota.mes:02d}/{quota.ano} marcada como paga'})


# ─── Cancelamento de Quota ──────────────────────────────────────────────────

@require_http_methods(['POST'])
@_requer_login
def api_quotas_cancelar(request, fatura_uuid):
    """Admin cancela uma quota. NUNCA apaga o registo."""
    from users.permissoes import usuario_tem_permissao
    papel = request.session['usuario']['papel']
    if papel != 'Administrador' and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    quota = get_object_or_404(QuotaGerada, fatura_uuid=fatura_uuid)
    if papel != 'Administrador':
        uid = request.session.get('banca_usuario_id') or request.session.get('usuario_id')
        if quota.despachante_id != uid:
            return JsonResponse({'erro': 'Sem permissão para cancelar quotas de outro despachante'}, status=403)
    if quota.status == 'Paga':
        return JsonResponse({'erro': 'Não é possível cancelar uma quota já paga'}, status=400)
    if quota.status == 'Cancelada':
        return JsonResponse({'erro': 'Quota já está cancelada'}, status=400)
    motivo = request.POST.get('motivo', '').strip()
    if not motivo:
        return JsonResponse({'erro': 'Motivo do cancelamento é obrigatório'}, status=400)
    quota.status = 'Cancelada'
    quota.observacoes = (quota.observacoes + '\n' + f'Cancelado: {motivo}').strip()
    quota.save(update_fields=['status', 'observacoes'])
    pags = list(PagamentoQuota.objects.filter(quota=quota, status='Pendente Confirmacao'))
    for pag in pags:
        pag.status = 'Rejeitado'
        pag.observacoes = (pag.observacoes + '\n' + f'Cancelado por admin: {motivo}').strip()
    PagamentoQuota.objects.bulk_update(pags, fields=['status', 'observacoes'])
    _registrar_historico(
        membro=quota.despachante, quota=quota, pagamento=None,
        acao='QUOTA_CANCELADA',
        descricao=f'Quota cancelada por admin. Motivo: {motivo}',
        request=request,
    )
    _atualizar_estado_financeiro(quota.despachante_id, registar_historico=True, request=request)
    Notificacao.objects.create(
        usuario=quota.despachante, tipo='quota_cancelada',
        titulo='Quota Cancelada',
        mensagem=f'A sua quota {quota.mes:02d}/{quota.ano} foi cancelada. Motivo: {motivo}',
        link='/governanca/quotas/',
    )
    if quota.despachante.email:
        _email('Quota Cancelada',
            f'Olá {quota.despachante.nome},\n\nA sua quota {quota.mes:02d}/{quota.ano} foi cancelada.\nMotivo: {motivo}\n\nCDOA Angola', None, [quota.despachante.email])
    return JsonResponse({'status': 'ok', 'mensagem': f'Quota {quota.mes:02d}/{quota.ano} cancelada'})


# ─── Isenções ────────────────────────────────────────────────────────────────

@_requer_login
def api_quotas_listar_isencoes(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    isencoes = IsencaoMembro.objects.select_related('despachante', 'tipo_quota', 'aprovado_por').order_by('-created_at')
    ano = request.GET.get('ano')
    if ano:
        isencoes = isencoes.filter(data_inicio__year=ano)
    data = []
    for i in isencoes:
        data.append({
            'id': i.id,
            'despachante_nome': i.despachante.nome,
            'despachante_id': i.despachante_id,
            'tipo_quota': i.tipo_quota.nome if i.tipo_quota else 'Todas',
            'data_inicio': str(i.data_inicio),
            'data_fim': str(i.data_fim) if i.data_fim else 'Indeterminado',
            'motivo': i.motivo,
            'criado_por': i.aprovado_por.nome if i.aprovado_por else '—',
            'created_at': i.created_at.strftime('%d/%m/%Y %H:%M'),
        })
    return JsonResponse({'isencoes': data})


@require_http_methods(['POST'])
@_requer_login
def api_quotas_criar_isencao(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    try:
        despachante_id = int(request.POST.get('despachante_id', 0))
        tipo_quota_id = request.POST.get('tipo_quota_id', '') or None
        data_inicio = request.POST.get('data_inicio', '')
        data_fim = request.POST.get('data_fim', '') or None
        motivo = request.POST.get('motivo', '').strip()
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Dados inválidos'}, status=400)
    if not despachante_id or not data_inicio or not motivo:
        return JsonResponse({'erro': 'Preencha todos os campos obrigatórios'}, status=400)
    try:
        data_inicio = _dt.datetime.strptime(data_inicio, '%Y-%m-%d').date()
        data_fim = _dt.datetime.strptime(data_fim, '%Y-%m-%d').date() if data_fim else None
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Datas inválidas (use AAAA-MM-DD)'}, status=400)
    if data_fim and data_inicio > data_fim:
        return JsonResponse({
            'erro': 'A data de início da isenção não pode ser posterior à data de fim.'
        }, status=400)
    isencao = IsencaoMembro.objects.create(
        despachante_id=despachante_id,
        tipo_quota_id=int(tipo_quota_id) if tipo_quota_id else None,
        data_inicio=data_inicio,
        data_fim=data_fim,
        motivo=motivo,
        aprovado_por_id=request.session['usuario_id'],
    )
    membro = Usuario.objects.filter(id=despachante_id).first()
    if membro:
        _registrar_historico(
            membro=membro, quota=None, pagamento=None,
            acao='ISENCAO_CRIADA',
            descricao=f'Isenção criada: {data_inicio} a {data_fim or "indeterminado"}. Motivo: {motivo}',
            request=request,
        )
    filtro_q = {
        'despachante_id': despachante_id,
        'status__in': ['Pendente', 'Atrasada'],
        'data_vencimento__gte': data_inicio,
    }
    if tipo_quota_id:
        filtro_q['tipo_id'] = int(tipo_quota_id)
    if data_fim:
        filtro_q['data_vencimento__lte'] = data_fim
    quotas_isentar = QuotaGerada.objects.filter(**filtro_q)
    for q in quotas_isentar:
        q.status = 'Isenta'
        q.valor_multa = _Decimal('0.00')
        q.valor_total = q.valor_original or q.valor
        q.save(update_fields=['status', 'valor_multa', 'valor_total'])
        if membro:
            _registrar_historico(
                membro=membro, quota=q, pagamento=None,
                acao='QUOTA_ISENTADA',
                descricao=f'Quota isentada por isenção #{isencao.id}',
                request=request,
            )
    _atualizar_estado_financeiro(despachante_id, registar_historico=True, request=request)
    return JsonResponse({'status': 'ok', 'mensagem': 'Isenção criada com sucesso'})


# ─── Histórico ───────────────────────────────────────────────────────────────

@_requer_login
def api_quotas_historico(request, fatura_uuid):
    quota = get_object_or_404(QuotaGerada, fatura_uuid=fatura_uuid)
    papel = request.session['usuario']['papel']
    usuario_id = request.session.get('banca_usuario_id') or request.session['usuario_id']
    if papel != 'Administrador' and quota.despachante_id != usuario_id:
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    historico = HistoricoQuota.objects.filter(quota=quota).select_related('utilizador').order_by('-created_at')
    data = []
    for h in historico:
        data.append({
            'acao': h.acao,
            'descricao': h.descricao,
            'utilizador': h.utilizador.nome if h.utilizador else '—',
            'ip': h.ip or '—',
            'created_at': h.created_at.strftime('%d/%m/%Y %H:%M'),
        })
    return JsonResponse({'historico': data})


# ─── Verificar Vencimentos (trigger manual) ────────────────────────────────

@require_http_methods(['POST'])
@_requer_login
def api_quotas_verificar_vencimentos(request):
    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_quotas'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    from governanca.management.commands.verificar_vencimentos import Command as VencCmd
    from io import StringIO
    out = StringIO()
    cmd = VencCmd()
    cmd.stdout = out
    cmd.handle()
    output = out.getvalue()
    return JsonResponse({'status': 'ok', 'output': output})


# ═══════════════════════════════════════════════════════════════════════════════
# Submódulo 3: Escuta Activa, Fórum & Transparência
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Páginas HTML ────────────────────────────────────────────────────────────

@_requer_login
def consulta_lista(request):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_consultas') and not usuario_tem_permissao(request, 'gerir_governanca'):
        return redirect('governanca_index')

    status_filtro = request.GET.get('status', '')
    qs = ConsultaPublica.objects.select_related('criado_por').all()
    if status_filtro:
        qs = qs.filter(status=status_filtro)
    STATUS_CHOICES = [
        ('Rascunho', 'Rascunhos'),
        ('Publicada', 'Publicadas'),
        ('EmVotacao', 'Em Votação'),
        ('Encerrada', 'Encerradas'),
        ('Aprovada', 'Aprovadas'),
        ('Rejeitada', 'Rejeitadas'),
    ]
    paginator = Paginator(qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'consulta',
        'page_obj': page_obj,
        'status_atual': status_filtro,
        'status_choices': STATUS_CHOICES,
    }
    return render(request, 'governanca/consulta/lista.html', context)


@_requer_login
def consulta_detalhe(request, pk):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_consultas'):
        return redirect('governanca_index')
    consulta = get_object_or_404(ConsultaPublica.objects.prefetch_related(
        'artigos__comentarios__autor', 'artigos__comentarios__respostas__autor',
        'votacoes__votos',
    ), pk=pk)
    total_comentarios = sum(len(a.comentarios.all()) for a in consulta.artigos.all())
    votacao_ativa = consulta.votacoes.filter(ativa=True).first()
    ja_votou = False
    resultados_votacao = {}
    if votacao_ativa:
        usuario_obj = request.usuario_obj
        ja_votou = votacao_ativa.votos.filter(usuario=usuario_obj).exists()
        from django.db.models import Count
        qs_votos = votacao_ativa.votos.values('voto').annotate(total=Count('id'))
        resultados_votacao = {r['voto']: r['total'] for r in qs_votos}
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'consulta',
        'consulta': consulta,
        'total_comentarios': total_comentarios,
        'votacao_ativa': votacao_ativa,
        'ja_votou': ja_votou,
        'resultados_votacao': resultados_votacao,
    }
    return render(request, 'governanca/consulta/detalhe.html', context)


@_requer_login
def consulta_criar(request):
    from users.permissoes import usuario_tem_permissao
    if not usuario_tem_permissao(request, 'gerir_consultas'):
        messages.error(request, 'Sem permissão para gerir consultas.')
        return redirect('governanca_consultas')
    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        if not titulo:
            messages.error(request, 'O título é obrigatório.')
        else:
            prazo_fim_raw = request.POST.get('prazo_fim', '').strip()
            prazo_fim = None
            if prazo_fim_raw:
                from django.utils.dateparse import parse_datetime
                dt = parse_datetime(prazo_fim_raw)
                if not dt or dt <= timezone.now():
                    messages.error(request, 'O prazo final deve ser posterior ao momento atual.')
                    return redirect('governanca_consultas')
                prazo_fim = dt
            consulta = ConsultaPublica.objects.create(
                titulo=titulo,
                descricao=request.POST.get('descricao', '').strip(),
                prazo_fim=prazo_fim,
                criado_por=request.usuario_obj,
            )
            if request.FILES.get('documento'):
                consulta.documento = request.FILES['documento']
                consulta.save()
            for key, value in request.POST.items():
                if key.startswith('artigo_numero_'):
                    idx = key.split('_')[-1]
                    numero = value
                    titulo_artigo = request.POST.get(f'artigo_titulo_{idx}', '')
                    conteudo = request.POST.get(f'artigo_conteudo_{idx}', '')
                    if numero:
                        ArtigoDocumento.objects.create(
                            consulta=consulta,
                            numero=int(numero),
                            titulo=titulo_artigo,
                            conteudo=conteudo,
                            ordem=int(numero),
                        )
            messages.success(request, 'Consulta criada com sucesso.')
            return redirect('governanca_consulta_detalhe', pk=consulta.id)
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'consulta',
    }
    return render(request, 'governanca/consulta/criar.html', context)


@_requer_login
def consulta_editar(request, pk):
    from users.permissoes import usuario_tem_permissao
    consulta = get_object_or_404(ConsultaPublica, pk=pk)
    if not usuario_tem_permissao(request, 'gerir_consultas'):
        messages.error(request, 'Sem permissão para gerir consultas.')
        return redirect('governanca_consultas')
    if consulta.status != 'Rascunho':
        messages.error(request, 'Apenas consultas em rascunho podem ser editadas.')
        return redirect('governanca_consulta_detalhe', pk=consulta.id)
    if request.method == 'POST':
        prazo_fim = request.POST.get('prazo_fim', '').strip()
        if prazo_fim:
            try:
                from django.utils.dateparse import parse_datetime
                dt = parse_datetime(prazo_fim)
                if not dt or dt <= timezone.now():
                    messages.error(request, 'O prazo final deve ser posterior ao momento atual.')
                    return redirect('governanca_consulta_editar', pk=consulta.id)
                consulta.prazo_fim = dt
            except (ValueError, TypeError):
                messages.error(request, 'Prazo final inválido. Use AAAA-MM-DD HH:MM.')
                return redirect('governanca_consulta_editar', pk=consulta.id)
        else:
            consulta.prazo_fim = None
        consulta.titulo = request.POST.get('titulo', consulta.titulo).strip()
        consulta.descricao = request.POST.get('descricao', '').strip()
        consulta.prazo_fim = request.POST.get('prazo_fim') or None
        if request.FILES.get('documento'):
            consulta.documento = request.FILES['documento']
        consulta.save()
        messages.success(request, 'Consulta atualizada com sucesso.')
        return redirect('governanca_consulta_detalhe', pk=consulta.id)
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'consulta',
        'consulta': consulta,
    }
    return render(request, 'governanca/consulta/editar.html', context)


@_requer_login
def consulta_relatorio(request, pk):
    from users.permissoes import usuario_tem_permissao
    papel = request.session.get('usuario', {}).get('papel', '')
    if papel not in ('Administrador', 'Despachante Oficial') and not usuario_tem_permissao(request, 'gerir_consultas'):
        return redirect('governanca_index')
    consulta = get_object_or_404(ConsultaPublica.objects.prefetch_related(
        'artigos__comentarios__autor', 'votacoes__votos'
    ), pk=pk)
    resultados = {}
    votacao = consulta.votacoes.filter(ativa=False).first()
    if votacao:
        from django.db.models import Count
        qs_votos = votacao.votos.values('voto').annotate(total=Count('id'))
        resultados = {r['voto']: r['total'] for r in qs_votos}
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'active_sub': 'consulta',
        'consulta': consulta,
        'resultados': resultados,
    }
    return render(request, 'governanca/consulta/relatorio.html', context)


# ─── API ─────────────────────────────────────────────────────────────────────

@_requer_login
def api_consulta_publicar(request, pk):
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='Rascunho')
    consulta.status = 'Publicada'
    consulta.publicado_em = timezone.now()
    consulta.save()
    from django.db.models import Q
    destinatarios = Usuario.objects.filter(
        Q(papel='Despachante Oficial') | Q(papel='Administrador'),
        status='Ativo'
    ).values_list('email', flat=True)
    Notificacao.objects.bulk_create([
        Notificacao(usuario=u, tipo='consulta_publicada',
            titulo='Nova Consulta Pública',
            mensagem=f'Foi publicada a consulta "{consulta.titulo}". Participe até {consulta.prazo_fim.strftime("%d/%m/%Y %H:%M") if consulta.prazo_fim else "ao prazo indicado"}.',
            link=f'/governanca/consulta/{consulta.id}/',
        ) for u in Usuario.objects.filter(
            Q(papel='Despachante Oficial') | Q(papel='Administrador'),
            status='Ativo'
        )
    ])
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_consulta_comentar(request, pk):
    consulta = get_object_or_404(ConsultaPublica, pk=pk)
    if consulta.status not in ('Publicada', 'EmVotacao'):
        return JsonResponse({'erro': 'Consulta não está aberta a comentários'}, status=400)
    import json
    data = json.loads(request.body)
    artigo_id = data.get('artigo_id')
    texto = data.get('texto', '').strip()
    if not artigo_id or not texto:
        return JsonResponse({'erro': 'artigo_id e texto são obrigatórios'}, status=400)
    artigo = get_object_or_404(ArtigoDocumento, pk=artigo_id, consulta=consulta)
    comentario = Comentario.objects.create(
        artigo=artigo,
        autor=request.usuario_obj,
        texto=texto,
    )
    if consulta.criado_por_id != request.usuario_obj.id:
        Notificacao.objects.create(
            usuario=consulta.criado_por, tipo='novo_comentario',
            titulo='Novo comentário',
            mensagem=f'{request.usuario_obj.nome} comentou no Artigo {artigo.numero} de "{consulta.titulo}".',
            link=f'/governanca/consulta/{consulta.id}/',
        )
    return JsonResponse({'status': 'ok', 'id': comentario.id})


@_requer_login
def api_consulta_responder(request, pk):
    import json
    data = json.loads(request.body)
    comentario_id = data.get('comentario_id')
    texto = data.get('texto', '').strip()
    if not comentario_id or not texto:
        return JsonResponse({'erro': 'comentario_id e texto são obrigatórios'}, status=400)
    comentario = get_object_or_404(Comentario, pk=comentario_id)
    if comentario.artigo.consulta.pk != pk:
        return JsonResponse({'erro': 'Comentário não pertence a esta consulta'}, status=400)
    resposta = Comentario.objects.create(
        artigo=comentario.artigo,
        autor=request.usuario_obj,
        texto=texto,
        resposta_a=comentario,
    )
    return JsonResponse({'status': 'ok', 'id': resposta.id})


@_requer_login
def api_consulta_abrir_votacao(request, pk):
    papel = request.session['usuario']['papel']
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='Publicada')
    if consulta.prazo_fim and consulta.prazo_fim <= timezone.now():
        return JsonResponse({
            'erro': 'O prazo final desta consulta já expirou. Não é possível abrir votação.'
        }, status=400)
    consulta.status = 'EmVotacao'
    consulta.save()
    VotacaoConsulta.objects.create(consulta=consulta)
    usuarios_ativos = Usuario.objects.filter(status='Ativo')
    Notificacao.objects.bulk_create([
        Notificacao(usuario=u, tipo='votacao_aberta',
            titulo='Votação Aberta',
            mensagem=f'A votação para "{consulta.titulo}" está aberta. Participe!',
            link=f'/governanca/consulta/{consulta.id}/',
        ) for u in usuarios_ativos
    ])
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_consulta_votar(request, pk):
    import json
    data = json.loads(request.body)
    voto = data.get('voto')
    if voto not in ('Favor', 'Contra', 'Abstencao'):
        return JsonResponse({'erro': 'Voto inválido'}, status=400)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='EmVotacao')
    votacao = consulta.votacoes.filter(ativa=True).first()
    if not votacao:
        return JsonResponse({'erro': 'Nenhuma votação ativa'}, status=400)
    if votacao.votos.filter(usuario=request.usuario_obj).exists():
        return JsonResponse({'erro': 'Já votou'}, status=400)
    VotoConsulta.objects.create(votacao=votacao, usuario=request.usuario_obj, voto=voto)
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_consulta_encerrar(request, pk):
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='EmVotacao')
    votacao = consulta.votacoes.filter(ativa=True).first()
    if votacao:
        votacao.ativa = False
        votacao.data_fim = timezone.now()
        votacao.save()
    consulta.status = 'Encerrada'
    consulta.save()
    usuarios_ativos = Usuario.objects.filter(status='Ativo')
    Notificacao.objects.bulk_create([
        Notificacao(usuario=u, tipo='consulta_encerrada',
            titulo='Consulta Encerrada',
            mensagem=f'A consulta "{consulta.titulo}" foi encerrada. O relatório final será publicado em breve.',
            link=f'/governanca/consulta/{consulta.id}/',
        ) for u in usuarios_ativos
    ])
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_consulta_gerar_relatorio(request, pk):
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='Encerrada')
    if hasattr(consulta, 'relatorio'):
        return JsonResponse({'erro': 'Relatório já gerado'}, status=400)
    import hashlib, json
    from django.utils import timezone
    sugestoes = {}
    for artigo in consulta.artigos.prefetch_related('comentarios__autor'):
        sugestoes[str(artigo.numero)] = {
            'titulo': artigo.titulo,
            'comentarios': [{'autor': c.autor.nome, 'texto': c.texto, 'data': c.created_at.isoformat()} for c in artigo.comentarios.all()],
        }
    votacao = consulta.votacoes.filter(ativa=False).first()
    resultados_votacao = {}
    if votacao:
        qs_votos = votacao.votos.values('voto').annotate(total=Count('id'))
        resultados_votacao = {r['voto']: r['total'] for r in qs_votos}
    relatorio_data = {
        'consulta': consulta.titulo,
        'criado_em': timezone.now().isoformat(),
        'sugestoes': sugestoes,
        'resultados_votacao': resultados_votacao,
    }
    relatorio = RelatorioConsulta.objects.create(
        consulta=consulta,
        conteudo=relatorio_data,
        criado_por=request.usuario_obj,
    )
    hash_str = json.dumps(relatorio_data, sort_keys=True)
    relatorio.assinatura_hash = hashlib.sha256(hash_str.encode()).hexdigest()
    relatorio.save()
    usuarios_ativos = Usuario.objects.filter(status='Ativo')
    Notificacao.objects.bulk_create([
        Notificacao(usuario=u, tipo='relatorio_publicado',
            titulo='Relatório Disponível',
            mensagem=f'O relatório da consulta "{consulta.titulo}" já está disponível.',
            link=f'/governanca/consulta/{consulta.id}/relatorio/',
        ) for u in usuarios_ativos
    ])
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_consulta_publicar_versao_final(request, pk):
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='Encerrada')
    consulta.status = 'Aprovada'
    consulta.save()
    usuarios_ativos = Usuario.objects.filter(status='Ativo')
    Notificacao.objects.bulk_create([
        Notificacao(usuario=u, tipo='versao_final_publicada',
            titulo='Versão Final Publicada',
            mensagem=f'A versão final da consulta "{consulta.titulo}" foi publicada no Repositório Digital.',
            link='/governanca/atas/',
        ) for u in usuarios_ativos
    ])
    return JsonResponse({'status': 'ok'})


@_requer_login
def api_consulta_rejeitar(request, pk):
    if not _pode_admin_ou_secretario(request.session['usuario']['papel'], request.usuario_obj, request):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)
    consulta = get_object_or_404(ConsultaPublica, pk=pk, status='Encerrada')
    consulta.status = 'Rejeitada'
    consulta.save()
    return JsonResponse({'status': 'ok'})


# ═══════════════════════════════════════════════════════════════════════════════
# Submódulo 1 — Convocatórias
# ═══════════════════════════════════════════════════════════════════════════════

@_requer_login
def lista_convocatorias(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if assembleia.status != 'Agendada':
        messages.error(request, 'Convocatórias só estão disponíveis para assembleias agendadas.')
        return redirect('governanca_detalhe', pk=pk)
    qs = assembleia.convocatorias.all()
    paginator = Paginator(qs, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    total = qs.count()
    publicadas = qs.filter(status='Publicada').count()
    rascunhos = qs.filter(status='Rascunho').count()
    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleia': assembleia,
        'convocatorias': page_obj,
        'page_obj': page_obj,
        'total_conv': total,
        'publicadas_conv': publicadas,
        'rascunhos_conv': rascunhos,
    }
    return render(request, 'governanca/convocatorias/lista.html', context)


@_requer_login
def criar_convocatoria(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if request.session['usuario']['papel'] != 'Administrador' and not usuario_tem_permissao(request, 'gerir_convocatorias'):
        messages.error(request, 'Sem permissão.')
        return redirect('governanca_detalhe', pk=pk)
    if assembleia.status != 'Agendada':
        messages.error(request, 'Só é possível criar convocatórias para assembleias agendadas.')
        return redirect('governanca_detalhe', pk=pk)

    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        descricao = request.POST.get('descricao', '').strip()
        prazo = request.POST.get('prazo_confirmacao', '')
        documento = request.FILES.get('documento')
        if not titulo:
            messages.error(request, 'Título obrigatório.')
            return render(request, 'governanca/convocatorias/criar.html', locals())
        conv = Convocatoria.objects.create(
            assembleia=assembleia, titulo=titulo, descricao=descricao,
            documento=documento or '',
        )
        if prazo:
            from django.utils.dateparse import parse_datetime
            dt = parse_datetime(prazo)
            if not dt:
                messages.error(request, 'Prazo de confirmação inválido. Use AAAA-MM-DD HH:MM.')
                return render(request, 'governanca/convocatorias/criar.html', locals())
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
            if dt <= timezone.now():
                messages.error(request, 'O prazo de confirmação deve ser posterior ao momento atual.')
                return render(request, 'governanca/convocatorias/criar.html', locals())
            if dt >= assembleia.data_hora:
                messages.error(request, 'O prazo de confirmação deve ser anterior à data da assembleia.')
                return render(request, 'governanca/convocatorias/criar.html', locals())
            conv.prazo_confirmacao = dt
            conv.save()
        messages.success(request, 'Convocatória criada como rascunho.')
        return redirect('governanca_convocatorias', pk=assembleia.pk)

    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleia': assembleia,
    }
    return render(request, 'governanca/convocatorias/criar.html', context)


@require_http_methods(['POST'])
@_requer_login
def api_convocatoria_publicar(request, pk):
    from users.permissoes import usuario_tem_permissao
    conv = get_object_or_404(Convocatoria, pk=pk)
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_convocatorias'):
        return JsonResponse({'status': 'error', 'message': 'Sem permissão.'}, status=403)
    if conv.assembleia.status != 'Agendada':
        return JsonResponse({'status': 'error', 'message': 'A assembleia já não está agendada.'}, status=400)
    if conv.assembleia.data_hora <= timezone.now():
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode publicar convocatória para uma assembleia cuja data já passou.'
        }, status=400)
    conv.status = 'Publicada'
    conv.save()
    _notificar_para_papel('Administrador', 'convocatoria_publicada',
        f'Convocatória: {conv.titulo}',
        f'Foi publicada a convocatória "{conv.titulo}" para {conv.assembleia.titulo}.',
        f'/governanca/assembleia/{conv.assembleia.pk}/')
    _notificar_para_papel('Despachante Oficial', 'convocatoria_publicada',
        f'Convocatória: {conv.titulo}',
        f'Foi publicada a convocatória "{conv.titulo}". Confirme a sua presença.',
        f'/governanca/assembleia/{conv.assembleia.pk}/')
    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
@_requer_login
def api_convocatoria_confirmar_rececao(request, pk):
    conv = get_object_or_404(Convocatoria, pk=pk)
    if conv.status != 'Publicada':
        return JsonResponse({'status': 'error', 'message': 'Convocatória não publicada.'}, status=400)
    if conv.assembleia.data_hora <= timezone.now():
        return JsonResponse({
            'status': 'error',
            'message': 'Não pode confirmar receção de convocatória para uma assembleia cuja data já passou.'
        }, status=400)
    return JsonResponse({'status': 'ok', 'message': 'Receção confirmada.'})


# ═══════════════════════════════════════════════════════════════════════════════
# RSVP — Confirmação de Participação
# ═══════════════════════════════════════════════════════════════════════════════

@require_http_methods(['POST'])
@_requer_login
def api_responder_presenca(request, pk):
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if assembleia.status not in ('Agendada', 'Em Curso'):
        return JsonResponse({'status': 'error', 'message': 'Assembleia já foi concluída ou cancelada.'}, status=400)
    data = json.loads(request.body)
    resposta = data.get('resposta', '')
    if resposta not in ('Sim', 'Nao', 'Talvez'):
        return JsonResponse({'status': 'error', 'message': 'Resposta inválida.'}, status=400)
    rp, created = RespostaPresenca.objects.update_or_create(
        assembleia=assembleia,
        usuario_id=request.session['usuario_id'],
        defaults={'resposta': resposta},
    )
    return JsonResponse({
        'status': 'ok',
        'resposta': rp.resposta,
        'quorum_previsto': assembleia.quorum_previsto,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Reabertura de Votação
# ═══════════════════════════════════════════════════════════════════════════════

@require_http_methods(['POST'])
@_requer_login
def api_reabrir_votacao(request, pk):
    from users.permissoes import usuario_tem_permissao
    pauta = get_object_or_404(PautaVotacao, pk=pk)
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_votacoes'):
        return JsonResponse({'status': 'error', 'message': 'Sem permissão.'}, status=403)
    if pauta.status != 'Concluida':
        return JsonResponse({'status': 'error', 'message': 'Apenas pautas concluídas podem ser reabertas.'}, status=400)
    if pauta.assembleia.status != 'Em Curso':
        return JsonResponse({'status': 'error', 'message': 'Assembleia precisa estar em curso.'}, status=400)

    votos_anteriores = pauta.votos.count()
    with transaction.atomic():
        pauta.votos.all().delete()
        pauta.status = 'Pendente'
        pauta.reaberta = True
        pauta.reaberta_em = timezone.now()
        pauta.resultado_final = ''
        pauta.save()

    _log_assembleia(pauta.assembleia_id, request.session['usuario_id'], 'votacao_reaberta', {
        'pauta_id': pauta.id, 'pauta_titulo': pauta.titulo,
        'votos_anteriores_apagados': votos_anteriores,
    }, ip=_get_client_ip(request))

    _notificar_para_papel('Administrador', 'votacao_reaberta',
        f'Votação reaberta: {pauta.titulo}',
        f'A votação da pauta "{pauta.titulo}" foi reaberta pelo administrador.',
        f'/governanca/assembleia/{pauta.assembleia.pk}/sala/')
    _notificar_para_papel('Despachante Oficial', 'votacao_reaberta',
        f'Votação reaberta: {pauta.titulo}',
        f'A votação da pauta "{pauta.titulo}" foi reaberta. Volte a votar!',
        f'/governanca/assembleia/{pauta.assembleia.pk}/sala/')

    _broadcast_ws(pauta.assembleia_id, 'votacao_reaberta', {
        'pauta_id': pauta.id, 'titulo': pauta.titulo,
        'votos_anteriores': votos_anteriores,
    })

    return JsonResponse({
        'status': 'ok',
        'message': f'Votação reaberta. {votos_anteriores} voto(s) anterior(es) arquivado(s).',
        'votos_anteriores': votos_anteriores,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Exportação de Resultados (PDF / Excel / CSV)
# ═══════════════════════════════════════════════════════════════════════════════

from io import BytesIO


@_requer_login
def exportar_resultados_pdf(request, pk):
    import logging as _log
    _logger = _log.getLogger(__name__)
    try:
        assembleia = get_object_or_404(Assembleia, pk=pk)
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        from reportlab.pdfgen import canvas
        from reportlab.platypus import Table, TableStyle

        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        w, h = A4
        c.setTitle(f'Resultados - {assembleia.titulo}')

        # Header CDOA
        cor_cdoa = HexColor('#1a3a5c')
        cor_cdoa_gold = HexColor('#c9a84c')
        estado_display = assembleia.get_status_display()
        c.saveState()
        c.setFillColor(cor_cdoa)
        c.rect(0, h - 50, w, 50, fill=1, stroke=0)
        c.setFillColor(HexColor('#ffffff'))
        c.setFont('Helvetica-Bold', 12)
        c.drawString(30, h - 35, 'REPÚBLICA DE ANGOLA')
        c.setFont('Helvetica', 9)
        c.drawString(30, h - 20, 'CÂMARA DOS DESPACHANTES OFICIAIS ADUANEIROS (CDOA)')
        c.setFillColor(cor_cdoa_gold)
        c.setFont('Helvetica-Bold', 11)
        c.drawRightString(w - 30, h - 30, estado_display)
        c.restoreState()

        y = h - 80
        # Título do documento
        c.saveState()
        c.setFont('Helvetica-Bold', 18)
        c.setFillColor(cor_cdoa)
        c.drawCentredString(w / 2, y, 'Resultados da Assembleia')
        y -= 25
        c.setFont('Helvetica', 10)
        c.setFillColor(HexColor('#333333'))
        c.drawCentredString(w / 2, y, assembleia.titulo)
        y -= 25
        c.restoreState()

        c.saveState()
        c.setFont('Helvetica-Bold', 11)
        c.drawString(40, y, f'Data: {assembleia.data_hora:%d/%m/%Y %H:%M}')
        y -= 16
        c.drawString(40, y, f'Status: {estado_display}')
        y -= 16
        c.drawString(40, y, f'Presentes: {assembleia.presentes_count} / Quórum: {assembleia.quorum_minimo}')
        y -= 25

        pautas_com_votos = assembleia.pautas.with_vote_counts()
        for pauta in pautas_com_votos:
            c.setFont('Helvetica-Bold', 12)
            c.setFillColor(HexColor('#1a3a5c'))
            c.drawString(40, y, f'{pauta.ordem}. {pauta.titulo}')
            y -= 18
            c.setFont('Helvetica', 10)
            c.setFillColor(HexColor('#333333'))
            c.drawString(60, y, f'Favor: {pauta.votos_favor}  |  Contra: {pauta.votos_contra}  |  Abstenção: {pauta.votos_abstencao}')
            y -= 14
            c.drawString(60, y, f'Total: {pauta.total_votos} votos  |  Resultado: {pauta.resultado_final or "---"}')
            y -= 22
            if y < 80:
                c.showPage()
                # Header CDOA na nova página
                c.saveState()
                c.setFillColor(cor_cdoa)
                c.rect(0, h - 50, w, 50, fill=1, stroke=0)
                c.setFillColor(HexColor('#ffffff'))
                c.setFont('Helvetica-Bold', 12)
                c.drawString(30, h - 35, 'REPÚBLICA DE ANGOLA')
                c.setFont('Helvetica', 9)
                c.drawString(30, h - 20, 'CÂMARA DOS DESPACHANTES OFICIAIS ADUANEIROS (CDOA)')
                c.setFillColor(cor_cdoa_gold)
                c.setFont('Helvetica-Bold', 11)
                c.drawRightString(w - 30, h - 30, estado_display)
                c.restoreState()
                y = h - 80

        c.saveState()
        c.setFont('Helvetica', 8)
        c.setFillColor(HexColor('#888888'))
        c.drawCentredString(w / 2, 30, f'Documento gerado em {timezone.now():%d/%m/%Y %H:%M}  |  Hash: {assembleia.hash_integridade[:20]}...')
        c.restoreState()

        c.showPage()
        c.save()
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="resultados_{assembleia.pk}_{assembleia.titulo[:30]}.pdf"'
        return response
    except Exception:
        _logger.exception("Erro ao gerar PDF de resultados da assembleia %s", pk)
        return HttpResponse('Erro ao gerar PDF.', status=500)


@_requer_login
def exportar_resultados_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    assembleia = get_object_or_404(Assembleia, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resultados'

    header_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    ws.cell(1, 1, 'Pauta').fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 2, 'Favor').fill = header_fill
    ws.cell(1, 2).font = header_font
    ws.cell(1, 3, 'Contra').fill = header_fill
    ws.cell(1, 3).font = header_font
    ws.cell(1, 4, 'Abstenção').fill = header_fill
    ws.cell(1, 4).font = header_font
    ws.cell(1, 5, 'Total').fill = header_fill
    ws.cell(1, 5).font = header_font
    ws.cell(1, 6, 'Resultado').fill = header_fill
    ws.cell(1, 6).font = header_font

    pautas_com_votos = assembleia.pautas.with_vote_counts()

    for i, pauta in enumerate(pautas_com_votos, 2):
        ws.cell(i, 1, pauta.titulo)
        ws.cell(i, 2, pauta.votos_favor)
        ws.cell(i, 3, pauta.votos_contra)
        ws.cell(i, 4, pauta.votos_abstencao)
        ws.cell(i, 5, pauta.total_votos)
        ws.cell(i, 6, pauta.resultado_final or '---')

    # Sheet detalhe
    ws2 = wb.create_sheet('Detalhe Votos')
    ws2.cell(1, 1, 'Pauta').fill = header_fill
    ws2.cell(1, 1).font = header_font
    ws2.cell(1, 2, 'Eleitor').fill = header_fill
    ws2.cell(1, 2).font = header_font
    ws2.cell(1, 3, 'Voto').fill = header_fill
    ws2.cell(1, 3).font = header_font
    ws2.cell(1, 4, 'Delegação').fill = header_fill
    ws2.cell(1, 4).font = header_font
    row = 2
    for pauta in pautas_com_votos:
        for voto in pauta.votos.select_related('usuario', 'delegado_de').all():
            if pauta.tipo_votacao == 'Secreta':
                opcao = '*** (voto secreto)'
            else:
                opcao = voto.opcao
            ws2.cell(row, 1, pauta.titulo)
            ws2.cell(row, 2, voto.usuario.nome if voto.usuario else '---')
            ws2.cell(row, 3, opcao)
            ws2.cell(row, 4, 'Sim' if voto.em_delegacao else 'Não')
            row += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="resultados_{assembleia.pk}_{assembleia.titulo[:30]}.xlsx"'
    wb.save(response)
    return response


@_requer_login
def exportar_resultados_csv(request, pk):
    import csv
    assembleia = get_object_or_404(Assembleia, pk=pk)
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="resultados_{assembleia.pk}_{assembleia.titulo[:30]}.csv"'
    w = csv.writer(response)
    w.writerow(['Pauta', 'Favor', 'Contra', 'Abstenção', 'Total', 'Resultado'])
    for pauta in assembleia.pautas.with_vote_counts():
        w.writerow([pauta.titulo, pauta.votos_favor, pauta.votos_contra, pauta.votos_abstencao, pauta.total_votos, pauta.resultado_final or '---'])
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# Assinatura Digital da Ata
# ═══════════════════════════════════════════════════════════════════════════════

@require_http_methods(['POST'])
@_requer_login
def api_assinar_ata(request, pk):
    ata = get_object_or_404(AtaDigital, pk=pk)
    assembleia = ata.assembleia
    usuario = request.usuario_obj
    papel = request.session['usuario']['papel']

    # Verificar se usuário é membro da mesa
    membro = MembroMesa.objects.filter(assembleia=assembleia, usuario=usuario).first()
    if not membro and papel != 'Administrador':
        return JsonResponse({'status': 'error', 'message': 'Apenas membros da mesa podem assinar a ata.'}, status=403)

    funcao = membro.funcao if membro else 'Administrador'
    is_presidente = funcao == 'Presidente'
    is_secretario = funcao in ('Secretário', '1º Secretário')

    if not is_presidente and not is_secretario and papel != 'Administrador':
        return JsonResponse({'status': 'error', 'message': 'Apenas Presidente ou Secretário podem assinar.'}, status=403)

    # Gerar hash da assinatura
    raw = f'{ata.id}-{usuario.id}-{timezone.now().isoformat()}'
    assinatura = hashlib.sha256(raw.encode()).hexdigest()

    from users.permissoes import _is_admin_ou_acesso_total
    if is_presidente or _is_admin_ou_acesso_total(request):
        if ata.assinatura_hash_presidente:
            return JsonResponse({'status': 'error', 'message': 'Presidente já assinou.'}, status=400)
        ata.assinatura_hash_presidente = assinatura
        ata.assinado_presidente_em = timezone.now()
        if not ata.assinatura_hash:
            ata.assinatura_hash = assinatura
            ata.assinado_por = usuario
            ata.assinado_em = timezone.now()

    if is_secretario:
        if ata.assinatura_hash_secretario:
            return JsonResponse({'status': 'error', 'message': 'Secretário já assinou.'}, status=400)
        ata.assinatura_hash_secretario = assinatura
        ata.assinado_secretario_em = timezone.now()

    # Atualizar status
    if ata.assinatura_hash_presidente and ata.assinatura_hash_secretario:
        ata.status_assinatura = 'Assinada'
    elif ata.assinatura_hash_presidente:
        ata.status_assinatura = 'Aguardando Secretario'
    elif ata.assinatura_hash_secretario:
        ata.status_assinatura = 'Aguardando Presidente'

    ata.save()

    _notificar_para_papel('Administrador', 'ata_assinada',
        f'Ata assinada: {assembleia.titulo}',
        f'{usuario.nome} assinou a ata da assembleia "{assembleia.titulo}". Status: {ata.get_status_assinatura_display()}.',
        f'/governanca/assembleia/{assembleia.pk}/')

    return JsonResponse({
        'status': 'ok',
        'assinatura': assinatura,
        'status_assinatura': ata.get_status_assinatura_display(),
        'funcao': funcao,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Logs da Assembleia
# ═══════════════════════════════════════════════════════════════════════════════

@_requer_login
def assembleia_logs(request, pk):
    from users.permissoes import usuario_tem_permissao
    assembleia = get_object_or_404(Assembleia, pk=pk)
    if request.session['usuario']['papel'] not in ('Administrador',) and not usuario_tem_permissao(request, 'gerir_assembleia'):
        messages.error(request, 'Sem permissão.')
        return redirect('governanca_detalhe', pk=pk)

    qs = assembleia.logs.select_related('usuario').all()
    paginator = Paginator(qs, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'usuario': request.session['usuario'],
        'nome': request.session['usuario']['nome'],
        'papel': request.session['usuario']['papel'],
        'active_menu': 'Governanca',
        'assembleia': assembleia,
        'logs': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'governanca/assembleia_logs.html', context)


# ═══════════════════════════════════════════════════════════════════════════════
# Notificações
# ═══════════════════════════════════════════════════════════════════════════════

def pagina_notificacoes(request):
    usuario = _get_usuario(request)
    if not usuario:
        return redirect('login')
    usuario_id = request.session['usuario_id']
    qs = Notificacao.objects.filter(usuario_id=usuario_id).order_by('-created_at')
    nao_lidas = qs.filter(lida=False).count()
    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'usuario': usuario,
        'nome': usuario['nome'],
        'papel': usuario['papel'],
        'active_menu': 'Governanca',
        'notificacoes': page_obj,
        'page_obj': page_obj,
        'nao_lidas': nao_lidas,
    }
    return render(request, 'governanca/notificacoes.html', context)


def api_notificacoes(request):
    usuario = _get_usuario(request)
    if not usuario:
        return JsonResponse({'error': 'Login required'}, status=401)
    qs = Notificacao.objects.filter(usuario_id=request.session['usuario_id'])
    nao_lidas = qs.filter(lida=False).count()
    data = [{
        'id': n.id, 'tipo': n.tipo, 'titulo': n.titulo,
        'mensagem': n.mensagem, 'link': n.link,
        'lida': n.lida, 'created_at': n.created_at.isoformat(),
    } for n in qs.order_by('-created_at')[:50]]
    return JsonResponse({'notificacoes': data, 'nao_lidas': nao_lidas})


@require_http_methods(['POST'])
def api_notificacao_marcar_lida(request, pk):
    usuario = _get_usuario(request)
    if not usuario:
        return JsonResponse({'error': 'Login required'}, status=401)
    n = get_object_or_404(Notificacao, pk=pk, usuario_id=request.session['usuario_id'])
    n.lida = True
    n.save(update_fields=['lida'])
    return JsonResponse({'status': 'ok'})


@require_http_methods(['POST'])
def api_notificacoes_marcar_todas_lidas(request):
    usuario = _get_usuario(request)
    if not usuario:
        return JsonResponse({'error': 'Login required'}, status=401)
    Notificacao.objects.filter(usuario_id=request.session['usuario_id'], lida=False).update(lida=True)
    return JsonResponse({'status': 'ok'})





# ═══════════════════════════════════════════════════════════════════════════════
# GESTÃO DE UTILIZADORES (hub central)
# ═══════════════════════════════════════════════════════════════════════════════

@_requer_login
def utilizador_novo_view(request):
    from users.permissoes import usuario_tem_permissao
    usuario = _get_usuario(request)
    if not usuario:
        return redirect('login')
    if usuario.get('papel') != 'Administrador' and not usuario_tem_permissao(request, 'gerir_utilizadores'):
        messages.error(request, 'Apenas administradores podem criar utilizadores.')
        return redirect('dashboard')

    erros = {}
    if request.method == 'POST':
        from users.models import Usuario, ColaboradorInstitucional
        from utils.email_utils import gerar_senha_aleatoria
        import bcrypt

        tipo = request.POST.get('tipo', '')
        nome = request.POST.get('nome', '').strip()
        email = request.POST.get('email', '').strip().lower()
        telefone = request.POST.get('telefone', '').strip()
        nif = request.POST.get('nif', '').strip()
        cedula = request.POST.get('cedula', '').strip()
        area_actuacao = request.POST.get('area_actuacao', '').strip()
        nome_tipo = request.POST.get('nome_tipo', '').strip()
        salario_base = request.POST.get('salario_base', '').strip()
        funcao_id = request.POST.get('funcao', '').strip()

        if not nome: erros['nome'] = 'O nome é obrigatório.'
        if not email: erros['email'] = 'O email é obrigatório.'
        elif Usuario.objects.filter(email=email).exists(): erros['email'] = 'Já existe um utilizador com este email.'
        if tipo == 'despachante':
            if not cedula: erros['cedula'] = 'A cédula CDOA é obrigatória.'
            if not nif: erros['nif'] = 'O NIF é obrigatório.'

        if not erros:
            if tipo == 'despachante': papel = 'Despachante Oficial'
            elif tipo == 'colaborador': papel = 'Colaborador Institucional'
            elif tipo == 'outro': papel = 'Visualizador'
            else: papel = 'Visualizador'

            base = email.split('@')[0]
            username = base
            c = 1
            while Usuario.objects.filter(username=username).exists():
                username = f'{base}{c}'; c += 1

            senha = gerar_senha_aleatoria(10)
            salt = bcrypt.gensalt()
            hash_senha = bcrypt.hashpw(senha.encode('utf-8'), salt).decode('utf-8').replace('$2b$', '$2y$')

            funcao_obj = None
            if funcao_id and funcao_id.isdigit():
                from users.models import Funcao
                funcao_obj = Funcao.objects.filter(pk=int(funcao_id)).first()

            user = Usuario.objects.create(
                username=username, password=hash_senha, nome=nome, email=email,
                telefone=telefone, nif=nif, cedula=cedula if tipo == 'despachante' else '',
                papel=papel, status='Ativo',
                area_actuacao=area_actuacao if tipo == 'colaborador' else '',
                cargo_personalizado=nome_tipo if tipo == 'outro' else '',
                funcao=funcao_obj if funcao_obj else None,
            )

            if tipo == 'colaborador':
                from decimal import Decimal
                salario_dec = Decimal(parse_kz(salario_base)) if salario_base else None
                ColaboradorInstitucional.objects.create(
                    usuario=user, nome=nome, email=email, telefone=telefone,
                    area_actuacao=area_actuacao, salario_base=salario_dec,
                )

            sucesso_email, msg_email = _enviar_credenciais_utilizador(user, senha)
            if sucesso_email:
                messages.success(request, f'Utilizador "{nome}" criado com sucesso. Credenciais enviadas para {email}.')
            else:
                messages.success(request, f'Utilizador "{nome}" criado com sucesso.')
                messages.warning(request, f'Não foi possível enviar o email de credenciais: {msg_email}')
            return redirect('governanca_gerir_utilizadores')

    from users.models import Funcao
    from users.permissoes import get_usuario_permissoes as _get_perms
    ctx = {
        'usuario': usuario, 'nome': usuario['nome'], 'papel': usuario['papel'],
        'active_menu': 'ADMIN_RH', 'active_sub': 'gerir_utilizadores', 'is_admin_sistema': True,
        'user_permissoes': _get_perms(request),
        'erros': erros, 'funcoes': Funcao.objects.all().order_by('nome'),
    }
    return render(request, 'governanca/utilizador_novo.html', ctx)


@_requer_login
def utilizador_editar_view(request, usuario_id):
    from users.permissoes import usuario_tem_permissao
    usuario = _get_usuario(request)
    if not usuario:
        return redirect('login')
    if usuario.get('papel') != 'Administrador' and not usuario_tem_permissao(request, 'gerir_utilizadores'):
        messages.error(request, 'Apenas administradores podem editar utilizadores.')
        return redirect('dashboard')

    from users.models import Usuario, ColaboradorInstitucional
    user_obj = get_object_or_404(Usuario, pk=usuario_id)

    erros = {}
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        email = request.POST.get('email', '').strip().lower()
        telefone = request.POST.get('telefone', '').strip()
        nif = request.POST.get('nif', '').strip()
        cedula = request.POST.get('cedula', '').strip()
        area_actuacao = request.POST.get('area_actuacao', '').strip()
        cargo_personalizado = request.POST.get('cargo_personalizado', '').strip()
        funcao_id = request.POST.get('funcao', '').strip()

        if not nome:
            erros['nome'] = 'O nome é obrigatório.'
        if not email:
            erros['email'] = 'O email é obrigatório.'
        elif email != user_obj.email and Usuario.objects.filter(email=email).exists():
            erros['email'] = 'Já existe um utilizador com este email.'

        if not erros:
            user_obj.nome = nome
            user_obj.email = email
            user_obj.telefone = telefone
            user_obj.nif = nif
            if user_obj.papel == 'Despachante Oficial':
                user_obj.cedula = cedula
            if user_obj.papel == 'Colaborador Institucional':
                user_obj.area_actuacao = area_actuacao
                # Atribuir função
                if funcao_id and funcao_id.isdigit():
                    from users.models import Funcao
                    funcao_obj = Funcao.objects.filter(pk=int(funcao_id)).first()
                    user_obj.funcao = funcao_obj
                else:
                    user_obj.funcao = None
            if user_obj.papel == 'Visualizador':
                user_obj.cargo_personalizado = cargo_personalizado
            user_obj.save()

            messages.success(request, f'Dados de "{user_obj.nome}" actualizados com sucesso.')
            return redirect('governanca_gerir_utilizadores')

    from users.models import Funcao
    from users.permissoes import get_usuario_permissoes as _get_perms
    ctx = {
        'usuario': usuario, 'nome': usuario['nome'], 'papel': usuario['papel'],
        'active_menu': 'ADMIN_RH', 'active_sub': 'gerir_utilizadores', 'is_admin_sistema': True,
        'user_permissoes': _get_perms(request),
        'user_obj': user_obj, 'erros': erros, 'funcoes': Funcao.objects.all().order_by('nome'),
    }
    return render(request, 'governanca/utilizador_editar.html', ctx)


@_requer_login
def utilizador_permissoes_view(request, usuario_id):
    messages.error(request, 'Permissões diretas desativadas. Atribua uma função ao utilizador para definir as permissões.')
    return redirect('governanca_gerir_utilizadores')


@_requer_login
def gerir_utilizadores(request):
    from users.permissoes import usuario_tem_permissao, get_usuario_permissoes
    usuario = _get_usuario(request)
    if not usuario:
        return redirect('login')
    if usuario.get('papel') != 'Administrador' and not usuario_tem_permissao(request, 'gerir_utilizadores'):
        messages.error(request, 'Apenas administradores podem gerir utilizadores.')
        return redirect('dashboard')

    from users.models import Usuario
    q = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    if not tipo:
        tipo = 'Colaborador Institucional'

    utilizadores_qs = Usuario.objects.exclude(papel='Administrador')
    if q:
        utilizadores_qs = utilizadores_qs.filter(nome__icontains=q)
    if tipo:
        utilizadores_qs = utilizadores_qs.filter(papel=tipo)
    utilizadores_qs = utilizadores_qs.order_by('nome').prefetch_related('permissoes_diretas').select_related('funcao')

    paginator = Paginator(utilizadores_qs, 12)
    page_number = request.GET.get('page')
    utilizadores = paginator.get_page(page_number)

    ids = [u.id for u in utilizadores_qs]
    from django.db.models import Count
    from rh.models import Banca
    bancas = Banca.objects.filter(usuario_id__in=ids).annotate(
        num_colaboradores=Count('colaboradores')
    )
    bancas_por_usuario = {}
    for b in bancas:
        bancas_por_usuario.setdefault(b.usuario_id, []).append(b)
    for u in utilizadores:
        u.bancas_list = bancas_por_usuario.get(u.id, [])
        u.bancas_total = len(u.bancas_list)
        u.bancas_colaboradores = sum(
            getattr(b, 'num_colaboradores', 0) or 0 for b in u.bancas_list
        )

    stats_total = utilizadores_qs.count()
    stats_ativos = utilizadores_qs.filter(status='Ativo').count()
    stats_inativos = utilizadores_qs.filter(status__in=['Inativo', 'Suspenso']).count()

    from users.models import Funcao
    extra_params = ''
    if q: extra_params = 'q=' + q
    if tipo: extra_params += ('&' if extra_params else '') + 'tipo=' + tipo

    return render(request, 'governanca/gerir_utilizadores.html', {
        'usuario': usuario,
        'nome': usuario['nome'],
        'papel': usuario['papel'],
        'active_menu': 'ADMIN_RH',
        'active_sub': 'gerir_utilizadores',
        'is_admin_sistema': True,
        'user_permissoes': get_usuario_permissoes(request),
        'utilizadores': utilizadores,
        'bancas_por_usuario': bancas_por_usuario,
        'stats_total': stats_total,
        'stats_ativos': stats_ativos,
        'stats_inativos': stats_inativos,
        'page_obj': utilizadores,
        'q': q,
        'tipo': tipo,
        'extra_params': extra_params,
        'funcoes': Funcao.objects.all().order_by('nome'),
    })


@require_http_methods(['POST'])
@_requer_login
def api_utilizador_criar(request):
    from users.models import Usuario
    from utils.email_utils import gerar_senha_aleatoria
    import bcrypt, json
    from django.utils.text import slugify

    from users.permissoes import usuario_tem_permissao
    if request.session['usuario']['papel'] != 'Administrador' and not usuario_tem_permissao(request, 'gerir_utilizadores'):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)

    data = json.loads(request.body)
    tipo_criacao = data.get('tipo', '')
    nome = data.get('nome', '').strip()
    email = data.get('email', '').strip().lower()
    telefone = data.get('telefone', '').strip()
    nif = data.get('nif', '').strip()
    cedula = data.get('cedula', '').strip()
    area_actuacao = data.get('area_actuacao', '').strip()
    nome_tipo = data.get('nome_tipo', '').strip()
    funcao_id = data.get('funcao_id', '')
    enviar_credenciais = data.get('enviar_credenciais', True)

    if not nome or not email:
        return JsonResponse({'erro': 'Nome e email são obrigatórios.'}, status=400)
    if Usuario.objects.filter(email=email).exists():
        return JsonResponse({'erro': 'Já existe um utilizador com este email.'}, status=400)

    # Definir papel baseado no tipo
    if tipo_criacao == 'despachante':
        papel = 'Despachante Oficial'
        if not cedula:
            return JsonResponse({'erro': 'Cédula CDOA é obrigatória para despachantes.'}, status=400)
        if not nif:
            return JsonResponse({'erro': 'NIF é obrigatório para despachantes.'}, status=400)
    elif tipo_criacao == 'colaborador':
        papel = 'Colaborador Institucional'
    elif tipo_criacao == 'outro':
        papel = 'Visualizador'
    else:
        return JsonResponse({'erro': 'Tipo de utilizador inválido.'}, status=400)

    base_username = email.split('@')[0]
    username = base_username
    contador = 1
    while Usuario.objects.filter(username=username).exists():
        username = f'{base_username}{contador}'
        contador += 1

    senha = gerar_senha_aleatoria(10)
    salt = bcrypt.gensalt()
    hash_senha = bcrypt.hashpw(senha.encode('utf-8'), salt).decode('utf-8').replace('$2b$', '$2y$')

    funcao_obj = None
    if funcao_id and str(funcao_id).isdigit():
        from users.models import Funcao
        funcao_obj = Funcao.objects.filter(pk=int(funcao_id)).first()

    usuario = Usuario.objects.create(
        username=username,
        password=hash_senha,
        nome=nome,
        email=email,
        telefone=telefone,
        nif=nif if nif else '',
        cedula=cedula if cedula else '',
        papel=papel,
        status='Ativo',
        area_actuacao=area_actuacao if tipo_criacao == 'colaborador' else '',
        cargo_personalizado=nome_tipo if tipo_criacao == 'outro' else '',
        funcao=funcao_obj if funcao_obj else None,
    )

    # Criar ColaboradorInstitucional se for colaborador institucional
    if tipo_criacao == 'colaborador':
        from users.models import ColaboradorInstitucional
        ColaboradorInstitucional.objects.create(
            usuario=usuario,
            nome=nome,
            email=email,
            telefone=telefone,
            area_actuacao=area_actuacao,
        )

    # Enviar credenciais
    msg_email = ''
    if enviar_credenciais:
        sucesso, msg_email = _enviar_credenciais_utilizador(usuario, senha)

    return JsonResponse({
        'status': 'ok',
        'usuario': {
            'id': usuario.id,
            'nome': usuario.nome,
            'email': usuario.email,
            'papel': usuario.papel,
            'cargo_personalizado': usuario.cargo_personalizado,
        },
        'email_enviado': enviar_credenciais,
        'msg_email': msg_email,
    })


def _enviar_credenciais_utilizador(usuario, senha):
    if not usuario.email:
        return False, "Utilizador não tem email registado"
    from django.conf import settings
    from django.urls import reverse
    from utils.email_utils import _enviar

    base = settings.SITE_URL.rstrip('/')
    link_login = f"{base}{reverse('login')}"

    assunto = 'As suas credenciais de acesso — SICDOA'
    texto = f"""Prezado(a) {usuario.nome},

A sua conta no Sistema SICDOA foi criada pelo Administrador.

Credenciais de acesso:
  Email : {usuario.email}
  Senha : {senha}

Inicie sessão em: {link_login}

Por segurança, altere a sua senha após o primeiro acesso.

Atenciosamente,
Administração SICDOA — CDOA Angola
"""

    html = f"""
<!DOCTYPE html>
<html lang="pt">
<body style="margin:0;padding:0;background:#f6f7f8;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:40px 20px;">
    <table width="560" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.08);">
      <tr><td style="background:linear-gradient(135deg,#137fec,#0ea5e9);padding:32px 40px;">
        <h1 style="margin:0;color:#fff;font-size:22px;font-weight:700;">CDOA Sistema</h1>
        <p style="margin:6px 0 0;color:rgba(255,255,255,.85);font-size:14px;">Credenciais de Acesso ao SICDOA</p>
      </td></tr>
      <tr><td style="padding:36px 40px;">
        <p style="margin:0 0 16px;color:#374151;font-size:15px;">Prezado(a) <strong>{usuario.nome}</strong>,</p>
        <p style="margin:0 0 24px;color:#6b7280;font-size:14px;line-height:1.6;">A sua conta no Sistema SICDOA foi criada. Utilize as credenciais abaixo para aceder à plataforma.</p>
        <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;margin-bottom:24px;">
          <tr><td style="padding:20px 24px;">
            <p style="margin:0 0 12px;font-size:13px;color:#0369a1;font-weight:600;text-transform:uppercase;letter-spacing:.05em;">As suas credenciais</p>
            <p style="margin:0 0 10px;font-size:14px;color:#374151;"><strong>Email:</strong>&nbsp;{usuario.email}</p>
            <p style="margin:0;font-size:14px;color:#374151;"><strong>Senha:</strong>&nbsp;<code style="background:#e0f2fe;padding:3px 10px;border-radius:5px;font-size:15px;letter-spacing:.08em;">{senha}</code></p>
          </td></tr>
        </table>
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:24px;">
          <tr><td align="center">
            <a href="{link_login}" style="display:inline-block;background:#137fec;color:#fff;text-decoration:none;font-size:15px;font-weight:600;padding:14px 36px;border-radius:10px;">Iniciar sessão no SICDOA</a>
          </td></tr>
        </table>
        <p style="margin:0;color:#ef4444;font-size:13px;font-weight:600;">Por segurança, altere a sua senha após o primeiro acesso.</p>
      </td></tr>
      <tr><td style="background:#f9fafb;padding:20px 40px;border-top:1px solid #e5e7eb;">
        <p style="margin:0;color:#9ca3af;font-size:12px;">© 2026 CDOA Sistema · Câmara dos Despachantes Oficiais de Angola</p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""
    return _enviar(assunto, texto, html, usuario.email)


@require_http_methods(['POST'])
@_requer_login
def api_utilizador_toggle_status(request):
    from users.models import Usuario
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'erro': 'Apenas administradores podem alterar estado de utilizadores'}, status=403)
    data = json.loads(request.body)
    usuario_id = data.get('usuario_id')
    if not usuario_id:
        return JsonResponse({'erro': 'ID do utilizador obrigatório.'}, status=400)
    usuario_obj = get_object_or_404(Usuario, pk=usuario_id)
    if usuario_obj.status == 'Ativo':
        usuario_obj.status = 'Suspenso'
    else:
        usuario_obj.status = 'Ativo'
    usuario_obj.save(update_fields=['status'])
    return JsonResponse({'status': 'ok', 'novo_status': usuario_obj.status})


@require_http_methods(['POST'])
@_requer_login
def api_utilizador_enviar_credenciais(request):
    from users.models import Usuario
    from utils.email_utils import gerar_senha_aleatoria
    import bcrypt
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'erro': 'Apenas administradores podem reenviar credenciais'}, status=403)
    data = json.loads(request.body)
    usuario_id = data.get('usuario_id')
    usuario_obj = get_object_or_404(Usuario, pk=usuario_id)
    senha = gerar_senha_aleatoria(10)
    salt = bcrypt.gensalt()
    hash_senha = bcrypt.hashpw(senha.encode('utf-8'), salt).decode('utf-8').replace('$2b$', '$2y$')
    from django.db import connection
    from django.utils import timezone
    with connection.cursor() as cursor:
        cursor.execute(
            'UPDATE usuarios SET password = %s, updated_at = %s WHERE id = %s',
            [hash_senha, timezone.now(), usuario_obj.id],
        )
    sucesso, msg = _enviar_credenciais_utilizador(usuario_obj, senha)
    return JsonResponse({'status': 'ok', 'email_enviado': sucesso, 'msg': msg})


@require_http_methods(['POST'])
@_requer_login
def api_utilizador_permissoes(request):
    return JsonResponse({
        'status': 'erro',
        'message': 'Permissões diretas desativadas. Atribua uma função ao utilizador para definir as permissões.'
    }, status=403)





def api_permissoes_usuario(request):
    """GET: desativado — permissões vêm exclusivamente da função."""
    return JsonResponse({
        'status': 'erro',
        'message': 'Permissões diretas desativadas. Atribua uma função ao utilizador para definir as permissões.'
    }, status=403)


@require_http_methods(['POST'])
@_requer_login
def api_utilizador_eliminar(request):
    """Elimina um utilizador."""
    from users.models import Usuario
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'erro': 'Apenas administradores podem eliminar utilizadores'}, status=403)
    data = json.loads(request.body)
    usuario_id = data.get('usuario_id')
    if not usuario_id:
        return JsonResponse({'erro': 'ID do utilizador obrigatório.'}, status=400)
    usuario_obj = get_object_or_404(Usuario, pk=usuario_id)
    nome = usuario_obj.nome
    usuario_obj.delete()
    return JsonResponse({'status': 'ok', 'message': f'Utilizador "{nome}" eliminado com sucesso.'})


@require_http_methods(['POST'])
@_requer_login
def api_utilizador_atribuir_funcao(request):
    """Atribui ou remove a função de um utilizador."""
    from users.models import Funcao, Usuario
    if request.session['usuario']['papel'] != 'Administrador':
        return JsonResponse({'erro': 'Apenas administradores podem atribuir funções'}, status=403)
    data = json.loads(request.body)
    usuario_id = data.get('usuario_id')
    funcao_id = data.get('funcao_id')
    if not usuario_id:
        return JsonResponse({'erro': 'ID do utilizador obrigatório.'}, status=400)
    usuario_obj = get_object_or_404(Usuario, pk=usuario_id)
    if funcao_id:
        funcao_obj = get_object_or_404(Funcao, pk=funcao_id)
        usuario_obj.funcao = funcao_obj
        msg = f'Função "{funcao_obj.nome}" atribuída a {usuario_obj.nome}.'
    else:
        usuario_obj.funcao = None
        msg = f'Função removida de {usuario_obj.nome}.'
    usuario_obj.save(update_fields=['funcao'])
    return JsonResponse({'status': 'ok', 'message': msg, 'funcao_nome': funcao_obj.nome if funcao_id else None})

